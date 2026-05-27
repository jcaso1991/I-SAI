from fastapi import FastAPI, APIRouter, HTTPException, Depends, Header, Query, UploadFile, File, Request
from fastapi.responses import RedirectResponse, HTMLResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import re
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Literal
import uuid
import secrets
from datetime import datetime, timezone, timedelta
import time
import bcrypt
import jwt as pyjwt
import msal
import httpx
import io
import math
import base64
from openpyxl import load_workbook
import pypdfium2 as pdfium
from PIL import Image
from fastapi.responses import Response
from pdf_filler import build_budget_pdf
from cryptography.fernet import Fernet

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# ---------------- Config ----------------
MONGO_URL = os.environ['MONGO_URL']
DB_NAME = os.environ['DB_NAME']
JWT_SECRET = os.environ.get('JWT_SECRET', '')
if not JWT_SECRET:
    import secrets
    JWT_SECRET = secrets.token_hex(32)
    logging.warning("JWT_SECRET not configured. Using auto-generated key. Set it in .env for persistence.")
JWT_ALGORITHM = os.environ.get('JWT_ALGORITHM', 'HS256')
if JWT_ALGORITHM not in ('HS256', 'HS384', 'HS512'):
    raise ValueError("JWT_ALGORITHM must be HS256, HS384, or HS512")
JWT_EXPIRE_HOURS = int(os.environ.get('JWT_EXPIRE_HOURS', '24'))

MS_TENANT_ID = os.environ.get('MS_TENANT_ID', '')
MS_CLIENT_ID = os.environ.get('MS_CLIENT_ID', '')
MS_CLIENT_SECRET = os.environ.get('MS_CLIENT_SECRET', '')
MS_REDIRECT_URI = os.environ.get('MS_REDIRECT_URI', '')
ONEDRIVE_FILE_PATH = os.environ.get('ONEDRIVE_FILE_PATH', '/Materiales.xlsx')
ONEDRIVE_SHARE_URL = os.environ.get('ONEDRIVE_SHARE_URL', '').strip()
INITIAL_EXCEL_PATH = os.environ.get('INITIAL_EXCEL_PATH', '/app/backend/Materiales.xlsx')

# Support personal + work accounts
MS_AUTHORITY = "https://login.microsoftonline.com/common"
MS_SCOPES = ["Files.ReadWrite.All", "User.Read"]  # offline_access added automatically by MSAL
GRAPH_BASE = "https://graph.microsoft.com/v1.0"

ONEDRIVE_TOKEN_ENCRYPTION_KEY = os.environ.get('ONEDRIVE_TOKEN_ENCRYPTION_KEY', '')

def _onedrive_fernet() -> Fernet:
    if not ONEDRIVE_TOKEN_ENCRYPTION_KEY:
        raise HTTPException(500, "ONEDRIVE_TOKEN_ENCRYPTION_KEY no configurada. Contactá al administrador del sistema.")
    try:
        return Fernet(ONEDRIVE_TOKEN_ENCRYPTION_KEY.encode())
    except Exception:
        raise HTTPException(500, "ONEDRIVE_TOKEN_ENCRYPTION_KEY inválida. Generá una clave Fernet válida.")

def _encrypt_onedrive_token(token: str) -> str:
    return _onedrive_fernet().encrypt(token.encode()).decode()

def _decrypt_onedrive_token(token_enc: str) -> str:
    try:
        return _onedrive_fernet().decrypt(token_enc.encode()).decode()
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(500, "No se pudo descifrar el token de OneDrive. Revisá la clave de cifrado o reconectá OneDrive.")

# Microsoft user-authentication redirect URIs
MS_AUTH_REDIRECT_URI = os.environ.get('MS_AUTH_REDIRECT_URI', 'http://localhost:8000/api/auth/microsoft/callback')
FRONTEND_URL = os.environ.get('FRONTEND_URL', 'http://localhost:8081')
CORS_ORIGINS = os.environ.get('CORS_ORIGINS', '')
TRUST_PROXY_HEADERS = os.environ.get('TRUST_PROXY_HEADERS', 'false').lower() == 'true'

def _microsoft_login_enabled() -> bool:
    return bool(MS_CLIENT_ID and MS_CLIENT_SECRET and MS_AUTH_REDIRECT_URI)

# Demo seed config
ENABLE_DEMO_SEED = os.environ.get('ENABLE_DEMO_SEED', 'false').lower() == 'true'
DEMO_ADMIN_EMAIL = os.environ.get('DEMO_ADMIN_EMAIL', 'admin@materiales.com')
DEMO_ADMIN_PASSWORD = os.environ.get('DEMO_ADMIN_PASSWORD', '')
DEMO_ADMIN_NAME = os.environ.get('DEMO_ADMIN_NAME', 'Administrador')
DEMO_USER_PASSWORD = os.environ.get('DEMO_USER_PASSWORD', '')

# Auto-sync config
AUTO_IMPORT_INTERVAL_SEC = 300  # re-import OneDrive file every 5 min on read
AUTO_PUSH_DELAY_SEC = 6         # debounce pushes: wait 6s after last edit

# ---------------- DB ----------------
client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

# In-memory one-time auth codes for Microsoft mobile login (no JWT in URL)
# code -> {jwt, state, expires_at}
_auth_codes: dict = {}

def _cleanup_expired_codes():
    now = time.time()
    for k in list(_auth_codes.keys()):
        if _auth_codes[k]["expires_at"] < now:
            del _auth_codes[k]

# ---------------- App ----------------
app = FastAPI(title="Materiales OneDrive App")
api_router = APIRouter(prefix="/api")

# ---------------- Models ----------------
class UserRegister(BaseModel):
    email: EmailStr
    password: str = Field(..., min_length=8, max_length=128)
    name: Optional[str] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserOut(BaseModel):
    id: str
    email: str
    name: Optional[str] = None
    role: str = "user"
    color: Optional[str] = None
    role_id: Optional[str] = None
    role_name: Optional[str] = None
    permissions: Optional[List[str]] = None

class TokenOut(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserOut

class Material(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    row_index: int  # 1-based row in excel (excluding header)
    materiales: Optional[str] = None
    cliente: Optional[str] = None
    ubicacion: Optional[str] = None
    horas_prev: Optional[str] = None
    comercial: Optional[str] = None
    gestor: Optional[str] = None
    # editable
    fecha: Optional[str] = None
    entrega_recogida: Optional[str] = None
    total_parcial: Optional[str] = None
    tecnico: Optional[str] = None
    tecnicos: Optional[List[str]] = None
    comentarios: Optional[str] = None
    # manager assignment
    manager_id: Optional[str] = None
    manager_name: Optional[str] = None
    # project status
    project_status: Optional[str] = "pendiente"
    # imputed hours from events
    horas_imputadas: Optional[float] = 0
    # geo coordinates
    lat: Optional[float] = None
    lng: Optional[float] = None
    direccion: Optional[str] = None
    # attachments (PDFs, presupuestos)
    attachments: Optional[List[dict]] = None
    # meta
    sync_status: str = "synced"  # synced | pending | error
    updated_at: Optional[str] = None
    updated_by: Optional[str] = None

class MaterialUpdate(BaseModel):
    fecha: Optional[str] = None
    entrega_recogida: Optional[str] = None
    total_parcial: Optional[str] = None
    tecnico: Optional[str] = None
    tecnicos: Optional[List[str]] = None
    comentarios: Optional[str] = None
    manager_id: Optional[str] = None
    project_status: Optional[str] = None

class OneDriveStatus(BaseModel):
    connected: bool
    admin_email: Optional[str] = None
    last_import_at: Optional[str] = None
    last_push_at: Optional[str] = None
    file_path: str = ONEDRIVE_FILE_PATH
    file_name: Optional[str] = None
    using_share_url: bool = bool(ONEDRIVE_SHARE_URL)

# ---------------- Helpers ----------------
def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

MIN_PASSWORD_LENGTH = 8

def validate_password_strength(pw: str) -> None:
    if len(pw or "") < MIN_PASSWORD_LENGTH:
        raise HTTPException(400, f"Contraseña demasiado corta (mín. {MIN_PASSWORD_LENGTH})")

def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False

def create_jwt(user: dict) -> str:
    payload = {
        "sub": user["id"],
        "email": user["email"],
        "role": user.get("role", "user"),
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRE_HOURS),
    }
    return pyjwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def current_user(authorization: Optional[str] = Header(None)) -> dict:
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(401, "Missing bearer token")
    token = authorization.split(" ", 1)[1]
    try:
        payload = pyjwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except Exception:
        raise HTTPException(401, "Invalid token")
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(401, "User not found")
    return user

async def current_admin(user: dict = Depends(current_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    return user

# ---------------- Roles & Permissions ----------------
PERMISSIONS_CATALOG = [
    {"key": "proyectos.view", "label": "Ver Proyectos", "module": "Proyectos"},
    {"key": "proyectos.edit", "label": "Editar Proyectos (completo)", "module": "Proyectos"},
    {"key": "proyectos.editar_campo", "label": "Editar solo recogida/total/observaciones", "module": "Proyectos"},
    {"key": "calendario.view", "label": "Ver Calendario", "module": "Calendario"},
    {"key": "calendario.edit", "label": "Crear/Editar eventos", "module": "Calendario"},
    {"key": "calendario.assign", "label": "Asignar técnicos a eventos", "module": "Calendario"},
    {"key": "events.edit", "label": "Editar eventos existentes", "module": "Calendario"},
    {"key": "planos.view", "label": "Ver Planos", "module": "Planos"},
    {"key": "planos.edit", "label": "Editar Planos", "module": "Planos"},
    {"key": "planos.download", "label": "Descargar planos", "module": "Planos"},
    {"key": "presupuestos.view", "label": "Ver Presupuestos", "module": "Presupuestos"},
    {"key": "presupuestos.edit", "label": "Editar Presupuestos", "module": "Presupuestos"},
    {"key": "presupuestos.export", "label": "Exportar presupuestos a PDF", "module": "Presupuestos"},
    {"key": "sat.view", "label": "Ver CRM SAT", "module": "CRM SAT"},
    {"key": "sat.edit", "label": "Gestionar incidencias SAT", "module": "CRM SAT"},
    {"key": "sat.export", "label": "Exportar SAT a Excel", "module": "CRM SAT"},
    {"key": "users.manage", "label": "Gestionar usuarios", "module": "Administración"},
    {"key": "roles.manage", "label": "Gestionar roles y permisos", "module": "Administración"},
    {"key": "onedrive.manage", "label": "Conectar/Sincronizar OneDrive", "module": "Administración"},
    {"key": "chat.view", "label": "Ver el chat", "module": "Chat"},
    {"key": "chat.edit", "label": "Enviar mensajes en el chat", "module": "Chat"},
    {"key": "dashboard.view", "label": "Ver panel de datos", "module": "Dashboard"},
    {"key": "materiales.view_hours", "label": "Ver horas de proyectos", "module": "Proyectos"},
    {"key": "preciario.ver_precios", "label": "Ver precios en preciario", "module": "Preciario"},
    {"key": "preciario.edit", "label": "Editar descuentos y stock del preciario", "module": "Preciario"},
    {"key": "preciario.view", "label": "Ver preciario y productos", "module": "Preciario"},
    {"key": "notas.view", "label": "Ver y crear notas personales", "module": "Notas"},
    {"key": "documentos.manage", "label": "Gestionar documentos (fichas/manuales)", "module": "Documentación"},
]
ALL_PERMS = [p["key"] for p in PERMISSIONS_CATALOG]

NOTIFICATION_CATALOG = [
    {"key": "event_completed", "label": "Proyecto terminado", "module": "Calendario"},
    {"key": "event_pending_completion", "label": "Pendiente de terminar", "module": "Calendario"},
    {"key": "event_updated", "label": "Estado de evento actualizado", "module": "Calendario"},
    {"key": "sat_new", "label": "Nueva incidencia SAT", "module": "CRM SAT"},
    {"key": "sat_revived", "label": "Incidencia SAT reactivada", "module": "CRM SAT"},
    {"key": "chat_message", "label": "Mensajes de chat", "module": "Chat"},
]
ALL_NOTIFS = [n["key"] for n in NOTIFICATION_CATALOG]

NON_ADMIN_PERMS = [p for p in ALL_PERMS if p not in ("users.manage", "roles.manage")]
TECNICO_PERMS = ["proyectos.view", "proyectos.edit", "calendario.view", "calendario.edit", "planos.view", "planos.edit", "chat.view", "chat.edit", "planos.download", "events.edit"]
COMERCIAL_PERMS = ["presupuestos.view", "presupuestos.edit", "presupuestos.export", "proyectos.view", "chat.view", "chat.edit"]
SAT_PERMS = ["sat.view", "sat.edit", "sat.export", "chat.view", "chat.edit"]

# System roles seeded on startup. Admin role can NEVER be modified or deleted.
DEFAULT_ROLES = [
    {"key": "admin", "name": "Administrador principal", "permissions": ALL_PERMS, "notification_prefs": ALL_NOTIFS, "system": True, "locked": True},
    {"key": "gestor", "name": "Gestor", "permissions": NON_ADMIN_PERMS, "notification_prefs": ALL_NOTIFS, "system": True, "locked": False},
    {"key": "tecnico", "name": "Técnico", "permissions": TECNICO_PERMS, "notification_prefs": ["event_completed", "event_pending_completion", "chat_message"], "system": True, "locked": False},
    {"key": "comercial", "name": "Comercial", "permissions": COMERCIAL_PERMS, "notification_prefs": ["event_completed", "chat_message"], "system": True, "locked": False},
    {"key": "sat", "name": "SAT", "permissions": SAT_PERMS, "notification_prefs": ["sat_new", "sat_revived", "chat_message"], "system": True, "locked": False},
]

# Map legacy `role` field -> new role key
LEGACY_ROLE_MAP = {
    "admin": "admin",
    "user": "tecnico",
    "comercial": "comercial",
}


class RoleOut(BaseModel):
    id: str
    key: str
    name: str
    permissions: List[str]
    system: bool = False
    locked: bool = False
    created_at: Optional[str] = None
    notification_prefs: List[str] = []


class RoleCreate(BaseModel):
    name: str
    permissions: List[str] = []
    notification_prefs: List[str] = []


class RoleUpdate(BaseModel):
    name: Optional[str] = None
    permissions: Optional[List[str]] = None
    notification_prefs: Optional[List[str]] = None


async def ensure_default_roles_and_migrate():
    """Bootstrap default system roles and migrate legacy users to role_id."""
    # 1. Seed system roles (idempotent by `key`)
    for r in DEFAULT_ROLES:
        existing = await db.roles.find_one({"key": r["key"]})
        if not existing:
            await db.roles.insert_one({
                "id": str(uuid.uuid4()),
                "key": r["key"],
                "name": r["name"],
                "permissions": r["permissions"],
                "notification_prefs": r.get("notification_prefs", []),
                "system": r["system"],
                "locked": r.get("locked", False),
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
        else:
            # Always force admin role to have ALL permissions (locked, source of truth).
            if r["key"] == "admin":
                await db.roles.update_one(
                    {"key": "admin"},
                    {"$set": {"permissions": ALL_PERMS, "system": True, "locked": True, "name": existing.get("name") or r["name"]}},
                )
    # 2. Migrate users without role_id
    users_no_role = await db.users.find(
        {"$or": [{"role_id": {"$exists": False}}, {"role_id": None}]},
        {"_id": 0, "id": 1, "role": 1},
    ).to_list(5000)
    for u in users_no_role:
        legacy = (u.get("role") or "user").lower()
        key = LEGACY_ROLE_MAP.get(legacy, "tecnico")
        role = await db.roles.find_one({"key": key})
        if role:
            await db.users.update_one({"id": u["id"]}, {"$set": {"role_id": role["id"]}})


async def get_user_permissions(user: dict) -> List[str]:
    """Resolve permissions for a user via role_id; falls back to legacy `role` field."""
    role_id = user.get("role_id")
    role = None
    if role_id:
        role = await db.roles.find_one({"id": role_id})
    if not role:
        legacy = (user.get("role") or "user").lower()
        key = LEGACY_ROLE_MAP.get(legacy, "tecnico")
        role = await db.roles.find_one({"key": key})
    return role.get("permissions", []) if role else []


async def get_user_role_info(user: dict) -> dict:
    """Return {role_id, role_name, permissions, notification_prefs} for a user."""
    role_id = user.get("role_id")
    role = None
    if role_id:
        role = await db.roles.find_one({"id": role_id})
    if not role:
        legacy = (user.get("role") or "user").lower()
        key = LEGACY_ROLE_MAP.get(legacy, "tecnico")
        role = await db.roles.find_one({"key": key})
    return {
        "role_id": role.get("id") if role else None,
        "role_name": role.get("name") if role else None,
        "permissions": role.get("permissions", []) if role else [],
        "notification_prefs": role.get("notification_prefs", []) if role else [],
    }


async def should_notify_user(user_id: str, notif_type: str) -> bool:
    """Check if a user's role allows receiving this notification type."""
    user = await db.users.find_one({"id": user_id})
    if not user:
        return False
    info = await get_user_role_info(user)
    return notif_type in info.get("notification_prefs", [])


def require_permission(perm: str):
    """FastAPI dependency factory: ensures the current user has `perm`."""
    async def _dep(user: dict = Depends(current_user)) -> dict:
        perms = await get_user_permissions(user)
        if perm not in perms:
            raise HTTPException(403, f"No tienes permiso ({perm})")
        return user
    return _dep


def require_any_permission(*perms: str):
    """FastAPI dependency: passes if user has at least one of `perms`."""
    async def _dep(user: dict = Depends(current_user)) -> dict:
        user_perms = await get_user_permissions(user)
        if not any(p in user_perms for p in perms):
            raise HTTPException(403, "No tienes permiso para esta acción")
        return user
    return _dep

def _clean(v):
    if v is None:
        return None
    if isinstance(v, float) and math.isnan(v):
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s if s and s.lower() != "nan" else None

def _safe_float(v) -> float:
    """Parse float safely, returning 0 on any error."""
    try:
        return float(v or 0)
    except (ValueError, TypeError):
        return 0.0

# ---------------- Excel parsing ----------------
# Column mapping (A..M in the Excel):
# A=Materiales, B=CLIENTE, C=Ubicación Cliente, D=Horas PREV, E=Comercial,
# F=Gestor/a, G=Fecha, H=Entrega/recogida, I=Recogida Total/Parcial,
# J=Técnico, K=Comentarios, L=Columna1, M=Columna2
EDITABLE_COL_MAP = {
    "fecha": "G",
    "entrega_recogida": "H",
    "total_parcial": "I",
    "tecnico": "J",
    "comentarios": "K",
}

def parse_workbook(xlsx_bytes: bytes) -> List[dict]:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    rows = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # skip empty
        if not any(row):
            continue
        r = list(row) + [None] * (13 - len(row))
        rows.append({
            "row_index": idx,
            "materiales": _clean(r[0]),
            "cliente": _clean(r[1]),
            "ubicacion": _clean(r[2]),
            "horas_prev": _clean(r[3]),
            "comercial": _clean(r[4]),
            "gestor": _clean(r[5]),
            "fecha": _clean(r[6]),
            "entrega_recogida": _clean(r[7]),
            "total_parcial": _clean(r[8]),
            "tecnico": _clean(r[9]),
            "comentarios": _clean(r[10]),
        })
    return rows

# ---------------- Microsoft Graph ----------------
def _msal_app():
    return msal.ConfidentialClientApplication(
        client_id=MS_CLIENT_ID,
        client_credential=MS_CLIENT_SECRET,
        authority=MS_AUTHORITY,
    )

async def _get_onedrive_token() -> str:
    """Return a fresh access token using the stored refresh token (encrypted)."""
    doc = await db.onedrive_tokens.find_one({"_id": "admin"})
    if not doc:
        raise HTTPException(400, "OneDrive no conectado. El admin debe vincular OneDrive primero.")

    access_token_enc = doc.get("access_token_enc")
    refresh_token_enc = doc.get("refresh_token_enc")

    if access_token_enc and refresh_token_enc:
        refresh_token = _decrypt_onedrive_token(refresh_token_enc)
    elif "access_token" in doc and "refresh_token" in doc:
        access_token = doc["access_token"]
        refresh_token = doc["refresh_token"]
        await db.onedrive_tokens.update_one(
            {"_id": "admin"},
            {"$set": {
                "access_token_enc": _encrypt_onedrive_token(access_token),
                "refresh_token_enc": _encrypt_onedrive_token(refresh_token),
            },
            "$unset": {"access_token": "", "refresh_token": ""}}
        )
    else:
        raise HTTPException(400, "OneDrive no conectado. El admin debe vincular OneDrive primero.")

    app_msal = _msal_app()
    result = app_msal.acquire_token_by_refresh_token(refresh_token, scopes=MS_SCOPES)
    if "error" in result:
        raise HTTPException(401, f"Error refrescando token OneDrive: {result.get('error_description')}")
    if not result.get("access_token"):
        raise HTTPException(401, "Microsoft no devolvió access token al refrescar OneDrive")
    new_access = result["access_token"]
    new_refresh = result.get("refresh_token")
    update = {"access_token_enc": _encrypt_onedrive_token(new_access)}
    if new_refresh:
        update["refresh_token_enc"] = _encrypt_onedrive_token(new_refresh)
    await db.onedrive_tokens.update_one({"_id": "admin"}, {"$set": update})
    return new_access

async def _graph_get(url: str, token: str) -> httpx.Response:
    async with httpx.AsyncClient(timeout=60) as c:
        return await c.get(url, headers={"Authorization": f"Bearer {token}"})

async def _graph_request(method: str, url: str, token: str, **kw) -> httpx.Response:
    async with httpx.AsyncClient(timeout=120) as c:
        headers = kw.pop("headers", {})
        headers["Authorization"] = f"Bearer {token}"
        return await c.request(method, url, headers=headers, **kw)

async def _resolve_share_url(token: str) -> dict:
    """Resolve ONEDRIVE_SHARE_URL to driveId + itemId + name (cached in Mongo)."""
    cached = await db.onedrive_share_cache.find_one({"_id": ONEDRIVE_SHARE_URL})
    if cached:
        return cached
    # Encode URL per Graph spec: u! + base64url(no padding)
    b64 = base64.urlsafe_b64encode(ONEDRIVE_SHARE_URL.encode()).decode().rstrip("=")
    share_id = f"u!{b64}"
    url = f"{GRAPH_BASE}/shares/{share_id}/driveItem"
    r = await _graph_get(url, token)
    if r.status_code != 200:
        raise HTTPException(r.status_code, f"No se pudo resolver el enlace compartido: {r.text}")
    data = r.json()
    drive_id = data.get("parentReference", {}).get("driveId")
    item_id = data.get("id")
    name = data.get("name")
    if not drive_id or not item_id:
        raise HTTPException(500, f"Respuesta inesperada del share: {data}")
    doc = {"_id": ONEDRIVE_SHARE_URL, "drive_id": drive_id, "item_id": item_id, "name": name}
    await db.onedrive_share_cache.update_one({"_id": ONEDRIVE_SHARE_URL}, {"$set": doc}, upsert=True)
    return doc

async def _download_excel_from_onedrive() -> bytes:
    token = await _get_onedrive_token()
    if ONEDRIVE_SHARE_URL:
        info = await _resolve_share_url(token)
        url = f"{GRAPH_BASE}/drives/{info['drive_id']}/items/{info['item_id']}/content"
    else:
        url = f"{GRAPH_BASE}/me/drive/root:{ONEDRIVE_FILE_PATH}:/content"
    r = await _graph_get(url, token)
    if r.status_code not in (200, 302):
        raise HTTPException(r.status_code, "Error al descargar el archivo de OneDrive")
    return r.content

async def _upload_excel_to_onedrive(xlsx_bytes: bytes) -> None:
    token = await _get_onedrive_token()
    if ONEDRIVE_SHARE_URL:
        info = await _resolve_share_url(token)
        url = f"{GRAPH_BASE}/drives/{info['drive_id']}/items/{info['item_id']}/content"
    else:
        url = f"{GRAPH_BASE}/me/drive/root:{ONEDRIVE_FILE_PATH}:/content"
    async with httpx.AsyncClient(timeout=120) as c:
        r = await c.put(
            url,
            content=xlsx_bytes,
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            },
        )
    if r.status_code not in (200, 201):
        raise HTTPException(r.status_code, f"No se pudo subir Excel: {r.text}")

# ---------------- Rate limiter (simple in-memory) ----------------
import time as _time
from collections import defaultdict

_rate_window = 10  # seconds
_rate_max = 60     # max requests per window per IP
_rate_store: dict[str, list[float]] = defaultdict(list)

async def _rate_limit(request: Request, max_req: int = _rate_max):
    if not hasattr(request, "client"):
        return
    ip = request.client.host if request.client else "unknown"
    now = _time.time()
    _rate_store[ip] = [t for t in _rate_store[ip] if t > now - _rate_window]
    if len(_rate_store[ip]) >= max_req:
        raise HTTPException(429, "Demasiadas peticiones. Espera unos segundos.")
    _rate_store[ip].append(now)


@app.middleware("http")
async def rate_limit_middleware(request: Request, call_next):
    await _rate_limit(request, 200)
    return await call_next(request)
import asyncio
_sync_lock = asyncio.Lock()
_push_task: Optional[asyncio.Task] = None
_import_task: Optional[asyncio.Task] = None

def _fire_and_forget(coro):
    """Lanza una corrutina en background con logging de errores."""
    task = asyncio.create_task(coro)
    task.add_done_callback(lambda t: logging.getLogger(__name__).error(f"Background task failed: {t.exception()}", exc_info=t.exception()) if t.exception() else None)
    return task

async def _has_onedrive_link() -> bool:
    return await db.onedrive_tokens.find_one({"_id": "admin"}) is not None

async def _do_import() -> int:
    """Internal import — no auth, used by background job."""
    async with _sync_lock:
        xlsx_bytes = await _download_excel_from_onedrive()
        rows = parse_workbook(xlsx_bytes)
        existing = {m["row_index"]: m for m in await db.materiales.find({}, {"_id": 0}).to_list(10000)}
        docs = []
        imported_row_indexes = []
        imported_at = datetime.now(timezone.utc).isoformat()
        for r in rows:
            old = existing.get(r["row_index"])
            preserved = {}
            if old:
                for key in ("manager_id", "manager_name", "project_status", "tecnicos", "demo_seed"):
                    if key in old:
                        preserved[key] = old[key]
            docs.append({
                "id": old["id"] if old else str(uuid.uuid4()),
                **r,
                **preserved,
                "sync_status": "synced",
                "updated_at": imported_at,
                "updated_by": old.get("updated_by", "onedrive") if old else "onedrive",
            })
            imported_row_indexes.append(r["row_index"])

        if not docs:
            raise HTTPException(400, "El Excel de OneDrive no contiene filas para importar; no se modificaron los materiales actuales.")

        # Sincronizar carpetas con concurrencia limitada
        _sync_sem = asyncio.Semaphore(5)
        async def _sync_one(doc):
            async with _sync_sem:
                await _sync_project_folder(doc)
        for doc in docs:
            await db.materiales.update_one(
                {"row_index": doc["row_index"]},
                {"$set": doc},
                upsert=True,
            )
            _fire_and_forget(_sync_one(doc))
        await db.materiales.delete_many({"row_index": {"$nin": imported_row_indexes}})
        await db.sync_meta.update_one(
            {"_id": "meta"},
            {"$set": {"_id": "meta", "last_import_at": imported_at}},
            upsert=True,
        )
        return len(docs)

async def _do_push() -> int:
    """Internal push — merges pending edits into the OneDrive Excel."""
    async with _sync_lock:
        xlsx_bytes = await _download_excel_from_onedrive()
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        materials = await db.materiales.find({}, {"_id": 0}).to_list(10000)
        for m in materials:
            row = m["row_index"]
            for field, col in EDITABLE_COL_MAP.items():
                val = m.get(field)
                if val is not None:
                    ws[f"{col}{row}"] = val
        out = io.BytesIO()
        wb.save(out)
        await _upload_excel_to_onedrive(out.getvalue())
        await db.materiales.update_many({}, {"$set": {"sync_status": "synced"}})
        await db.sync_meta.update_one(
            {"_id": "meta"},
            {"$set": {"_id": "meta", "last_push_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True,
        )
        return len(materials)

async def _delayed_push():
    try:
        await asyncio.sleep(AUTO_PUSH_DELAY_SEC)
        if not await _has_onedrive_link():
            return
        logger = logging.getLogger(__name__)
        try:
            n = await _do_push()
            logger.info(f"Auto-push: {n} filas sincronizadas con OneDrive")
        except Exception as e:
            logger.error(f"Auto-push falló: {e}")
    except asyncio.CancelledError:
        pass

def schedule_auto_push():
    global _push_task
    if _push_task and not _push_task.done():
        _push_task.cancel()
    _push_task = _fire_and_forget(_delayed_push())

async def maybe_auto_import():
    """If OneDrive linked and last import older than AUTO_IMPORT_INTERVAL_SEC, import in background."""
    global _import_task
    if not await _has_onedrive_link():
        return
    meta = await db.sync_meta.find_one({"_id": "meta"})
    last = meta.get("last_import_at") if meta else None
    if last:
        try:
            last_dt = datetime.fromisoformat(last)
            if (datetime.now(timezone.utc) - last_dt).total_seconds() < AUTO_IMPORT_INTERVAL_SEC:
                return
        except Exception:
            pass
    # Don't run two imports simultaneously
    if _import_task and not _import_task.done():
        return
    async def _runner():
        try:
            n = await _do_import()
            logging.getLogger(__name__).info(f"Auto-import: {n} filas traídas de OneDrive")
        except Exception as e:
            logging.getLogger(__name__).error(f"Auto-import falló: {e}")
    _import_task = _fire_and_forget(_runner())

# ---------------- Auth routes ----------------
@api_router.post("/auth/register", response_model=TokenOut)
async def register(payload: UserRegister):
    count = await db.users.count_documents({})
    # Only allow public register when DB is empty (bootstrap first admin).
    if count > 0:
        raise HTTPException(403, "Registro público deshabilitado. Pide a un administrador que cree tu cuenta.")
    validate_password_strength(payload.password)
    user = {
        "id": str(uuid.uuid4()),
        "email": payload.email.lower(),
        "name": payload.name or payload.email.split("@")[0],
        "password": hash_password(payload.password),
        "role": "admin",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(user)
    token = create_jwt(user)
    return TokenOut(access_token=token, user=UserOut(id=user["id"], email=user["email"], name=user["name"], role=user["role"], color=user.get("color")))

@api_router.post("/auth/login", response_model=TokenOut)
async def login(payload: UserLogin):
    user = await db.users.find_one({"email": payload.email.lower()})
    if not user or not verify_password(payload.password, user["password"]):
        raise HTTPException(401, "Credenciales inválidas")
    token = create_jwt(user)
    info = await get_user_role_info(user)
    return TokenOut(
        access_token=token,
        user=UserOut(
            id=user["id"], email=user["email"], name=user.get("name"),
            role=user.get("role", "user"), color=user.get("color"),
            role_id=info["role_id"], role_name=info["role_name"], permissions=info["permissions"],
        ),
    )

class TechnicianOut(BaseModel):
    id: str
    name: str
    email: str

@api_router.get("/technicians", response_model=List[TechnicianOut])
async def list_technicians(user: dict = Depends(current_user)):
    """List all app users as potential technicians (any authenticated user can read)."""
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    users.sort(key=lambda u: (u.get("name") or u.get("email") or "").lower())
    return [
        TechnicianOut(
            id=u["id"],
            name=u.get("name") or u.get("email", ""),
            email=u.get("email", ""),
        ) for u in users
    ]

@api_router.get("/managers", response_model=List[TechnicianOut])
async def list_managers(user: dict = Depends(current_user)):
    """List admin users (potential project managers / gestores)."""
    users = await db.users.find({"role": "admin"}, {"_id": 0, "password": 0}).to_list(500)
    users.sort(key=lambda u: (u.get("name") or u.get("email") or "").lower())
    return [
        TechnicianOut(
            id=u["id"],
            name=u.get("name") or u.get("email", ""),
            email=u.get("email", ""),
        ) for u in users
    ]



@api_router.get("/auth/me", response_model=UserOut)
async def me(user: dict = Depends(current_user)):
    info = await get_user_role_info(user)
    return UserOut(
        id=user["id"], email=user["email"], name=user.get("name"),
        role=user.get("role", "user"), color=user.get("color"),
        role_id=info["role_id"], role_name=info["role_name"], permissions=info["permissions"],
    )

# ---------------- User management (admin only) ----------------
class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: Optional[str] = None
    role: Optional[Literal["admin", "user", "comercial"]] = None  # legacy
    role_id: Optional[str] = None
    color: Optional[str] = None

class UserPatch(BaseModel):
    name: Optional[str] = None
    role: Optional[Literal["admin", "user", "comercial"]] = None  # legacy
    role_id: Optional[str] = None
    color: Optional[str] = None

class PasswordReset(BaseModel):
    password: str

class UserListItem(BaseModel):
    id: str
    email: str
    name: Optional[str] = None
    role: str
    role_id: Optional[str] = None
    role_name: Optional[str] = None
    color: Optional[str] = None
    created_at: Optional[str] = None

# Default color palette rotated per newly created user
DEFAULT_USER_COLORS = [
    "#3B82F6",  # blue
    "#10B981",  # green
    "#F59E0B",  # amber
    "#EF4444",  # red
    "#8B5CF6",  # violet
    "#EC4899",  # pink
    "#06B6D4",  # cyan
    "#F97316",  # orange
    "#84CC16",  # lime
    "#14B8A6",  # teal
]

def _next_default_color(existing_colors: List[str]) -> str:
    used = set(existing_colors or [])
    for c in DEFAULT_USER_COLORS:
        if c not in used:
            return c
    return DEFAULT_USER_COLORS[len(used) % len(DEFAULT_USER_COLORS)]

@api_router.get("/users", response_model=List[UserListItem])
async def list_users(admin: dict = Depends(require_permission("users.manage"))):
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    users.sort(key=lambda u: u.get("created_at", ""))
    # Enrich with role_name
    role_ids = list({u.get("role_id") for u in users if u.get("role_id")})
    role_map = {}
    if role_ids:
        async for r in db.roles.find({"id": {"$in": role_ids}}):
            role_map[r["id"]] = r.get("name")
    out = []
    for u in users:
        u["role_name"] = role_map.get(u.get("role_id"))
        out.append(u)
    return out

async def _resolve_role_for_user(role_id: Optional[str], legacy_role: Optional[str]) -> dict:
    """Return the role doc to assign. Prefers role_id; falls back to legacy mapping."""
    role = None
    if role_id:
        role = await db.roles.find_one({"id": role_id})
        if not role:
            raise HTTPException(400, "Rol inválido")
    elif legacy_role:
        key = LEGACY_ROLE_MAP.get(legacy_role.lower(), "tecnico")
        role = await db.roles.find_one({"key": key})
    else:
        role = await db.roles.find_one({"key": "tecnico"})
    if not role:
        raise HTTPException(500, "Roles del sistema no inicializados")
    return role

@api_router.post("/users", response_model=UserListItem)
async def create_user(payload: UserCreate, admin: dict = Depends(require_permission("users.manage"))):
    existing = await db.users.find_one({"email": payload.email.lower()})
    if existing:
        raise HTTPException(400, "Email ya registrado")
    validate_password_strength(payload.password)
    role = await _resolve_role_for_user(payload.role_id, payload.role)
    # Pick a unique default color if none provided
    if payload.color:
        color = payload.color
    else:
        existing_colors = [u.get("color") for u in await db.users.find({}, {"_id": 0, "color": 1}).to_list(1000)]
        color = _next_default_color([c for c in existing_colors if c])
    # Legacy `role` field: gestor also marked as "admin" so it passes inline
    # `role == "admin"` checks scattered in the codebase. The real ACL is on role_id.
    legacy_role = "admin" if role["key"] in ("admin", "gestor") else ("comercial" if role["key"] == "comercial" else "user")
    user = {
        "id": str(uuid.uuid4()),
        "email": payload.email.lower(),
        "name": payload.name or payload.email.split("@")[0],
        "password": hash_password(payload.password),
        "role": legacy_role,
        "role_id": role["id"],
        "color": color,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(user)
    return UserListItem(
        id=user["id"], email=user["email"], name=user["name"],
        role=user["role"], role_id=user["role_id"], role_name=role["name"],
        color=user["color"], created_at=user["created_at"],
    )

@api_router.patch("/users/{uid}", response_model=UserListItem)
async def update_user(uid: str, payload: UserPatch, admin: dict = Depends(require_permission("users.manage"))):
    target = await db.users.find_one({"id": uid})
    if not target:
        raise HTTPException(404, "Usuario no encontrado")
    upd = {k: v for k, v in payload.dict().items() if v is not None and k != "role_id" and k != "role"}
    new_role = None
    if payload.role_id is not None or payload.role is not None:
        new_role = await _resolve_role_for_user(payload.role_id, payload.role)
        upd["role_id"] = new_role["id"]
        upd["role"] = "admin" if new_role["key"] in ("admin", "gestor") else ("comercial" if new_role["key"] == "comercial" else "user")
    if not upd:
        raise HTTPException(400, "Nada que actualizar")
    # Prevent removing last admin (super)
    if new_role and new_role["key"] != "admin":
        if (target.get("role") == "admin") or (await db.roles.find_one({"id": target.get("role_id"), "key": "admin"})):
            remaining = await db.users.count_documents({"role": "admin", "id": {"$ne": uid}})
            if remaining == 0:
                raise HTTPException(400, "No puedes quitar el rol Administrador principal al último administrador")
    await db.users.update_one({"id": uid}, {"$set": upd})
    updated = await db.users.find_one({"id": uid}, {"_id": 0, "password": 0})
    if updated.get("role_id"):
        rdoc = await db.roles.find_one({"id": updated["role_id"]})
        updated["role_name"] = rdoc.get("name") if rdoc else None
    return updated

@api_router.post("/users/{uid}/reset-password")
async def reset_password(uid: str, payload: PasswordReset, admin: dict = Depends(require_permission("users.manage"))):
    target = await db.users.find_one({"id": uid})
    if not target:
        raise HTTPException(404, "Usuario no encontrado")
    validate_password_strength(payload.password)
    await db.users.update_one({"id": uid}, {"$set": {"password": hash_password(payload.password)}})
    return {"ok": True}

@api_router.delete("/users/{uid}")
async def delete_user(uid: str, admin: dict = Depends(require_permission("users.manage"))):
    if uid == admin["id"]:
        raise HTTPException(400, "No puedes eliminarte a ti mismo")
    target = await db.users.find_one({"id": uid})
    if not target:
        raise HTTPException(404, "Usuario no encontrado")
    # Prevent deleting last admin
    if target.get("role") == "admin":
        remaining = await db.users.count_documents({"role": "admin", "id": {"$ne": uid}})
        if remaining == 0:
            raise HTTPException(400, "No puedes eliminar al último administrador")
    await db.users.delete_one({"id": uid})
    return {"ok": True}

# ---------------- Roles & Permissions API ----------------
@api_router.get("/permissions")
async def list_permissions(user: dict = Depends(current_user)):
    """Public catalog of all permissions in the app, grouped by module."""
    return {"permissions": PERMISSIONS_CATALOG}

@api_router.get("/notification-prefs")
async def list_notification_prefs(user: dict = Depends(current_user)):
    """Public catalog of all notification preferences."""
    return {"notifications": NOTIFICATION_CATALOG}

@api_router.get("/roles", response_model=List[RoleOut])
async def list_roles(user: dict = Depends(current_user)):
    """Anyone authenticated can see roles (used by user-edit dropdown)."""
    roles = await db.roles.find({}, {"_id": 0}).to_list(200)
    # Stable order: system first by key order, then customs by created_at
    sys_order = {"admin": 0, "gestor": 1, "tecnico": 2, "comercial": 3, "sat": 4}
    roles.sort(key=lambda r: (0 if r.get("system") else 1, sys_order.get(r.get("key"), 99), r.get("created_at", "")))
    return roles

@api_router.post("/roles", response_model=RoleOut)
async def create_role(payload: RoleCreate, admin: dict = Depends(require_permission("roles.manage"))):
    name = payload.name.strip()
    if not name:
        raise HTTPException(400, "Nombre obligatorio")
    bad = [p for p in payload.permissions if p not in ALL_PERMS]
    if bad:
        raise HTTPException(400, f"Permisos inválidos: {bad}")
    role_notifs = [n for n in payload.notification_prefs if n in ALL_NOTIFS]
    base_key = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_") or "role"
    key = base_key
    suffix = 1
    while await db.roles.find_one({"key": key}):
        suffix += 1
        key = f"{base_key}_{suffix}"
    role = {
        "id": str(uuid.uuid4()),
        "key": key,
        "name": name,
        "permissions": list(dict.fromkeys(payload.permissions)),
        "notification_prefs": role_notifs,
        "system": False,
        "locked": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.roles.insert_one(role)
    return role

@api_router.patch("/roles/{rid}", response_model=RoleOut)
async def update_role(rid: str, payload: RoleUpdate, admin: dict = Depends(require_permission("roles.manage"))):
    role = await db.roles.find_one({"id": rid})
    if not role:
        raise HTTPException(404, "Rol no encontrado")
    if role.get("locked"):
        raise HTTPException(400, "El rol Administrador principal no se puede modificar")
    upd = {}
    if payload.name is not None:
        nm = payload.name.strip()
        if not nm:
            raise HTTPException(400, "Nombre vacío")
        upd["name"] = nm
    if payload.permissions is not None:
        bad = [p for p in payload.permissions if p not in ALL_PERMS]
        if bad:
            raise HTTPException(400, f"Permisos inválidos: {bad}")
        upd["permissions"] = list(dict.fromkeys(payload.permissions))
    if payload.notification_prefs is not None:
        upd["notification_prefs"] = [n for n in payload.notification_prefs if n in ALL_NOTIFS]
    if not upd:
        raise HTTPException(400, "Nada que actualizar")
    await db.roles.update_one({"id": rid}, {"$set": upd})
    updated = await db.roles.find_one({"id": rid}, {"_id": 0})
    return updated

@api_router.delete("/roles/{rid}")
async def delete_role(rid: str, admin: dict = Depends(require_permission("roles.manage"))):
    role = await db.roles.find_one({"id": rid})
    if not role:
        raise HTTPException(404, "Rol no encontrado")
    if role.get("system"):
        raise HTTPException(400, "No puedes eliminar un rol del sistema")
    # Reassign any user using this role to "tecnico"
    fallback = await db.roles.find_one({"key": "tecnico"})
    fallback_id = fallback["id"] if fallback else None
    if fallback_id:
        await db.users.update_many(
            {"role_id": rid},
            {"$set": {"role_id": fallback_id, "role": "user"}},
        )
    await db.roles.delete_one({"id": rid})
    return {"ok": True}

# ---------------- Plans & Stamps ----------------
class PlanCreate(BaseModel):
    title: str
    data: Optional[dict] = None  # {shapes: [...]}
    material_id: Optional[str] = None
    source_event_id: Optional[str] = None
    source_attachment_id: Optional[str] = None

class PlanPatch(BaseModel):
    title: Optional[str] = None
    data: Optional[dict] = None
    source_attachment_id: Optional[str] = None
    material_id: Optional[str] = None
    source_event_id: Optional[str] = None

class PlanOut(BaseModel):
    id: str
    title: str
    data: dict
    created_at: str
    updated_at: str
    created_by: str
    material_id: Optional[str] = None
    source_event_id: Optional[str] = None
    source_attachment_id: Optional[str] = None

class PlanListItem(BaseModel):
    id: str
    title: str
    created_at: str
    updated_at: str
    created_by: str
    shape_count: int = 0
    thumbnail: Optional[str] = None
    material_id: Optional[str] = None
    source_event_id: Optional[str] = None

class StampCreate(BaseModel):
    name: str
    image_base64: str  # data URI: "data:image/png;base64,...."

class StampOut(BaseModel):
    id: str
    name: str
    is_builtin: bool
    image_base64: Optional[str] = None  # null for builtin (frontend knows icons)
    icon_key: Optional[str] = None      # key for builtin SVG on frontend

BUILTIN_STAMPS = [
    # Aberturas
    {"id": "builtin_door", "name": "Puerta", "is_builtin": True, "icon_key": "door"},
    {"id": "builtin_door_double", "name": "Puerta doble", "is_builtin": True, "icon_key": "door_double"},
    {"id": "builtin_door_sliding", "name": "Puerta corredera", "is_builtin": True, "icon_key": "door_sliding"},
    {"id": "builtin_window", "name": "Ventana", "is_builtin": True, "icon_key": "window"},
    {"id": "builtin_stairs", "name": "Escalera", "is_builtin": True, "icon_key": "stairs"},
    {"id": "builtin_door_handle", "name": "Manilla puerta", "is_builtin": True, "icon_key": "door_handle"},
    # Control de accesos
    {"id": "builtin_card_reader", "name": "Lector tarjeta", "is_builtin": True, "icon_key": "card_reader"},
    {"id": "builtin_keypad", "name": "Teclado", "is_builtin": True, "icon_key": "keypad"},
    {"id": "builtin_fingerprint", "name": "Lector huella", "is_builtin": True, "icon_key": "fingerprint"},
    {"id": "builtin_face_reader", "name": "Lector facial", "is_builtin": True, "icon_key": "face_reader"},
    {"id": "builtin_maglock", "name": "Electroimán", "is_builtin": True, "icon_key": "maglock"},
    {"id": "builtin_electric_strike", "name": "Cerradura eléc.", "is_builtin": True, "icon_key": "electric_strike"},
    {"id": "builtin_exit_button", "name": "Pulsador salida", "is_builtin": True, "icon_key": "exit_button"},
    {"id": "builtin_emergency_button", "name": "Botón emergencia", "is_builtin": True, "icon_key": "emergency_button"},
    {"id": "builtin_intercom", "name": "Interfono", "is_builtin": True, "icon_key": "intercom"},
    {"id": "builtin_video_intercom", "name": "Videoportero", "is_builtin": True, "icon_key": "video_intercom"},
    {"id": "builtin_controller", "name": "Controladora", "is_builtin": True, "icon_key": "controller"},
    {"id": "builtin_door_contact", "name": "Contacto mag.", "is_builtin": True, "icon_key": "door_contact"},
    {"id": "builtin_turnstile", "name": "Torniquete", "is_builtin": True, "icon_key": "turnstile"},
    {"id": "builtin_bollard", "name": "Bolardo retráctil", "is_builtin": True, "icon_key": "bollard"},
    {"id": "builtin_barrier", "name": "Barrera vehículo", "is_builtin": True, "icon_key": "barrier"},
    {"id": "builtin_gate_motor", "name": "Motor puerta", "is_builtin": True, "icon_key": "gate_motor"},
    # Seguridad
    {"id": "builtin_camera", "name": "Cámara", "is_builtin": True, "icon_key": "camera"},
    {"id": "builtin_motion_sensor", "name": "Sensor mov.", "is_builtin": True, "icon_key": "motion_sensor"},
    {"id": "builtin_smoke_detector", "name": "Detector humo", "is_builtin": True, "icon_key": "smoke_detector"},
    {"id": "builtin_siren", "name": "Sirena", "is_builtin": True, "icon_key": "siren"},
    # Electricidad
    {"id": "builtin_outlet", "name": "Enchufe", "is_builtin": True, "icon_key": "outlet"},
    {"id": "builtin_switch", "name": "Interruptor", "is_builtin": True, "icon_key": "switch"},
    {"id": "builtin_light", "name": "Luminaria", "is_builtin": True, "icon_key": "light"},
    {"id": "builtin_light_wall", "name": "Luz pared", "is_builtin": True, "icon_key": "light_wall"},
    # Referencia
    {"id": "builtin_north_arrow", "name": "Norte", "is_builtin": True, "icon_key": "north_arrow"},
    {"id": "builtin_column", "name": "Columna", "is_builtin": True, "icon_key": "column"},
    {"id": "builtin_column_round", "name": "Columna redonda", "is_builtin": True, "icon_key": "column_round"},
    {"id": "builtin_dimension", "name": "Cota", "is_builtin": True, "icon_key": "dimension"},
]

@api_router.get("/plans", response_model=List[PlanListItem])
async def list_plans(user: dict = Depends(require_permission("planos.view"))):
    plans = await db.plans.find({}, {"_id": 0, "data": 0}).to_list(2000)
    all_data = await db.plans.find({}, {"_id": 0, "id": 1, "data": 1}).to_list(2000)
    counts = {p["id"]: len((p.get("data") or {}).get("shapes", [])) for p in all_data}
    # Generar thumbnails desde background (en thread pool para no bloquear)
    async def _gen_thumbnail(pl: dict) -> Optional[tuple]:
        bg = (pl.get("data") or {}).get("background")
        if not bg:
            return None
        b64_data = bg if isinstance(bg, str) else bg.get("data_uri", "")
        if not b64_data:
            return None
        if "," in b64_data:
            b64 = b64_data.split(",", 1)[1]
        else:
            b64 = b64_data
        def _process():
            raw = base64.b64decode(b64)
            img = Image.open(io.BytesIO(raw))
            img.thumbnail((200, 200), Image.LANCZOS)
            if img.mode == "RGBA":
                img = img.convert("RGB")
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=70)
            return base64.b64encode(buf.getvalue()).decode()
        try:
            thumb_b64 = await asyncio.to_thread(_process)
            return (pl["id"], "data:image/jpeg;base64," + thumb_b64)
        except Exception:
            return None

    tasks = [_gen_thumbnail(p) for p in all_data]
    results = await asyncio.gather(*tasks)
    thumbs = {tid: thumb for r in results if r for tid, thumb in [r]}
    plans.sort(key=lambda p: p.get("updated_at", ""), reverse=True)
    out: List[PlanListItem] = []
    for p in plans:
        out.append(PlanListItem(
            id=p["id"], title=p["title"],
            created_at=p.get("created_at", ""), updated_at=p.get("updated_at", ""),
            created_by=p.get("created_by", ""),
            shape_count=counts.get(p["id"], 0),
            thumbnail=thumbs.get(p["id"]),
            material_id=p.get("material_id"),
            source_event_id=p.get("source_event_id"),
        ))
    return out

@api_router.post("/plans", response_model=PlanOut)
async def create_plan(payload: PlanCreate, user: dict = Depends(require_permission("planos.edit"))):
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "title": payload.title.strip() or "Plano sin título",
        "data": payload.data or {"shapes": []},
        "material_id": payload.material_id,
        "source_event_id": payload.source_event_id,
        "source_attachment_id": payload.source_attachment_id,
        "created_at": now,
        "updated_at": now,
        "created_by": user["email"],
    }
    await db.plans.insert_one(doc)
    return PlanOut(**{k: v for k, v in doc.items() if k != "_id"})

@api_router.get("/plans/{pid}", response_model=PlanOut)
async def get_plan(pid: str, user: dict = Depends(require_permission("planos.view"))):
    p = await db.plans.find_one({"id": pid}, {"_id": 0})
    if not p:
        raise HTTPException(404, "Plano no encontrado")
    # Ensure optional fields are present in response
    p.setdefault("material_id", None)
    p.setdefault("source_event_id", None)
    p.setdefault("source_attachment_id", None)
    return p

@api_router.patch("/plans/{pid}", response_model=PlanOut)
async def update_plan(pid: str, payload: PlanPatch, user: dict = Depends(require_permission("planos.edit"))):
    upd = {k: v for k, v in payload.dict().items() if v is not None}
    if not upd:
        raise HTTPException(400, "Nada que actualizar")
    upd["updated_at"] = datetime.now(timezone.utc).isoformat()
    res = await db.plans.update_one({"id": pid}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(404, "Plano no encontrado")
    p = await db.plans.find_one({"id": pid}, {"_id": 0})
    return p

@api_router.delete("/plans/{pid}")
async def delete_plan(pid: str, user: dict = Depends(require_permission("planos.edit"))):
    res = await db.plans.delete_one({"id": pid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Plano no encontrado")
    return {"ok": True}

MAX_BACKGROUND_MB = 25

class BackgroundUpload(BaseModel):
    file_base64: str  # raw base64 (no data URI prefix)
    mime_type: str   # "image/jpeg" | "image/png" | "application/pdf"

@api_router.post("/plans/{pid}/background")
async def upload_background(pid: str, payload: BackgroundUpload, user: dict = Depends(require_permission("planos.edit"))):
    plan = await db.plans.find_one({"id": pid})
    if not plan:
        raise HTTPException(404, "Plano no encontrado")
    # Size check before decoding (approximate: base64 → raw is ~ 3/4 ratio)
    est_raw = (len(payload.file_base64) * 3) // 4
    if est_raw > MAX_BACKGROUND_MB * 1024 * 1024:
        raise HTTPException(413, f"El archivo excede {MAX_BACKGROUND_MB}MB")
    try:
        raw = base64.b64decode(payload.file_base64)
    except Exception:
        raise HTTPException(400, "Base64 inválido")
    if len(raw) > MAX_BACKGROUND_MB * 1024 * 1024:
        raise HTTPException(413, f"El archivo excede {MAX_BACKGROUND_MB}MB")
    mime = payload.mime_type.lower()
    if mime == "application/pdf":
        try:
            pdf = pdfium.PdfDocument(raw)
            page = pdf[0]
            pil = page.render(scale=2).to_pil()
        except Exception as e:
            raise HTTPException(400, f"No se pudo leer el PDF: {e}")
        buf = io.BytesIO()
        pil.save(buf, format="PNG")
        png_b64 = base64.b64encode(buf.getvalue()).decode()
        bg = {"type": "image", "data_uri": f"data:image/png;base64,{png_b64}",
              "width": pil.width, "height": pil.height}
    elif mime in ("image/jpeg", "image/jpg", "image/png", "image/webp"):
        try:
            pil = Image.open(io.BytesIO(raw))
            width, height = pil.size
        except Exception as e:
            raise HTTPException(400, f"Imagen inválida: {e}")
        ext = "jpeg" if "jpeg" in mime or "jpg" in mime else "png" if "png" in mime else "webp"
        data_uri = f"data:image/{ext};base64,{payload.file_base64}"
        bg = {"type": "image", "data_uri": data_uri, "width": width, "height": height}
    else:
        raise HTTPException(400, "Tipo no soportado (usar JPG, PNG o PDF)")
    # update plan.data.background
    data = plan.get("data") or {"shapes": []}
    data["background"] = bg
    await db.plans.update_one({"id": pid}, {
        "$set": {"data": data, "updated_at": datetime.now(timezone.utc).isoformat()}
    })
    return {"ok": True, "background": bg}

@api_router.delete("/plans/{pid}/background")
async def remove_background(pid: str, user: dict = Depends(require_permission("planos.edit"))):
    plan = await db.plans.find_one({"id": pid})
    if not plan:
        raise HTTPException(404, "Plano no encontrado")
    data = plan.get("data") or {"shapes": []}
    if "background" in data:
        del data["background"]
    await db.plans.update_one({"id": pid}, {
        "$set": {"data": data, "updated_at": datetime.now(timezone.utc).isoformat()}
    })
    return {"ok": True}

# ---------------- Calendar Events ----------------
class RecurrenceRule(BaseModel):
    type: Literal["none", "daily", "weekly", "monthly"] = "none"
    until: Optional[str] = None  # ISO date (YYYY-MM-DD) inclusive

class EventCreate(BaseModel):
    title: str
    start_at: str  # ISO
    end_at: str    # ISO
    description: Optional[str] = None
    material_id: Optional[str] = None
    assigned_user_ids: List[str] = []
    manager_id: Optional[str] = None  # gestor del proyecto (admin)
    recurrence: Optional[RecurrenceRule] = None
    # Status del trabajo. "in_progress" por defecto. Valores aceptados:
    #   in_progress       — todavía en curso
    #   completed         — proyecto terminado (oscurece el evento)
    #   pending_completion — pendiente de terminar (resalta el evento)
    status: Optional[str] = "in_progress"
    seguimiento: Optional[str] = None  # observaciones del técnico
    hours: Optional[float] = None  # horas asignadas al evento

class EventPatch(BaseModel):
    title: Optional[str] = None
    start_at: Optional[str] = None
    end_at: Optional[str] = None
    description: Optional[str] = None
    material_id: Optional[str] = None
    assigned_user_ids: Optional[List[str]] = None
    manager_id: Optional[str] = None
    recurrence: Optional[RecurrenceRule] = None
    status: Optional[str] = None
    seguimiento: Optional[str] = None
    hours: Optional[float] = None
    budget_id: Optional[str] = None

class EventOut(BaseModel):
    id: str
    title: str
    start_at: str
    end_at: str
    description: Optional[str] = None
    material_id: Optional[str] = None
    material: Optional[dict] = None
    assigned_user_ids: List[str] = []
    assigned_users: List[dict] = []
    manager_id: Optional[str] = None
    manager: Optional[dict] = None
    recurrence: Optional[dict] = None
    attachments: List[dict] = []  # metadata only (no base64) for list views
    base_event_id: Optional[str] = None  # for expanded occurrences
    created_by: str
    created_at: str
    status: Optional[str] = "in_progress"
    seguimiento: Optional[str] = None
    hours: Optional[float] = None
    budget_id: Optional[str] = None

class AttachmentUpload(BaseModel):
    filename: str
    mime_type: str
    base64: str  # data only, no data: prefix

# ---------------------------------------------------------------------------
# Guardias (técnicos de guardia por día). Independientes de eventos.
# No cuentan como "día ocupado" en el cálculo de planificación mensual.
# ---------------------------------------------------------------------------
class GuardCreate(BaseModel):
    date: str  # YYYY-MM-DD
    user_id: str
    note: Optional[str] = None

class GuardOut(BaseModel):
    id: str
    date: str
    user_id: str
    user_name: Optional[str] = None
    user_color: Optional[str] = None
    note: Optional[str] = None
    created_at: str
    created_by: Optional[str] = None


MAX_ATTACHMENT_MB = 15  # per file

def _strip_attachments(ev: dict) -> dict:
    """Return event with attachments metadata only (no base64)."""
    atts = ev.get("attachments") or []
    meta = []
    for a in atts:
        meta.append({
            "id": a.get("id"),
            "filename": a.get("filename"),
            "mime_type": a.get("mime_type"),
            "size": a.get("size"),
            "uploaded_at": a.get("uploaded_at"),
            "uploaded_by": a.get("uploaded_by"),
        })
    ev["attachments"] = meta
    return ev

async def _attach_material(ev: dict) -> dict:
    mid = ev.get("material_id")
    if mid:
        m = await db.materiales.find_one({"id": mid}, {"_id": 0})
        if m:
            ev["material"] = m
    return ev

async def _attach_users(ev: dict) -> dict:
    ids = ev.get("assigned_user_ids") or []
    if ids:
        users = await db.users.find({"id": {"$in": ids}}, {"_id": 0, "password": 0}).to_list(200)
        ev["assigned_users"] = [
            {"id": u["id"], "email": u["email"], "name": u.get("name"), "color": u.get("color")}
            for u in users
        ]
    else:
        ev["assigned_users"] = []
    # Attach manager (gestor)
    mgr_id = ev.get("manager_id")
    if mgr_id:
        mgr = await db.users.find_one({"id": mgr_id}, {"_id": 0, "password": 0})
        if mgr:
            ev["manager"] = {
                "id": mgr["id"],
                "email": mgr["email"],
                "name": mgr.get("name"),
                "color": mgr.get("color"),
                "role": mgr.get("role"),
            }
        else:
            ev["manager"] = None
    else:
        ev["manager"] = None
    return ev

def _expand_recurrence(ev: dict, from_dt: datetime, to_dt: datetime) -> List[dict]:
    """Return list of virtual occurrences inside [from_dt, to_dt)."""
    rec = ev.get("recurrence") or {}
    rtype = rec.get("type", "none")
    if rtype == "none":
        start = datetime.fromisoformat(ev["start_at"].replace("Z", "+00:00"))
        # Defensive: normalise tz-naive ISO strings (some legacy rows) to UTC
        # so comparisons with the tz-aware from_dt/to_dt never fail.
        if start.tzinfo is None:
            start = start.replace(tzinfo=timezone.utc)
        if from_dt <= start < to_dt:
            return [ev]
        return []
    # compute step
    base_start = datetime.fromisoformat(ev["start_at"].replace("Z", "+00:00"))
    base_end = datetime.fromisoformat(ev["end_at"].replace("Z", "+00:00"))
    if base_start.tzinfo is None:
        base_start = base_start.replace(tzinfo=timezone.utc)
    if base_end.tzinfo is None:
        base_end = base_end.replace(tzinfo=timezone.utc)
    duration = base_end - base_start
    until_str = rec.get("until")
    until_dt = None
    if until_str:
        try:
            until_dt = datetime.fromisoformat(until_str + "T23:59:59+00:00")
        except Exception:
            until_dt = None
    results: List[dict] = []
    cur = base_start
    # Skip ahead if base_start far before from_dt
    # Iterate up to max 500 occurrences to prevent runaway
    for i in range(500):
        if until_dt and cur > until_dt:
            break
        if cur >= to_dt:
            break
        if cur + duration > from_dt:  # occurrence overlaps window
            occ = dict(ev)
            occ["id"] = f"{ev['id']}:{cur.date().isoformat()}"
            occ["base_event_id"] = ev["id"]
            occ["start_at"] = cur.isoformat().replace("+00:00", "Z")
            occ["end_at"] = (cur + duration).isoformat().replace("+00:00", "Z")
            results.append(occ)
        # next
        if rtype == "daily":
            cur = cur + timedelta(days=1)
        elif rtype == "weekly":
            cur = cur + timedelta(days=7)
        elif rtype == "monthly":
            # advance ~1 month keeping day of month
            y = cur.year
            m = cur.month + 1
            if m > 12:
                m -= 12
                y += 1
            try:
                cur = cur.replace(year=y, month=m)
            except ValueError:
                # day out of range (e.g. 31 of feb) → last day of month
                import calendar as _cal
                last = _cal.monthrange(y, m)[1]
                cur = cur.replace(year=y, month=m, day=last)
        else:
            break
    return results

@api_router.get("/events", response_model=List[EventOut])
async def list_events(
    user: dict = Depends(require_permission("calendario.view")),
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = None,
):
    query: dict = {}
    # Users who can edit calendar see the full team calendar; others see assigned work.
    perms = await get_user_permissions(user)
    can_edit_calendar = "calendario.edit" in perms
    can_edit_events = can_edit_calendar or "events.edit" in perms
    if not can_edit_events:
        query["assigned_user_ids"] = user["id"]
    events = await db.events.find(query, {"_id": 0}).to_list(2000)
    # Expand recurrences
    from_dt = datetime.fromisoformat(from_.replace("Z", "+00:00")) if from_ else datetime.min.replace(tzinfo=timezone.utc)
    if from_dt.tzinfo is None:
        from_dt = from_dt.replace(tzinfo=timezone.utc)
    to_dt = datetime.fromisoformat(to.replace("Z", "+00:00")) if to else datetime.max.replace(tzinfo=timezone.utc)
    if to_dt.tzinfo is None:
        to_dt = to_dt.replace(tzinfo=timezone.utc)
    expanded: List[dict] = []
    for e in events:
        expanded.extend(_expand_recurrence(e, from_dt, to_dt))
    expanded.sort(key=lambda e: e.get("start_at", ""))
    out = []
    for e in expanded:
        e = await _attach_material(e)
        e = await _attach_users(e)
        e = _strip_attachments(e)
        out.append(e)
    return out

async def _sync_material_hours(material_id: str):
    """Recalcula y persiste horas_imputadas en db.materiales desde db.events."""
    if not material_id:
        return
    events = await db.events.find(
        {"material_id": material_id},
        {"hours": 1}
    ).to_list(10000)
    total = round(sum(_safe_float(e.get("hours")) for e in events), 1)
    await db.materiales.update_one(
        {"id": material_id},
        {"$set": {"horas_imputadas": total}},
    )

@api_router.post("/events", response_model=EventOut)
async def create_event(payload: EventCreate, admin: dict = Depends(require_permission("calendario.edit"))):
    if payload.end_at <= payload.start_at:
        raise HTTPException(400, "end_at debe ser posterior a start_at")
    doc = {
        "id": str(uuid.uuid4()),
        "title": payload.title.strip() or "Evento sin título",
        "start_at": payload.start_at,
        "end_at": payload.end_at,
        "description": payload.description,
        "material_id": payload.material_id,
        "assigned_user_ids": payload.assigned_user_ids or [],
        "manager_id": payload.manager_id,
        "recurrence": payload.recurrence.dict() if payload.recurrence else None,
        "hours": payload.hours,
        "status": payload.status or "in_progress",
        "seguimiento": payload.seguimiento or "",
        "attachments": [],
        "created_by": admin["email"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.events.insert_one(doc)
    if doc.get("material_id"):
        await _sync_material_hours(doc["material_id"])
    doc = await _attach_material(doc)
    doc = await _attach_users(doc)
    doc = _strip_attachments(doc)
    return doc

@api_router.patch("/events/{eid}", response_model=EventOut)
async def update_event(eid: str, payload: EventPatch, user: dict = Depends(current_user)):
    # Strip virtual-occurrence suffix if present
    real_id = eid.split(":")[0]
    upd: dict = {}
    # Use exclude_unset so explicit nulls ARE applied (to clear optional fields).
    data = payload.dict(exclude_unset=True)

    # Fetch current event for permission check and notification baseline.
    ev_current = await db.events.find_one({"id": real_id}, {"_id": 0})
    if not ev_current:
        raise HTTPException(404, "Evento no encontrado")

    perms = await get_user_permissions(user)
    can_edit_calendar = "calendario.edit" in perms
    can_edit_events = can_edit_calendar or "events.edit" in perms
    can_assign = can_edit_calendar or "calendario.assign" in perms
    is_assigned = user["id"] in (ev_current.get("assigned_user_ids") or [])
    if not can_edit_events:
        # Users without calendario.edit or events.edit can only modify `status` and `seguimiento`,
        # and only if they are assigned to this event (i.e. the technician in charge).
        if not is_assigned:
            raise HTTPException(403, "No autorizado")
        allowed = {"status", "seguimiento"}
        if any(k not in allowed for k in data.keys()):
            raise HTTPException(403, "Técnicos solo pueden modificar estado y seguimiento")

    for k, v in data.items():
        if k == "recurrence":
            upd["recurrence"] = v  # dict from pydantic or None (clears)
        elif k == "assigned_user_ids":
            if not can_assign:
                raise HTTPException(403, "No tienes permiso para asignar técnicos")
            upd[k] = v or []
        else:
            upd[k] = v
    if not upd:
        raise HTTPException(400, "Nada que actualizar")
    res = await db.events.update_one({"id": real_id}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(404, "Evento no encontrado")

    # Notifications: whenever status changes, create an in-app notification
    # for the event's manager (if set). The seguimiento text is included so
    # the manager sees the technician's observations right in the list.
    new_status = upd.get("status")
    prev_status = ev_current.get("status") or "in_progress"
    if new_status and new_status != prev_status and ev_current.get("manager_id"):
        notif_type = f"event_{new_status}"
        if await should_notify_user(ev_current["manager_id"], notif_type):
            title = ev_current.get("title") or "Evento"
            mid = ev_current.get("material_id")
            project_status_opts = None
            if new_status == "completed":
                notif_title = f"Proyecto terminado: {title}"
                notif_msg = f"{user.get('name') or user.get('email')} marcó el evento como completado."
                if mid:
                    project_status_opts = ["planificado", "a_facturar", "facturado", "terminado", "bloqueado", "anulado"]
                # Generar PDF del presupuesto y adjuntarlo al evento y al proyecto
                budget_id = ev_current.get("budget_id") or upd.get("budget_id")
                if budget_id:
                    budget = await db.budgets.find_one({"id": budget_id})
                    if budget:
                        try:
                            pdf_bytes = build_budget_pdf(budget)
                            pdf_b64 = base64.b64encode(pdf_bytes).decode()
                            att = {
                                "id": str(uuid.uuid4()),
                                "filename": f"presupuesto_{budget.get('n_proyecto','')}.pdf",
                                "mime_type": "application/pdf",
                                "size": len(pdf_bytes),
                                "base64": pdf_b64,
                                "uploaded_at": datetime.now(timezone.utc).isoformat(),
                                "uploaded_by": "sistema",
                            }
                            await db.events.update_one({"id": real_id}, {"$push": {"attachments": att}})
                            notif_msg += f"\n📎 Presupuesto adjunto: {att['filename']}"
                            # Adjuntar también al proyecto vinculado
                            if mid:
                                await db.materiales.update_one(
                                    {"id": mid},
                                    {"$push": {"attachments": att}},
                                )
                                mat = await db.materiales.find_one({"id": mid})
                                if mat:
                                    _fire_and_forget(_sync_project_folder(mat))
                        except Exception as e:
                            logging.getLogger(__name__).error(f"Failed to attach budget PDF: {e}")
                    # Copiar adjuntos de planos vinculados al proyecto
                    if mid:
                        try:
                            linked_plans = await db.plans.find(
                                {"source_event_id": real_id, "source_attachment_id": {"$exists": True, "$ne": None}}
                            ).to_list(length=200)
                            existing_att_ids = {att.get("id") for att in ev_current.get("attachments", [])}
                            for plan in linked_plans:
                                plan_att_id = plan.get("source_attachment_id")
                                if not plan_att_id or plan_att_id not in existing_att_ids:
                                    continue
                                plan_att = next((att for att in ev_current.get("attachments", []) if att.get("id") == plan_att_id), None)
                                if not plan_att:
                                    continue
                                project = await db.materiales.find_one({"id": mid}, {"attachments": 1})
                                if any(att.get("id") == plan_att_id for att in (project.get("attachments") or [])):
                                    continue
                                await db.materiales.update_one(
                                    {"id": mid},
                                    {"$push": {"attachments": plan_att}},
                                )
                                notif_msg += f"\n📎 Plano «{plan.get('title','')}» → {plan_att['filename']}"
                            # Sincronizar carpeta del proyecto
                            mat = await db.materiales.find_one({"id": mid})
                            if mat:
                                _fire_and_forget(_sync_project_folder(mat))
                        except Exception as e:
                            logging.getLogger(__name__).error(f"Failed to copy plan attachments to project: {e}")
            elif new_status == "pending_completion":
                notif_title = f"Pendiente de terminar: {title}"
                seg = upd.get("seguimiento") or ev_current.get("seguimiento") or ""
                notif_msg = (
                    f"{user.get('name') or user.get('email')} dejó el evento pendiente de terminar."
                    + (f"\nObservaciones: {seg}" if seg else "")
                )
                if mid:
                    project_status_opts = ["pendiente", "planificado", "a_facturar", "bloqueado", "anulado"]
            else:
                notif_title = f"Estado actualizado: {title}"
                notif_msg = f"El estado del evento pasó a '{new_status}'."
            notif_doc = {
                "id": str(uuid.uuid4()),
                "user_id": ev_current["manager_id"],
                "event_id": real_id,
                "type": notif_type,
                "title": notif_title,
                "message": notif_msg,
                "read": False,
                "created_at": datetime.utcnow().isoformat(),
                "from_user_id": user["id"],
                "from_user_name": user.get("name") or user.get("email"),
            }
            if mid:
                notif_doc["material_id"] = mid
            if project_status_opts:
                notif_doc["project_status_opts"] = project_status_opts
            await db.notifications.insert_one(notif_doc)

    # If status changed and event has a project linked, cascade to all
    # events that share the same project (material_id).
    if new_status and new_status != prev_status:
        mid = upd.get("material_id") or ev_current.get("material_id")
        if mid:
            await db.events.update_many(
                {"material_id": mid, "id": {"$ne": real_id}},
                {"$set": {"status": new_status}},
            )

    ev = await db.events.find_one({"id": real_id}, {"_id": 0})
    # Sync hours to the linked project(s) if material changed
    mids = set()
    if ev_current.get("material_id"):
        mids.add(ev_current["material_id"])
    if ev and ev.get("material_id"):
        mids.add(ev["material_id"])
    for mid in mids:
        await _sync_material_hours(mid)
    ev = await _attach_material(ev)
    ev = await _attach_users(ev)
    ev = _strip_attachments(ev)
    return ev

@api_router.delete("/events/{eid}")
async def delete_event(eid: str, admin: dict = Depends(require_permission("calendario.edit"))):
    real_id = eid.split(":")[0]
    ev = await db.events.find_one({"id": real_id}, {"material_id": 1})
    mid = ev.get("material_id") if ev else None
    res = await db.events.delete_one({"id": real_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Evento no encontrado")
    if mid:
        await _sync_material_hours(mid)
    return {"ok": True}

# ---------------------------------------------------------------------------
# Guardias (técnicos de guardia por día)
# ---------------------------------------------------------------------------
async def _enrich_guard(g: dict) -> dict:
    """Adjunta datos del usuario a un objeto guardia."""
    uid = g.get("user_id")
    if uid:
        u = await db.users.find_one({"id": uid}, {"_id": 0, "id": 1, "name": 1, "email": 1, "color": 1})
        if u:
            g["user_name"] = u.get("name") or u.get("email") or "Usuario"
            g["user_color"] = u.get("color") or "#3B82F6"
    return g


@api_router.get("/guards")
async def list_guards(
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = None,
    user: dict = Depends(current_user),
):
    """Lista guardias en un rango de fechas (YYYY-MM-DD)."""
    q: dict = {}
    if from_ or to:
        q["date"] = {}
        if from_:
            q["date"]["$gte"] = from_
        if to:
            q["date"]["$lte"] = to
    items = await db.guards.find(q, {"_id": 0}).sort([("date", 1)]).to_list(5000)
    out = []
    for g in items:
        out.append(await _enrich_guard(g))
    return out


@api_router.post("/guards")
async def create_guard(
    payload: GuardCreate,
    user: dict = Depends(require_permission("calendario.edit")),
):
    """Asigna UN técnico de guardia para una fecha. Reemplaza al existente si ya hay alguno."""
    # Validar usuario
    u = await db.users.find_one({"id": payload.user_id}, {"_id": 0})
    if not u:
        raise HTTPException(404, "Usuario no encontrado")
    # Borrar cualquier guardia anterior de ese día (solo permitimos 1 por día)
    await db.guards.delete_many({"date": payload.date})
    new_id = str(uuid.uuid4())
    doc = {
        "id": new_id,
        "date": payload.date,
        "user_id": payload.user_id,
        "note": payload.note,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"],
    }
    await db.guards.insert_one(doc)
    doc.pop("_id", None)
    return await _enrich_guard(doc)


@api_router.delete("/guards/{gid}")
async def delete_guard(
    gid: str,
    user: dict = Depends(require_permission("calendario.edit")),
):
    res = await db.guards.delete_one({"id": gid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Guardia no encontrada")
    return {"ok": True}


@api_router.get("/events/{eid}", response_model=EventOut)
async def get_event(eid: str, user: dict = Depends(current_user)):
    """Fetch a single event by id. Used by the notifications bell to open an
    event that may not be inside the currently-visible calendar range."""
    real_id = eid.split(":")[0]
    ev = await db.events.find_one({"id": real_id}, {"_id": 0})
    if not ev:
        raise HTTPException(404, "Evento no encontrado")
    # Permission: calendario.edit, events.edit OR assigned user OR manager of the event
    perms = await get_user_permissions(user)
    can_edit_calendar = "calendario.edit" in perms or "events.edit" in perms
    allowed_ids = set((ev.get("assigned_user_ids") or []))
    if ev.get("manager_id"):
        allowed_ids.add(ev.get("manager_id"))
    if not can_edit_calendar and user["id"] not in allowed_ids:
        raise HTTPException(403, "No autorizado")
    ev = await _attach_material(ev)
    ev = await _attach_users(ev)
    ev = _strip_attachments(ev)
    return ev

# ---------------- Notifications ----------------
@api_router.get("/notifications")
async def list_notifications(user: dict = Depends(current_user), unread_only: bool = False):
    """Returns notifications addressed to the current user, newest first."""
    q: dict = {"user_id": user["id"]}
    if unread_only:
        q["read"] = False
    cursor = db.notifications.find(q, {"_id": 0}).sort("created_at", -1).limit(100)
    items = await cursor.to_list(length=100)
    unread = await db.notifications.count_documents({"user_id": user["id"], "read": False})
    return {"items": items, "unread": unread}

@api_router.post("/notifications/{nid}/read")
async def mark_notification_read(nid: str, user: dict = Depends(current_user)):
    res = await db.notifications.update_one(
        {"id": nid, "user_id": user["id"]},
        {"$set": {"read": True}},
    )
    if res.matched_count == 0:
        raise HTTPException(404, "Notificación no encontrada")
    return {"ok": True}

@api_router.post("/notifications/read-all")
async def mark_all_read(user: dict = Depends(current_user)):
    await db.notifications.update_many({"user_id": user["id"], "read": False}, {"$set": {"read": True}})
    return {"ok": True}

@api_router.delete("/notifications/{nid}")
async def delete_notification(nid: str, user: dict = Depends(current_user)):
    res = await db.notifications.delete_one({"id": nid, "user_id": user["id"]})
    if res.deleted_count == 0:
        raise HTTPException(404, "Notificación no encontrada")
    return {"ok": True}

@api_router.delete("/notifications")
async def delete_all_notifications(
    user: dict = Depends(current_user),
    only_read: bool = False,
):
    """Bulk-delete notifications for the current user.
    If only_read=true, only read notifications are deleted; otherwise all."""
    q: dict = {"user_id": user["id"]}
    if only_read:
        q["read"] = True
    res = await db.notifications.delete_many(q)
    return {"ok": True, "deleted": res.deleted_count}

class ProjectStatusBody(BaseModel):
    notification_id: str
    project_status: str

class EventGestorBody(BaseModel):
    gestor_list: str  # "general", "archivados", "pendiente_reagendar", "pendiente_revisar"

@api_router.patch("/events/{eid}/gestor-list")
async def set_event_gestor_list(
    eid: str,
    body: EventGestorBody,
    user: dict = Depends(current_user),
):
    """Mueve un evento entre las listas del gestor."""
    valid = {"general", "archivados", "pendiente_reagendar", "pendiente_revisar"}
    if body.gestor_list not in valid:
        raise HTTPException(400, f"Lista no válida. Opciones: {valid}")
    real_id = eid.split(":")[0]
    # Si mueve a Terminados, actualizar también el proyecto vinculado
    if body.gestor_list == "archivados":
        ev = await db.events.find_one({"id": real_id}, {"material_id": 1})
        if ev and ev.get("material_id"):
            await db.materiales.update_one(
                {"id": ev["material_id"]},
                {"$set": {"project_status": "terminado", "sync_status": "pending", "updated_at": datetime.now(timezone.utc).isoformat()}},
            )
    res = await db.events.update_one({"id": real_id}, {"$set": {"gestor_list": body.gestor_list}})
    if res.matched_count == 0:
        raise HTTPException(404, "Evento no encontrado")
    return {"ok": True}

@api_router.post("/notifications/project-status")
async def set_project_status_from_notification(
    body: ProjectStatusBody,
    user: dict = Depends(current_user),
):
    """Cambia el estado del proyecto vinculado a la notificación."""
    notif = await db.notifications.find_one({"id": body.notification_id, "user_id": user["id"]})
    if not notif:
        raise HTTPException(404, "Notificación no encontrada")
    mid = notif.get("material_id")
    if not mid:
        raise HTTPException(400, "La notificación no tiene proyecto vinculado")
    valid_statuses = notif.get("project_status_opts", [])
    if body.project_status not in valid_statuses:
        raise HTTPException(400, f"Estado no permitido. Opciones: {valid_statuses}")
    await db.materiales.update_one(
        {"id": mid},
        {"$set": {"project_status": body.project_status, "sync_status": "pending", "updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    # Marcar notificación como leída
    await db.notifications.update_one({"id": body.notification_id}, {"$set": {"read": True}})
    # Sincronizar carpeta del proyecto
    mat = await db.materiales.find_one({"id": mid}, {"_id": 0})
    if mat:
        _fire_and_forget(_sync_project_folder(mat))
    return {"ok": True, "project_status": body.project_status}

# ---------------- Event attachments ----------------
@api_router.post("/events/{eid}/attachments")
async def upload_event_attachment(eid: str, payload: AttachmentUpload, user: dict = Depends(current_user)):
    real_id = eid.split(":")[0]
    ev = await db.events.find_one({"id": real_id}, {"_id": 0})
    if not ev:
        raise HTTPException(404, "Evento no encontrado")
    # Permission: calendario.edit, events.edit OR assigned user
    perms = await get_user_permissions(user)
    can_edit_calendar = "calendario.edit" in perms or "events.edit" in perms
    if not can_edit_calendar and user["id"] not in (ev.get("assigned_user_ids") or []):
        raise HTTPException(403, "No autorizado")
    # Validate base64 size
    b64 = (payload.base64 or "").split(",")[-1].strip()
    if not b64:
        raise HTTPException(400, "Archivo vacío")
    try:
        raw_size = (len(b64) * 3) // 4
    except Exception:
        raw_size = 0
    if raw_size > MAX_ATTACHMENT_MB * 1024 * 1024:
        raise HTTPException(413, f"El archivo excede {MAX_ATTACHMENT_MB}MB")
    mime = payload.mime_type or ""
    allowed = ("application/pdf", "image/jpeg", "image/jpg", "image/png")
    if not any(mime.lower().startswith(a) for a in allowed):
        raise HTTPException(400, "Tipo no soportado. Solo PDF, JPEG o PNG")
    att = {
        "id": str(uuid.uuid4()),
        "filename": payload.filename[:160] or "archivo",
        "mime_type": mime,
        "size": raw_size,
        "base64": b64,
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "uploaded_by": user["email"],
    }
    await db.events.update_one({"id": real_id}, {"$push": {"attachments": att}})
    # Return metadata only
    return {
        "id": att["id"],
        "filename": att["filename"],
        "mime_type": att["mime_type"],
        "size": att["size"],
        "uploaded_at": att["uploaded_at"],
        "uploaded_by": att["uploaded_by"],
    }

@api_router.get("/events/{eid}/attachments/{aid}")
async def get_event_attachment(eid: str, aid: str, user: dict = Depends(current_user)):
    real_id = eid.split(":")[0]
    ev = await db.events.find_one({"id": real_id}, {"_id": 0})
    if not ev:
        raise HTTPException(404, "Evento no encontrado")
    perms = await get_user_permissions(user)
    can_edit_calendar = "calendario.edit" in perms or "events.edit" in perms
    if not can_edit_calendar and user["id"] not in (ev.get("assigned_user_ids") or []):
        raise HTTPException(403, "No autorizado")
    for a in (ev.get("attachments") or []):
        if a.get("id") == aid:
            return {
                "id": a["id"],
                "filename": a["filename"],
                "mime_type": a["mime_type"],
                "size": a["size"],
                "base64": a["base64"],
                "uploaded_at": a.get("uploaded_at"),
                "uploaded_by": a.get("uploaded_by"),
            }
    raise HTTPException(404, "Adjunto no encontrado")

@api_router.get("/events/{eid}/attachments/{aid}/share-token")
async def get_attachment_share_token(eid: str, aid: str, user: dict = Depends(current_user)):
    """Genera un token JWT para compartir un adjunto públicamente (válido 7 días)."""
    real_id = eid.split(":")[0]
    ev = await db.events.find_one({"id": real_id}, {"_id": 0})
    if not ev:
        raise HTTPException(404, "Evento no encontrado")
    found = any(a.get("id") == aid for a in (ev.get("attachments") or []))
    if not found:
        raise HTTPException(404, "Adjunto no encontrado")
    token = pyjwt.encode(
        {"eid": real_id, "aid": aid, "exp": datetime.now(timezone.utc) + timedelta(days=7)},
        JWT_SECRET, algorithm=JWT_ALGORITHM,
    )
    return {"token": token}

@api_router.delete("/events/{eid}/attachments/{aid}")
async def delete_event_attachment(eid: str, aid: str, user: dict = Depends(current_user)):
    real_id = eid.split(":")[0]
    ev = await db.events.find_one({"id": real_id}, {"_id": 0})
    if not ev:
        raise HTTPException(404, "Evento no encontrado")
    perms = await get_user_permissions(user)
    can_edit_calendar = "calendario.edit" in perms or "events.edit" in perms
    if not can_edit_calendar and user["id"] not in (ev.get("assigned_user_ids") or []):
        raise HTTPException(403, "No autorizado")
    res = await db.events.update_one(
        {"id": real_id},
        {"$pull": {"attachments": {"id": aid}}}
    )
    if res.modified_count == 0:
        raise HTTPException(404, "Adjunto no encontrado")
    return {"ok": True}

@api_router.get("/stamps", response_model=List[StampOut])
async def list_stamps(user: dict = Depends(current_user)):
    customs = await db.stamps.find({}, {"_id": 0}).to_list(500)
    out = [StampOut(**s) for s in BUILTIN_STAMPS]
    for c in customs:
        out.append(StampOut(
            id=c["id"], name=c["name"], is_builtin=False,
            image_base64=c["image_base64"],
        ))
    return out

@api_router.post("/stamps", response_model=StampOut)
async def create_stamp(payload: StampCreate, admin: dict = Depends(require_permission("planos.edit"))):
    doc = {
        "id": str(uuid.uuid4()),
        "name": payload.name.strip() or "Sello personalizado",
        "image_base64": payload.image_base64,
        "is_builtin": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": admin["email"],
    }
    await db.stamps.insert_one(doc)
    return StampOut(
        id=doc["id"], name=doc["name"], is_builtin=False,
        image_base64=doc["image_base64"],
    )

@api_router.delete("/stamps/{sid}")
async def delete_stamp(sid: str, admin: dict = Depends(require_permission("planos.edit"))):
    if sid.startswith("builtin_"):
        raise HTTPException(400, "No puedes eliminar sellos predefinidos")
    res = await db.stamps.delete_one({"id": sid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Sello no encontrado")
    return {"ok": True}

# ---------------- OneDrive routes ----------------
@api_router.get("/auth/onedrive/login")
async def onedrive_login(user: dict = Depends(require_permission("onedrive.manage"))):
    app_msal = _msal_app()
    state = pyjwt.encode(
        {
            "sub": user["id"],
            "purpose": "onedrive_link",
            "exp": datetime.now(timezone.utc) + timedelta(minutes=10),
        },
        JWT_SECRET,
        algorithm=JWT_ALGORITHM,
    )
    auth_url = app_msal.get_authorization_request_url(
        scopes=MS_SCOPES,
        redirect_uri=MS_REDIRECT_URI,
        state=state,
        prompt="select_account",
    )
    return {"auth_url": auth_url}

@api_router.get("/auth/onedrive/callback")
async def onedrive_callback(code: str, state: Optional[str] = None, error: Optional[str] = None, error_description: Optional[str] = None):
    if error:
        return HTMLResponse(f"<h2>Error</h2><p>{error}: {error_description}</p>", status_code=400)
    if not state:
        return HTMLResponse("<h2>Error</h2><p>Falta state de seguridad. Volvé a iniciar la vinculación.</p>", status_code=400)
    try:
        state_payload = pyjwt.decode(state, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if state_payload.get("purpose") != "onedrive_link":
            raise ValueError("invalid purpose")
    except pyjwt.ExpiredSignatureError:
        return HTMLResponse("<h2>Error</h2><p>State expirado. Volvé a iniciar la vinculación de OneDrive.</p>", status_code=400)
    except Exception:
        return HTMLResponse("<h2>Error</h2><p>State inválido. Volvé a iniciar la vinculación de OneDrive.</p>", status_code=400)
    app_msal = _msal_app()
    result = app_msal.acquire_token_by_authorization_code(
        code, scopes=MS_SCOPES, redirect_uri=MS_REDIRECT_URI
    )
    if "error" in result:
        return HTMLResponse(f"<h2>Error</h2><p>{result.get('error_description')}</p>", status_code=400)
    admin_email = result.get("id_token_claims", {}).get("preferred_username") or result.get("id_token_claims", {}).get("email", "unknown")
    access_token = result["access_token"]
    refresh_token = result.get("refresh_token", "")
    if not refresh_token:
        return HTMLResponse("<h2>Error</h2><p>Microsoft no devolvió refresh token. Volvé a vincular OneDrive.</p>", status_code=400)
    try:
        access_token_enc = _encrypt_onedrive_token(access_token)
        refresh_token_enc = _encrypt_onedrive_token(refresh_token)
    except HTTPException as exc:
        return HTMLResponse(f"<h2>Error del servidor</h2><p>{exc.detail}</p>", status_code=exc.status_code)
    await db.onedrive_tokens.update_one(
        {"_id": "admin"},
        {"$set": {
            "_id": "admin",
            "access_token_enc": access_token_enc,
            "refresh_token_enc": refresh_token_enc,
            "admin_email": admin_email,
            "connected_at": datetime.now(timezone.utc).isoformat(),
        },
        "$unset": {"access_token": "", "refresh_token": ""}},
        upsert=True,
    )
    return HTMLResponse(
        f"""<html><body style='font-family:sans-serif;text-align:center;padding:40px;background:#F8FAFC'>
        <h1 style='color:#EA580C'>✓ OneDrive conectado</h1>
        <p>Cuenta: <b>{admin_email}</b></p>
        <p>Ya puedes cerrar esta ventana y volver a la app.</p>
        </body></html>"""
    )

@api_router.get("/auth/onedrive/status", response_model=OneDriveStatus)
async def onedrive_status(user: dict = Depends(current_user)):
    doc = await db.onedrive_tokens.find_one({"_id": "admin"}, {"_id": 0, "access_token": 0, "refresh_token": 0, "access_token_enc": 0, "refresh_token_enc": 0})
    if not doc:
        return OneDriveStatus(connected=False)
    meta = await db.sync_meta.find_one({"_id": "meta"}, {"_id": 0}) or {}
    file_name = None
    if ONEDRIVE_SHARE_URL:
        share_doc = await db.onedrive_share_cache.find_one({"_id": ONEDRIVE_SHARE_URL})
        if share_doc:
            file_name = share_doc.get("name")
    return OneDriveStatus(
        connected=True,
        admin_email=doc.get("admin_email"),
        last_import_at=meta.get("last_import_at"),
        last_push_at=meta.get("last_push_at"),
        file_name=file_name,
    )

@api_router.post("/auth/onedrive/disconnect")
async def onedrive_disconnect(user: dict = Depends(require_permission("onedrive.manage"))):
    await db.onedrive_tokens.delete_one({"_id": "admin"})
    return {"ok": True}

# ---------------- Microsoft user authentication (Entra ID / Azure AD) ----------------
@api_router.get("/auth/microsoft/status")
async def microsoft_status():
    return {"enabled": _microsoft_login_enabled()}

@api_router.get("/auth/microsoft/login")
async def microsoft_login():
    """Return the Microsoft OAuth URL for user login (Entra ID / Azure AD).
    Generates a signed, expirable state JWT to protect the callback."""
    if not _microsoft_login_enabled():
        raise HTTPException(503, "Login Microsoft no configurado en este entorno")
    app_msal = _msal_app()
    state_id = str(uuid.uuid4())
    state_jwt = pyjwt.encode(
        {"jti": state_id, "exp": datetime.now(timezone.utc) + timedelta(minutes=5)},
        JWT_SECRET,
        algorithm=JWT_ALGORITHM,
    )
    auth_url = app_msal.get_authorization_request_url(
        scopes=["User.Read"],
        redirect_uri=MS_AUTH_REDIRECT_URI,
        state=state_jwt,
        prompt="select_account",
    )
    return {"auth_url": auth_url, "state": state_jwt}

@api_router.get("/auth/microsoft/callback", response_class=HTMLResponse)
async def microsoft_callback(code: str, state: Optional[str] = None, error: Optional[str] = None):
    """Handle Microsoft OAuth callback for user login.
    Creates the user on first login (auto-register), then returns HTML that
    posts the JWT back to the frontend (popup message on web, redirect on native)."""
    if not _microsoft_login_enabled():
        return HTMLResponse("<h2>Error de autenticación</h2><p>Login Microsoft no configurado en este entorno.</p>", status_code=503)
    if error:
        return HTMLResponse("<h2>Error</h2><p>Autenticación cancelada</p>", status_code=400)
    if not state:
        return HTMLResponse("<h2>Error de autenticación</h2><p>Falta el parámetro state.</p>", status_code=400)
    try:
        pyjwt.decode(state, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except pyjwt.ExpiredSignatureError:
        return HTMLResponse("<h2>Error de autenticación</h2><p>State expirado. Volvé a iniciar sesión.</p>", status_code=400)
    except Exception:
        return HTMLResponse("<h2>Error de autenticación</h2><p>State inválido.</p>", status_code=400)

    app_msal = _msal_app()
    result = app_msal.acquire_token_by_authorization_code(
        code,
        scopes=["User.Read"],
        redirect_uri=MS_AUTH_REDIRECT_URI,
    )
    if "error" in result:
        return HTMLResponse(
            "<h2>Error de autenticación</h2><p>No se pudo completar el inicio de sesión.</p>",
            status_code=400,
        )

    claims = result.get("id_token_claims", {})
    email = (claims.get("preferred_username") or claims.get("email") or "").lower().strip()
    name = claims.get("name") or ""

    if not email and result.get("access_token"):
        try:
            async with httpx.AsyncClient(timeout=15) as c:
                r = await c.get(
                    f"{GRAPH_BASE}/me",
                    headers={"Authorization": f"Bearer {result['access_token']}"},
                )
                if r.status_code == 200:
                    profile = r.json()
                    email = (profile.get("mail") or profile.get("userPrincipalName") or "").lower().strip()
                    name = profile.get("displayName") or name
        except Exception:
            pass

    if not email:
        return HTMLResponse(
            "<h2>Error de autenticación</h2><p>No se pudo obtener tu email de Microsoft.</p>",
            status_code=400,
        )

    user = await db.users.find_one({"email": email})
    if not user:
        count = await db.users.count_documents({})
        role_key = "admin" if count == 0 else "tecnico"
        role = await db.roles.find_one({"key": role_key})
        existing_colors = [u.get("color") for u in await db.users.find({}, {"_id": 0, "color": 1}).to_list(1000)]
        color = _next_default_color([c for c in existing_colors if c])

        legacy_role = "admin" if role_key == "admin" else "user"
        user = {
            "id": str(uuid.uuid4()),
            "email": email,
            "name": name or email.split("@")[0],
            "password": hash_password(str(uuid.uuid4())),
            "role": legacy_role,
            "role_id": role["id"] if role else None,
            "color": color,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "microsoft_id": claims.get("oid"),
        }
        await db.users.insert_one(user)
        logging.getLogger(__name__).info(f"Auto-registered Microsoft user: {email}")
    elif not user.get("name") and name:
        await db.users.update_one({"id": user["id"]}, {"$set": {"name": name}})

    jwt_token = create_jwt(user)

    _cleanup_expired_codes()
    code = secrets.token_hex(32)
    _auth_codes[code] = {
        "jwt": jwt_token,
        "state": state,
        "expires_at": time.time() + 300,
    }

    return HTMLResponse(content=f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Autenticado — i-SAI</title>
<style>
  * {{ box-sizing:border-box;margin:0;padding:0 }}
  body {{
    font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;
    display:flex;align-items:center;justify-content:center;
    min-height:100vh;background:#F8FAFC;
  }}
  .card {{
    background:#fff;border-radius:16px;padding:40px;text-align:center;
    box-shadow:0 2px 20px rgba(0,0,0,0.08);max-width:400px
  }}
  h2 {{ color:#10B981;margin-bottom:8px;font-size:20px }}
  p {{ color:#475569;margin-bottom:20px;font-size:14px }}
  .spinner {{
    width:24px;height:24px;border:3px solid #E2E8F0;border-top-color:#1976D2;
    border-radius:50%;animation:spin 0.6s linear infinite;margin:0 auto 16px
  }}
  @keyframes spin {{ to {{ transform:rotate(360deg) }} }}
  .fallback {{ font-size:12px;color:#94A3B8;margin-top:20px }}
  .fallback a {{ color:#1976D2;font-weight:600;text-decoration:none }}
</style>
</head>
<body>
<div class="card">
  <h2>✓ Autenticado</h2>
  <p>Redirigiendo a i-SAI como <strong>{email}</strong>...</p>
  <div class="spinner"></div>
  <p class="fallback">
    Si no te redirige automáticamente,<br/>
    <a href="{FRONTEND_URL}/login">volvé a la pantalla de inicio de sesión</a>
  </p>
</div>
<script>
  var code = '{code}';
  var state = '{state}';
  try {{
    if (window.opener && !window.opener.closed) {{
      window.opener.postMessage({{ type:'microsoft_auth', code:code, state:state }}, '{FRONTEND_URL}');
      setTimeout(function(){{ window.close(); }}, 500);
    }} else {{
      window.location.replace('frontend://microsoft-callback?code=' + encodeURIComponent(code) + '&state=' + encodeURIComponent(state));
    }}
  }} catch(e) {{
    window.location.replace('{FRONTEND_URL}/login');
  }}
</script>
</body>
</html>""")


class MicrosoftExchangeRequest(BaseModel):
    code: str
    state: str


class MicrosoftExchangeResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"


@api_router.post("/auth/microsoft/exchange", response_model=MicrosoftExchangeResponse)
async def microsoft_exchange(payload: MicrosoftExchangeRequest):
    """Exchange a one-time code for a JWT access token. Code is single-use."""
    _cleanup_expired_codes()
    record = _auth_codes.get(payload.code)
    if not record:
        raise HTTPException(401, "Código inválido o ya usado")
    if record["state"] != payload.state:
        raise HTTPException(401, "State no coincide")
    if time.time() > record["expires_at"]:
        del _auth_codes[payload.code]
        raise HTTPException(401, "Código expirado")
    jwt_token = record["jwt"]
    del _auth_codes[payload.code]
    return MicrosoftExchangeResponse(access_token=jwt_token, token_type="bearer")


# ---------------- Sync routes (manual override — uses internal helpers) ----------------
@api_router.post("/sync/import-from-onedrive")
async def sync_import(user: dict = Depends(require_permission("onedrive.manage"))):
    n = await _do_import()
    return {"imported": n}

@api_router.post("/sync/push-to-onedrive")
async def sync_push(user: dict = Depends(require_permission("onedrive.manage"))):
    n = await _do_push()
    return {"pushed": n}

# ---------------- Materials routes ----------------
@api_router.get("/materiales", response_model=List[Material])
async def list_materiales(user: dict = Depends(require_permission("proyectos.view")), q: Optional[str] = None, pending_only: bool = False, limit: int = 2000, manager_id: Optional[str] = None, unassigned: bool = False, project_status: Optional[str] = None, year: Optional[str] = None, month: Optional[str] = None):
    # fire-and-forget auto-import if stale
    await maybe_auto_import()
    user_perms = await get_user_permissions(user)
    es_editor_completo = "proyectos.edit" in user_perms
    conditions = []
    if pending_only:
        conditions.append({"sync_status": "pending"})
    if q:
        rx = {"$regex": q, "$options": "i"}
        conditions.append({"$or": [
            {"materiales": rx}, {"cliente": rx}, {"ubicacion": rx},
            {"tecnico": rx}, {"comentarios": rx}, {"comercial": rx}, {"gestor": rx},
        ]})
    if manager_id:
        ids = [i.strip() for i in manager_id.split(",") if i.strip()]
        conditions.append({"manager_id": ids[0] if len(ids) == 1 else {"$in": ids}})
    if unassigned:
        conditions.append({"manager_id": {"$in": [None, ""]}})
    if project_status:
        status_list = [s.strip() for s in project_status.split(",") if s.strip()]
        if len(status_list) == 1:
            st = status_list[0]
            if st == "pendiente":
                conditions.append({"$or": [{"project_status": "pendiente"}, {"project_status": {"$exists": False}}]})
            else:
                conditions.append({"project_status": st})
        elif len(status_list) > 1:
            q_list = [{"project_status": s} for s in status_list]
            if "pendiente" in status_list:
                q_list.append({"project_status": {"$exists": False}})
            conditions.append({"$or": q_list})
    if year and year != "todos":
        conditions.append({"fecha": {"$regex": year}})
    if month and month != "todos":
        # month viene como "01" a "12", coincidir con "-MM" en fecha ISO o "/MM/" en fecha ES
        conditions.append({"$or": [
            {"fecha": {"$regex": f"-{month}"}},
            {"fecha": {"$regex": f"/{month}/"}},
        ]})
    if not es_editor_completo:
        conditions.append({"project_status": {"$ne": "terminado"}})
    query = conditions[0] if len(conditions) == 1 else ({"$and": conditions} if len(conditions) > 1 else {})
    items = await db.materiales.find(query, {"_id": 0}).limit(limit).to_list(limit)
    items.sort(key=lambda x: x.get("row_index", 0))
    # Enrich with manager names and imputed hours from events
    manager_ids = list({m.get("manager_id") for m in items if m.get("manager_id")})
    if manager_ids:
        mgrs = await db.users.find({"id": {"$in": manager_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(500)
        mgr_map = {m["id"]: m.get("name") or m.get("email", "") for m in mgrs}
        for item in items:
            mid = item.get("manager_id")
            if mid and mid in mgr_map:
                item["manager_name"] = mgr_map[mid]
    # Compute imputed hours from linked events
    material_ids = [m["id"] for m in items]
    events = await db.events.find(
        {"material_id": {"$in": material_ids}},
        {"_id": 0, "material_id": 1, "hours": 1},
    ).to_list(10000)
    hours_by_material = {}
    for ev in events:
        mid = ev.get("material_id")
        h = _safe_float(ev.get("hours"))
        hours_by_material[mid] = hours_by_material.get(mid, 0) + h
    for item in items:
        item["horas_imputadas"] = round(hours_by_material.get(item["id"], 0), 1)
    return items

@api_router.get("/materiales/export-excel")
async def materiales_export_excel(user: dict = Depends(current_user)):
    """Export all projects as an Excel file."""
    items = await db.materiales.find({}, {"_id": 0}).sort("row_index", 1).to_list(10000)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proyectos"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    summary_fill = PatternFill(start_color="E6F0FB", end_color="E6F0FB", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # === Summary section ===
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row=row, column=1, value="INFORME DE PROYECTOS")
    cell.font = Font(bold=True, size=16, color="0B2545")
    row += 2

    # Projects by status
    cell = ws.cell(row=row, column=1, value="PROYECTOS POR ESTADO")
    cell.font = Font(bold=True, size=12, color="1976D2")
    row += 1
    for col, h in enumerate(["Estado", "Cantidad", "Horas totales"], 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
        cell.border = thin_border
    row += 1
    statuses = ["pendiente", "planificado", "a_facturar", "facturado", "terminado", "bloqueado", "anulado"]
    status_counts = {}
    for st in statuses:
        mats = [m for m in items if m.get("project_status") == st or (st == "pendiente" and not m.get("project_status"))]
        status_counts[st] = {"count": len(mats), "hours": round(sum(_safe_float(m.get("horas_prev")) for m in mats), 1)}
    for st in statuses:
        sc = status_counts[st]
        for col, v in enumerate([st.replace("_", " ").capitalize(), sc["count"], sc["hours"]], 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border
            cell.fill = summary_fill
        row += 1
    row += 1

    # Projects by manager
    cell = ws.cell(row=row, column=1, value="PROYECTOS POR GESTOR")
    cell.font = Font(bold=True, size=12, color="1976D2")
    row += 1
    for col, h in enumerate(["Gestor", "Total proyectos", "Horas", "Pendientes", "Planificados", "A facturar", "Facturados", "Terminados"], 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
        cell.border = thin_border
    row += 1
    manager_ids = list({m.get("manager_id") for m in items if m.get("manager_id")})
    managers = []
    for mid in manager_ids:
        mats = [m for m in items if m.get("manager_id") == mid]
        if not mats:
            continue
        name = (mats[0].get("manager_name") or mats[0].get("gestor") or "Sin gestor")
        total_h = round(sum(_safe_float(m.get("horas_prev")) for m in mats), 1)
        pend = len([m for m in mats if m.get("project_status") in (None, "pendiente")])
        plan = len([m for m in mats if m.get("project_status") == "planificado"])
        fact = len([m for m in mats if m.get("project_status") == "a_facturar"])
        factu = len([m for m in mats if m.get("project_status") == "facturado"])
        term = len([m for m in mats if m.get("project_status") == "terminado"])
        managers.append((name, len(mats), total_h, pend, plan, fact, factu, term))
    managers.sort(key=lambda x: x[1], reverse=True)
    for mgr in managers[:15]:
        for col, v in enumerate(mgr, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border
            cell.fill = summary_fill
        row += 1

    row += 2

    row += 1
    # Charts
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint

    # Bar chart: projects by status (count + hours)
    chart_sheet = wb.create_sheet("Gráficos")

    # Status data for chart
    chart_sheet.cell(row=1, column=1, value="Estado")
    chart_sheet.cell(row=1, column=2, value="Cantidad")
    chart_sheet.cell(row=1, column=3, value="Horas")
    for i, st in enumerate(statuses):
        sc = status_counts[st]
        chart_sheet.cell(row=i + 2, column=1, value=st.replace("_", " ").capitalize())
        chart_sheet.cell(row=i + 2, column=2, value=sc["count"])
        chart_sheet.cell(row=i + 2, column=3, value=sc["hours"])

    bar = BarChart()
    bar.type = "col"
    bar.title = "Proyectos por estado"
    bar.y_axis.title = "Cantidad"
    bar.x_axis.title = "Estado"
    bar.style = 10
    data_ref = Reference(chart_sheet, min_col=2, min_row=1, max_col=3, max_row=len(statuses) + 1)
    cats_ref = Reference(chart_sheet, min_col=1, min_row=2, max_row=len(statuses) + 1)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.width = 22
    bar.height = 14
    colors = ["F59E0B", "3B82F6", "8B5CF6", "10B981", "6366F1", "EF4444", "6B7280"]
    for idx, color in enumerate(colors):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        bar.series[0].data_points.append(pt)
        pt2 = DataPoint(idx=idx)
        pt2.graphicalProperties.solidFill = color
        bar.series[1].data_points.append(pt2)
    chart_sheet.add_chart(bar, "E1")

    # Manager data for pie chart
    mgr_row = len(statuses) + 4
    chart_sheet.cell(row=mgr_row, column=1, value="Gestor")
    chart_sheet.cell(row=mgr_row, column=2, value="Proyectos")
    for i, mgr in enumerate(managers[:8]):
        chart_sheet.cell(row=mgr_row + 1 + i, column=1, value=mgr[0])
        chart_sheet.cell(row=mgr_row + 1 + i, column=2, value=mgr[1])

    pie = PieChart()
    pie.title = "Proyectos por gestor"
    pie_data = Reference(chart_sheet, min_col=2, min_row=mgr_row, max_row=mgr_row + min(8, len(managers)))
    pie_cats = Reference(chart_sheet, min_col=1, min_row=mgr_row + 1, max_row=mgr_row + min(8, len(managers)))
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    pie.width = 18
    pie.height = 14
    chart_sheet.add_chart(pie, "E18")

    # === Detail section ===
    cell = ws.cell(row=row, column=1, value="DETALLE DE PROYECTOS")
    cell.font = Font(bold=True, size=12, color="1976D2")
    row += 1

    headers = ["Nº Proyecto", "Cliente", "Ubicación", "Horas PREV", "Horas imputadas", "Comercial", "Gestor", "Técnicos", "Fecha", "Entrega/Recogida", "Total/Parcial", "Estado", "Comentarios"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    status_map = {"pendiente": "Pendiente", "a_facturar": "A facturar", "planificado": "Planificado",
                  "facturado": "Facturado", "terminado": "Terminado", "bloqueado": "Bloqueado", "anulado": "Anulado"}

    for row, item in enumerate(items, 2):
        prev_h = _safe_float(item.get("horas_prev"))
        imp_h = item.get("horas_imputadas", 0) or 0
        values = [
            item.get("materiales", ""),
            item.get("cliente", ""),
            item.get("ubicacion", ""),
            prev_h,
            imp_h,
            item.get("comercial", ""),
            item.get("manager_name") or item.get("gestor", ""),
            ", ".join(item.get("tecnicos") or [item.get("tecnico", "")]) if item.get("tecnicos") else (item.get("tecnico") or ""),
            item.get("fecha", ""),
            item.get("entrega_recogida", ""),
            item.get("total_parcial", ""),
            status_map.get(item.get("project_status", ""), item.get("project_status", "Pendiente")),
            item.get("comentarios", ""),
        ]
        red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") if imp_h > prev_h > 0 else None
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if red_fill:
                cell.fill = red_fill

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"proyectos_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@api_router.get("/materiales/{mid}", response_model=Material)
async def get_material(mid: str, user: dict = Depends(require_permission("proyectos.view"))):
    doc = await db.materiales.find_one({"id": mid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Material no encontrado")
    # Imputed hours from events
    perms_get = await get_user_permissions(user)
    es_editor_completo = "proyectos.edit" in perms_get
    if "materiales.view_hours" in perms_get or es_editor_completo:
        events = await db.events.find({"material_id": mid, "hours": {"$exists": True}}, {"hours": 1}).to_list(1000)
        doc["horas_imputadas"] = round(sum(_safe_float(ev.get("hours")) for ev in events), 1)
    # Ocultar project_status a usuarios sin edición completa
    if not es_editor_completo:
        doc.pop("project_status", None)
    return doc

@api_router.patch("/materiales/{mid}", response_model=Material)
async def update_material(mid: str, payload: MaterialUpdate, user: dict = Depends(current_user)):
    user_perms = await get_user_permissions(user)
    es_editor_completo = "proyectos.edit" in user_perms
    es_editor_limitado = "proyectos.editar_campo" in user_perms
    if not es_editor_completo and not es_editor_limitado:
        raise HTTPException(403, "No tienes permiso para editar proyectos (proyectos.edit o proyectos.editar_campo)")
    old = await db.materiales.find_one({"id": mid})
    if not old:
        raise HTTPException(404, "Material no encontrado")
    upd = {k: v for k, v in payload.dict().items() if v is not None}
    if not upd:
        raise HTTPException(400, "Nada que actualizar")
    # Editor limitado solo puede tocar entrega_recogida, total_parcial, comentarios
    CAMPOS_LIMITADOS = {"entrega_recogida", "total_parcial", "comentarios"}
    if not es_editor_completo:
        for campo in list(upd.keys()):
            if campo not in CAMPOS_LIMITADOS and campo != "sync_status":
                raise HTTPException(403, f"No puedes editar el campo '{campo}'. Solo recogida, total/parcial y observaciones.")
    upd["sync_status"] = "pending"
    upd["updated_at"] = datetime.now(timezone.utc).isoformat()
    upd["updated_by"] = user["email"]
    res = await db.materiales.update_one({"id": mid}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(404, "Material no encontrado")
    # Log changes to history
    for field, new_val in upd.items():
        if field in ("sync_status", "updated_at", "updated_by"):
            continue
        old_val = old.get(field)
        if str(old_val) != str(new_val):
            await db.project_history.insert_one({
                "id": str(uuid.uuid4()),
                "project_id": mid,
                "field": field,
                "old_value": str(old_val or ""),
                "new_value": str(new_val or ""),
                "changed_by": user.get("name") or user.get("email"),
                "changed_by_id": user["id"],
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
    doc = await db.materiales.find_one({"id": mid}, {"_id": 0})
    if await _has_onedrive_link():
        schedule_auto_push()
    # Sincronizar carpeta del proyecto
    _fire_and_forget(_sync_project_folder(doc))
    return doc

@api_router.get("/materiales/{mid}/history")
async def get_material_history(mid: str, user: dict = Depends(current_user)):
    items = await db.project_history.find({"project_id": mid}, {"_id": 0}).sort("created_at", -1).limit(50).to_list(50)
    return items

# ---------------- Stats ----------------
@api_router.get("/stats")
async def stats(user: dict = Depends(require_permission("proyectos.view"))):
    total = await db.materiales.count_documents({})
    pending = await db.materiales.count_documents({"sync_status": "pending"})
    return {"total": total, "pending": pending, "synced": total - pending}


@api_router.get("/stats/by-manager")
async def stats_by_manager(
    user: dict = Depends(require_permission("proyectos.view")),
    year: Optional[str] = Query(None, description="Filtrar por año (ej: 2025)"),
):
    """Devuelve estadísticas de proyectos por gestor y estado."""
    user_perms = await get_user_permissions(user)
    es_editor_completo = "proyectos.edit" in user_perms
    statuses = ["pendiente", "planificado", "a_facturar", "facturado", "terminado", "bloqueado", "anulado"]
    STATUS_LABELS = {
        "pendiente": "Pendiente", "planificado": "Planif.", "a_facturar": "Facturar",
        "facturado": "Facturado", "terminado": "Terminado", "bloqueado": "Bloqueado", "anulado": "Anulado",
    }
    STATUS_COLORS = {
        "pendiente": "#F59E0B", "planificado": "#3B82F6", "a_facturar": "#8B5CF6",
        "facturado": "#10B981", "terminado": "#6366F1", "bloqueado": "#EF4444", "anulado": "#6B7280",
    }
    managers = await db.users.find(
        {"$or": [{"role": "admin"}, {"role": "gestor"}]},
        {"_id": 0, "id": 1, "name": 1, "email": 1, "color": 1},
    ).to_list(100)
    result = []
    for mgr in managers:
        query: dict = {"manager_id": mgr["id"]}
        if year and year != "todos":
            query["fecha"] = {"$regex": year}
        if not es_editor_completo:
            query["project_status"] = {"$ne": "terminado"}
        mats = await db.materiales.find(query, {"project_status": 1}).to_list(5000)
        if not mats:
            continue
        by_status: dict = {}
        for st in statuses:
            count = sum(1 for m in mats if (m.get("project_status") or "pendiente") == st)
            if count > 0:
                by_status[st] = {"count": count, "label": STATUS_LABELS[st], "color": STATUS_COLORS[st]}
        result.append({
            "id": mgr["id"],
            "name": mgr.get("name") or mgr.get("email", ""),
            "color": mgr.get("color", "#3B82F6"),
            "total": len(mats),
            "by_status": by_status,
        })
    result.sort(key=lambda x: x["total"], reverse=True)
    return result


@api_router.get("/dashboard")
async def dashboard(user: dict = Depends(current_user)):
    perms = await get_user_permissions(user)
    if "dashboard.view" not in perms:
        raise HTTPException(403, "No tienes permiso para ver el panel de datos")
    # Projects by status — una sola agregación en vez de 7 count_documents
    projects_by_status = {st: 0 for st in ["pendiente", "planificado", "a_facturar", "facturado", "terminado", "bloqueado", "anulado"]}
    async for row in db.materiales.aggregate([
        {"$group": {"_id": "$project_status", "count": {"$sum": 1}}}
    ]):
        st = row["_id"] or "pendiente"
        if st in projects_by_status:
            projects_by_status[st] = row["count"]
    total_hours = 0

    # Hours by manager
    manager_hours = []
    managers = await db.users.find({"role": "admin"}, {"_id": 0, "id": 1, "name": 1, "email": 1, "color": 1}).to_list(50)
    status_order = ["pendiente", "planificado", "a_facturar", "facturado", "terminado"]
    for mgr in managers:
        mats = await db.materiales.find({"manager_id": mgr["id"]}, {"horas_prev": 1, "project_status": 1}).to_list(5000)
        total_h = round(sum(_safe_float(m.get("horas_prev")) for m in mats), 1)
        if total_h <= 0:
            continue
        entry = {
            "name": mgr.get("name") or mgr.get("email", ""),
            "color": mgr.get("color", "#3B82F6"),
            "hours": total_h,
            "count": len(mats),
            "by_status": {},
        }
        for st in status_order:
            st_mats = [m for m in mats if (m.get("project_status") or "pendiente") == st]
            entry["by_status"][st] = round(sum(_safe_float(m.get("horas_prev")) for m in st_mats), 1)
        manager_hours.append(entry)
    manager_hours.sort(key=lambda x: x["hours"], reverse=True)

    # SAT incidents by month (last 6 months)
    sat_by_month = []
    for i in range(5, -1, -1):
        dt = datetime.now(timezone.utc) - timedelta(days=30 * i)
        month_start = dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if i == 0:
            month_end = datetime.now(timezone.utc)
        else:
            if month_start.month == 12:
                month_end = month_start.replace(year=month_start.year + 1, month=1)
            else:
                month_end = month_start.replace(month=month_start.month + 1)
        count = await db.sat_incidents.count_documents({
            "created_at": {"$gte": month_start.isoformat(), "$lt": month_end.isoformat()}
        })
        resolved = await db.sat_incidents.count_documents({
            "status": "resuelta",
            "created_at": {"$gte": month_start.isoformat(), "$lt": month_end.isoformat()},
        })
        mes = month_start.strftime("%b").capitalize()
        sat_by_month.append({"month": mes, "total": count, "resolved": resolved, "year": str(month_start.year), "month_num": f"{month_start.month:02d}"})

    # Projects completed by month (last 6 months)
    projects_by_month = []
    for i in range(5, -1, -1):
        dt = datetime.now(timezone.utc) - timedelta(days=30 * i)
        month_start = dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if i == 0:
            month_end = datetime.now(timezone.utc)
        else:
            if month_start.month == 12:
                month_end = month_start.replace(year=month_start.year + 1, month=1)
            else:
                month_end = month_start.replace(month=month_start.month + 1)
        proj_count = await db.materiales.count_documents({
            "project_status": {"$in": ["terminado", "facturado"]},
            "updated_at": {"$gte": month_start.isoformat(), "$lt": month_end.isoformat()},
        })
        # Total hours for completed projects this month
        completed = await db.materiales.find({
            "project_status": {"$in": ["terminado", "facturado"]},
            "updated_at": {"$gte": month_start.isoformat(), "$lt": month_end.isoformat()},
        }, {"horas_prev": 1, "manager_id": 1}).to_list(5000)
        total_hours_month = sum(_safe_float(c.get("horas_prev")) for c in completed)
        # Desglose por gestor
        by_manager: dict = {}
        for c in completed:
            mgr = c.get("manager_id") or "sin_gestor"
            if mgr not in by_manager:
                by_manager[mgr] = {"count": 0, "hours": 0}
            by_manager[mgr]["count"] += 1
            by_manager[mgr]["hours"] = round(by_manager[mgr]["hours"] + _safe_float(c.get("horas_prev")), 1)
        mes = month_start.strftime("%b").capitalize()
        projects_by_month.append({"month": mes, "count": proj_count, "hours": round(total_hours_month, 1), "year": str(month_start.year), "month_num": f"{month_start.month:02d}", "by_manager": by_manager})

    # Resolver nombres de gestores para el desglose
    all_mgr_ids = set()
    for entry in projects_by_month:
        all_mgr_ids.update(entry["by_manager"].keys())
    all_mgr_ids.discard("sin_gestor")
    mgr_names: dict = {}
    if all_mgr_ids:
        mgrs = await db.users.find({"id": {"$in": list(all_mgr_ids)}}, {"_id": 0, "id": 1, "name": 1, "email": 1}).to_list(100)
        mgr_names = {m["id"]: m.get("name") or m.get("email", m["id"]) for m in mgrs}
    mgr_names["sin_gestor"] = "Sin gestor"
    for entry in projects_by_month:
        entry["by_manager"] = {mgr_names.get(k, k): v for k, v in entry["by_manager"].items()}

    # Total hours for all pending/active projects
    active_projects = await db.materiales.find({"project_status": {"$in": ["pendiente", "planificado", "a_facturar"]}}, {"horas_prev": 1}).to_list(10000)
    total_active_hours = round(sum(_safe_float(m.get("horas_prev")) for m in active_projects), 1)

    # Projects that have exceeded their planned hours (imputadas > previstas)
    all_with_hours = await db.materiales.find(
        {"project_status": {"$nin": ["anulado"]}},
        {"horas_prev": 1, "horas_imputadas": 1, "materiales": 1, "cliente": 1, "id": 1}
    ).to_list(10000)

    # Compute real imputed hours from events (like GET /materiales does)
    mids = [m["id"] for m in all_with_hours]
    events_hours = await db.events.find(
        {"material_id": {"$in": mids}},
        {"_id": 0, "material_id": 1, "hours": 1},
    ).to_list(50000)
    live_hours: dict[str, float] = {}
    for ev in events_hours:
        mid = ev.get("material_id")
        live_hours[mid] = live_hours.get(mid, 0) + _safe_float(ev.get("hours"))

    total_imputadas = 0.0
    total_previstas = 0.0
    over_hours_list = []
    for m in all_with_hours:
        prev = _safe_float(m.get("horas_prev"))
        imp = round(live_hours.get(m["id"], 0), 1)
        total_imputadas += imp
        total_previstas += prev
        if prev > 0 and imp > prev:
            over_hours_list.append({
                "id": m["id"],
                "materiales": m.get("materiales", ""),
                "cliente": m.get("cliente", ""),
                "previstas": prev,
                "imputadas": imp,
                "exceso": round(imp - prev, 1),
            })
    over_hours_list.sort(key=lambda x: x["exceso"], reverse=True)
    projects_over_hours = len(over_hours_list)
    total_over_hours = round(sum(x["exceso"] for x in over_hours_list), 1)
    top_over = over_hours_list[:5]
    total_imputadas = round(total_imputadas, 1)
    total_previstas = round(total_previstas, 1)

    # Today's pending
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    today_end = datetime.now(timezone.utc).replace(hour=23, minute=59, second=59).isoformat()
    today_events = await db.events.count_documents({
        "$or": [
            {"start_at": {"$gte": today_start, "$lte": today_end}},
            {"end_at": {"$gte": today_start, "$lte": today_end}},
        ],
        "status": {"$ne": "completed"},
    })
    pending_sat = await db.sat_incidents.count_documents({"status": "pendiente"})
    pending_budgets = await db.budgets.count_documents({"$or": [{"status": "pendiente"}, {"status": {"$exists": False}}]})

    # ====================== EXTENDED DASHBOARD DATA ======================

    # ---- 1. Week summary ----
    now_utc = datetime.now(timezone.utc)
    week_start = (now_utc - timedelta(days=now_utc.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    week_end = week_start + timedelta(days=7)
    today_evt_count = await db.events.count_documents({
        "start_at": {"$gte": today_start, "$lte": today_end},
    })
    week_events = await db.events.find(
        {"start_at": {"$gte": week_start.isoformat(), "$lt": week_end.isoformat()}},
        {"_id": 0, "hours": 1, "status": 1, "assigned_user_ids": 1, "start_at": 1}
    ).to_list(5000)
    week_events_count = len(week_events)
    week_hours_planned = round(sum(_safe_float(e.get("hours")) for e in week_events), 1)
    week_hours_real = round(sum(_safe_float(e.get("hours")) for e in week_events if e.get("status") == "completed"), 1)
    # Técnicos disponibles vs ocupados (hoy)
    today_busy_users: set = set()
    for e in week_events:
        st = e.get("start_at", "")
        if today_start <= st <= today_end and (e.get("status") != "completed"):
            for uid in (e.get("assigned_user_ids") or []):
                today_busy_users.add(uid)
    all_tech_users = await db.users.find(
        {"role": {"$in": ["user", "tecnico", "technician", "admin", "manager"]}},
        {"id": 1}
    ).to_list(500)
    total_users = len(all_tech_users)
    busy_users = len(today_busy_users)
    week_summary = {
        "events_today": today_evt_count,
        "events_week": week_events_count,
        "hours_planned_week": week_hours_planned,
        "hours_real_week": week_hours_real,
        "technicians_busy": busy_users,
        "technicians_total": total_users,
        "technicians_free": max(0, total_users - busy_users),
    }

    # ---- 1.b Disponibilidad técnicos · Planificación mensual (6 semanas) ----
    # 6 semanas (incluida la actual) = 30 días laborables.
    three_w_start = week_start  # lunes de esta semana
    three_w_end = week_start + timedelta(days=42)  # 6 semanas después
    # Lista de días laborales (lun-vie)
    work_days: list[datetime] = []
    cursor_d = three_w_start
    while cursor_d < three_w_end:
        if cursor_d.weekday() < 5:  # 0=lunes ... 4=viernes
            work_days.append(cursor_d)
        cursor_d += timedelta(days=1)

    # Cargar todos los eventos del rango
    three_w_events = await db.events.find(
        {
            "start_at": {"$gte": three_w_start.isoformat(), "$lt": three_w_end.isoformat()},
        },
        {"_id": 0, "start_at": 1, "end_at": 1, "assigned_user_ids": 1, "status": 1, "manager_id": 1}
    ).to_list(20000)

    # Cargar lista completa de técnicos / usuarios
    all_users_full = await db.users.find(
        {"active": {"$ne": False}},
        {"_id": 0, "id": 1, "name": 1, "email": 1, "role": 1, "color": 1}
    ).to_list(500)
    # Solo consideramos usuarios "operativos" (excluir admin principal si quieres puro)
    operative_users = [u for u in all_users_full if (u.get("role") or "").lower() in ("tecnico", "technician", "user")]

    # Mapa técnico_id -> { day_key: total_hours } con horas acumuladas por día
    busy_hours_by_user: dict[str, dict[str, float]] = {u["id"]: {} for u in operative_users}
    for e in three_w_events:
        if (e.get("status") or "") == "cancelled":
            continue
        st = e.get("start_at")
        et = e.get("end_at")
        if not st or not et:
            continue
        try:
            sd = datetime.fromisoformat(st.replace("Z", "+00:00"))
            ed = datetime.fromisoformat(et.replace("Z", "+00:00"))
        except Exception:
            continue
        hours = max(0.0, (ed - sd).total_seconds() / 3600.0)
        day_key = sd.strftime("%Y-%m-%d")
        for uid in (e.get("assigned_user_ids") or []):
            if uid in busy_hours_by_user:
                busy_hours_by_user[uid][day_key] = busy_hours_by_user[uid].get(day_key, 0.0) + hours

    # Construir respuesta por técnico
    tech_availability: list[dict] = []
    for u in operative_users:
        uid = u["id"]
        hours_map = busy_hours_by_user.get(uid, {})
        days_detail = []
        free_count = 0
        half_count = 0
        for d in work_days:
            key = d.strftime("%Y-%m-%d")
            h = hours_map.get(key, 0.0)
            if h <= 0:
                status = "free"
                free_count += 1
            elif h <= 4:
                status = "half"
                half_count += 1
            else:
                status = "busy"
            days_detail.append({
                "date": key,
                "weekday": d.weekday(),
                "free": status == "free",
                "status": status,
                "hours": round(h, 1),
            })
        tech_availability.append({
            "id": uid,
            "name": u.get("name") or u.get("email", "Sin nombre"),
            "color": u.get("color") or "#3B82F6",
            "free_days": free_count,
            "half_days": half_count,
            "total_days": len(work_days),
            "days": days_detail,
        })
    # Ordenar: técnicos con MÁS días libres primero (los que están "menos ocupados")
    tech_availability.sort(key=lambda x: -(x["free_days"] + x["half_days"] * 0.5))

    # Metadatos del rango
    tech_three_weeks = {
        "from": three_w_start.strftime("%Y-%m-%d"),
        "to": (three_w_end - timedelta(days=1)).strftime("%Y-%m-%d"),
        "total_workdays": len(work_days),
        "weeks_meta": [
            {
                "label": "Esta semana",
                "monday": three_w_start.strftime("%Y-%m-%d"),
            },
            {
                "label": "Próxima",
                "monday": (three_w_start + timedelta(days=7)).strftime("%Y-%m-%d"),
            },
            {
                "label": "+2 sem.",
                "monday": (three_w_start + timedelta(days=14)).strftime("%Y-%m-%d"),
            },
            {
                "label": "+3 sem.",
                "monday": (three_w_start + timedelta(days=21)).strftime("%Y-%m-%d"),
            },
            {
                "label": "+4 sem.",
                "monday": (three_w_start + timedelta(days=28)).strftime("%Y-%m-%d"),
            },
            {
                "label": "+5 sem.",
                "monday": (three_w_start + timedelta(days=35)).strftime("%Y-%m-%d"),
            },
        ],
        "technicians": tech_availability,
    }

    # ---- 2. Top técnicos del mes ----
    month_start = now_utc.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month_events = await db.events.find(
        {"start_at": {"$gte": month_start.isoformat(), "$lt": now_utc.isoformat()}},
        {"_id": 0, "assigned_user_ids": 1, "hours": 1, "status": 1}
    ).to_list(10000)
    tech_stats: dict = {}
    for e in month_events:
        hours = _safe_float(e.get("hours"))
        completed = e.get("status") == "completed"
        for uid in (e.get("assigned_user_ids") or []):
            if uid not in tech_stats:
                tech_stats[uid] = {"hours_planned": 0.0, "hours_real": 0.0, "events": 0, "completed": 0}
            tech_stats[uid]["hours_planned"] += hours
            tech_stats[uid]["events"] += 1
            if completed:
                tech_stats[uid]["hours_real"] += hours
                tech_stats[uid]["completed"] += 1
    user_map = {u["id"]: u for u in await db.users.find({}, {"_id": 0, "id": 1, "name": 1, "email": 1, "color": 1}).to_list(500)}
    top_technicians = []
    for uid, s in tech_stats.items():
        u = user_map.get(uid, {})
        planned = round(s["hours_planned"], 1)
        real = round(s["hours_real"], 1)
        rate = round((real / planned) * 100, 0) if planned > 0 else 0
        top_technicians.append({
            "id": uid,
            "name": u.get("name") or u.get("email") or "Sin nombre",
            "color": u.get("color") or "#3B82F6",
            "hours_planned": planned,
            "hours_real": real,
            "events": s["events"],
            "completed": s["completed"],
            "completion_rate": rate,
        })
    top_technicians.sort(key=lambda x: x["hours_planned"], reverse=True)
    top_technicians = top_technicians[:8]

    # ---- 3. Alerts críticas ----
    sat_urgent_open = await db.sat_incidents.count_documents({
        "status": {"$in": ["abierta", "en_proceso", "pendiente", None]},
        "prioridad": {"$in": ["alta", "urgente"]},
    })
    events_no_tech = await db.events.count_documents({
        "$or": [
            {"assigned_user_ids": {"$exists": False}},
            {"assigned_user_ids": {"$size": 0}},
            {"assigned_user_ids": None},
        ],
        "status": {"$ne": "completed"},
        "start_at": {"$gte": now_utc.isoformat()},
    })
    thirty_days_ago = (now_utc - timedelta(days=30)).isoformat()
    budgets_old_pending = await db.budgets.count_documents({
        "status": {"$in": ["pendiente", "enviado", "borrador", None]},
        "$or": [
            {"created_at": {"$lt": thirty_days_ago}},
            {"fecha": {"$lt": thirty_days_ago[:10]}},
        ],
    })
    alerts = {
        "sat_urgent_open": sat_urgent_open,
        "events_no_tech": events_no_tech,
        "budgets_pending_30d": budgets_old_pending,
        "total": sat_urgent_open + events_no_tech + budgets_old_pending,
    }

    # ---- 4. Budget pipeline ----
    pipeline_stages = ["borrador", "enviado", "aceptado", "rechazado"]
    pipeline = {}
    for stage in pipeline_stages:
        # Acepta tanto "status" como "estado"
        count = await db.budgets.count_documents({
            "$or": [{"status": stage}, {"estado": stage}]
        })
        # Sumar importes (campo "total" o "importe_total")
        cur = db.budgets.aggregate([
            {"$match": {"$or": [{"status": stage}, {"estado": stage}]}},
            {"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$total", {"$ifNull": ["$importe_total", 0]}]}}}}
        ])
        amt = 0.0
        async for row in cur:
            amt = row.get("total", 0) or 0
        pipeline[stage] = {"count": count, "amount": round(float(amt), 2)}
    # Conversion rate = aceptados / (aceptados + rechazados)
    accepted = pipeline.get("aceptado", {}).get("count", 0)
    rejected = pipeline.get("rechazado", {}).get("count", 0)
    conv_rate = round((accepted / (accepted + rejected)) * 100, 1) if (accepted + rejected) > 0 else 0
    budget_pipeline = {
        "stages": pipeline,
        "conversion_rate": conv_rate,
        "total_count": sum(p["count"] for p in pipeline.values()),
        "total_amount": round(sum(p["amount"] for p in pipeline.values()), 2),
    }

    # ---- 5. SAT health ----
    # Tiempo medio de resolución (horas)
    resolved_incidents = await db.sat_incidents.find(
        {"status": {"$in": ["resuelta", "cerrada"]}, "resolved_at": {"$exists": True, "$ne": None}, "created_at": {"$exists": True}},
        {"_id": 0, "created_at": 1, "resolved_at": 1, "history": 1, "cliente": 1, "client_name": 1, "lat": 1, "lng": 1, "direccion": 1}
    ).to_list(5000)
    resolution_hours = []
    first_visit_solved = 0
    by_client: dict = {}
    heatmap_points = []
    for inc in resolved_incidents:
        try:
            ca = datetime.fromisoformat(inc["created_at"].replace("Z", "+00:00"))
            ra = datetime.fromisoformat(inc["resolved_at"].replace("Z", "+00:00"))
            delta_h = (ra - ca).total_seconds() / 3600.0
            if delta_h >= 0:
                resolution_hours.append(delta_h)
        except Exception:
            pass
        # First visit: history length <= 1 visita
        if len(inc.get("history") or []) <= 1:
            first_visit_solved += 1
        # Por cliente
        name = inc.get("cliente") or inc.get("client_name") or "Sin cliente"
        by_client[name] = by_client.get(name, 0) + 1
        # Heatmap
        if inc.get("lat") and inc.get("lng"):
            heatmap_points.append({"lat": inc["lat"], "lng": inc["lng"]})
    # Añadir TODAS las incidencias al heatmap (no solo las resueltas)
    all_incidents = await db.sat_incidents.find(
        {"lat": {"$exists": True, "$ne": None}},
        {"_id": 0, "lat": 1, "lng": 1, "prioridad": 1, "status": 1, "cliente": 1}
    ).to_list(5000)
    heatmap_full = [{"lat": i["lat"], "lng": i["lng"], "priority": i.get("prioridad") or "media"} for i in all_incidents]

    avg_resolution = round(sum(resolution_hours) / len(resolution_hours), 1) if resolution_hours else 0
    first_visit_rate = round((first_visit_solved / len(resolved_incidents)) * 100, 1) if resolved_incidents else 0
    top_clients = sorted([{"name": k, "count": v} for k, v in by_client.items()], key=lambda x: x["count"], reverse=True)[:5]
    sat_health = {
        "avg_resolution_hours": avg_resolution,
        "first_visit_rate": first_visit_rate,
        "top_clients": top_clients,
        "total_resolved": len(resolved_incidents),
        "heatmap": heatmap_full,
    }

    # ---- 6. Active projects for mini map ----
    active_projects_map = await db.materiales.find(
        {
            "lat": {"$exists": True, "$ne": None},
            "lng": {"$exists": True, "$ne": None},
            "project_status": {"$in": ["pendiente", "planificado", "a_facturar"]},
        },
        {"_id": 0, "id": 1, "lat": 1, "lng": 1, "cliente": 1, "ubicacion": 1, "project_status": 1}
    ).to_list(2000)

    # ---- 7. YoY comparison ----
    current_year = now_utc.year
    last_year = current_year - 1
    def year_range(y: int) -> tuple[str, str]:
        return f"{y}-01-01T00:00:00+00:00", f"{y+1}-01-01T00:00:00+00:00"
    cy_start, cy_end = year_range(current_year)
    ly_start, ly_end = year_range(last_year)
    closed_query = {"project_status": {"$in": ["terminado", "facturado"]}}
    this_year_closed = await db.materiales.count_documents({**closed_query, "updated_at": {"$gte": cy_start, "$lt": cy_end}})
    last_year_closed = await db.materiales.count_documents({**closed_query, "updated_at": {"$gte": ly_start, "$lt": ly_end}})
    growth_pct = round(((this_year_closed - last_year_closed) / last_year_closed) * 100, 1) if last_year_closed > 0 else (100.0 if this_year_closed > 0 else 0)
    # Quarters
    quarters_this = []
    quarters_last = []
    for q in range(4):
        q_start_month = q * 3 + 1
        q_end_month = q_start_month + 3
        def q_range(y: int, m_start: int, m_end: int) -> tuple[str, str]:
            s = f"{y}-{m_start:02d}-01T00:00:00+00:00"
            if m_end > 12:
                e = f"{y+1}-01-01T00:00:00+00:00"
            else:
                e = f"{y}-{m_end:02d}-01T00:00:00+00:00"
            return s, e
        s1, e1 = q_range(current_year, q_start_month, q_end_month)
        s2, e2 = q_range(last_year, q_start_month, q_end_month)
        quarters_this.append(await db.materiales.count_documents({**closed_query, "updated_at": {"$gte": s1, "$lt": e1}}))
        quarters_last.append(await db.materiales.count_documents({**closed_query, "updated_at": {"$gte": s2, "$lt": e2}}))
    yoy_comparison = {
        "this_year": current_year,
        "last_year": last_year,
        "closed_this_year": this_year_closed,
        "closed_last_year": last_year_closed,
        "growth_pct": growth_pct,
        "quarters_this": quarters_this,
        "quarters_last": quarters_last,
    }

    # ---- 8. Geo distribution ----
    pipe_cities = [
        {"$match": {"ubicacion": {"$exists": True, "$ne": None, "$ne": ""}, "project_status": {"$nin": ["anulado", "facturado", "terminado"]}}},
        {"$group": {"_id": "$ubicacion", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 8},
    ]
    top_cities = []
    async for row in db.materiales.aggregate(pipe_cities):
        top_cities.append({"city": row["_id"], "count": row["count"]})
    # Facturación por provincia (aproximación: mapeo de ciudades -> provincia)
    province_map = {
        "BILBAO": "Bizkaia", "BARAKALDO": "Bizkaia", "GETXO": "Bizkaia", "PORTUGALETE": "Bizkaia", "SESTAO": "Bizkaia",
        "BASAURI": "Bizkaia", "DURANGO": "Bizkaia", "AMOREBIETA": "Bizkaia", "MUNGIA": "Bizkaia", "ERMUA": "Bizkaia",
        "GERNIKA": "Bizkaia", "GALDAKAO": "Bizkaia", "BERMEO": "Bizkaia", "LEKEITIO": "Bizkaia",
        "DONOSTIA": "Gipuzkoa", "IRUN": "Gipuzkoa", "ERRENTERIA": "Gipuzkoa", "EIBAR": "Gipuzkoa", "ARRASATE": "Gipuzkoa",
        "BERGARA": "Gipuzkoa", "HONDARRIBIA": "Gipuzkoa", "ZARAUTZ": "Gipuzkoa", "AZPEITIA": "Gipuzkoa",
        "BEASAIN": "Gipuzkoa", "ORDIZIA": "Gipuzkoa", "ELGOIBAR": "Gipuzkoa", "ZUMAIA": "Gipuzkoa",
        "TOLOSA": "Gipuzkoa", "LASARTE": "Gipuzkoa", "ASTIGARRAGA": "Gipuzkoa", "ONATI": "Gipuzkoa",
        "VITORIA": "Araba", "GASTEIZ": "Araba", "AMURRIO": "Araba", "LAUDIO": "Araba", "LLODIO": "Araba",
        "PAMPLONA": "Navarra", "IRUNEA": "Navarra", "TUDELA": "Navarra", "TAFALLA": "Navarra", "ESTELLA": "Navarra",
        "LOGRONO": "La Rioja", "CALAHORRA": "La Rioja", "HARO": "La Rioja",
        "SANTANDER": "Cantabria", "TORRELAVEGA": "Cantabria",
        "MADRID": "Madrid", "BARCELONA": "Barcelona", "VALENCIA": "Valencia",
        "SEVILLA": "Sevilla", "ZARAGOZA": "Zaragoza", "BURGOS": "Burgos",
    }
    def to_province(ubic: str) -> str:
        if not ubic:
            return "Otros"
        u = ubic.upper().strip()
        for key, prov in province_map.items():
            if key in u:
                return prov
        return "Otros"
    cur = db.materiales.find({"project_status": {"$nin": ["anulado"]}}, {"_id": 0, "ubicacion": 1, "total_parcial": 1})
    by_province: dict = {}
    async for m in cur:
        prov = to_province(m.get("ubicacion") or "")
        # parsear total_parcial "1.234,56 €" -> 1234.56
        raw = (m.get("total_parcial") or "").replace("€", "").replace(" ", "").replace(".", "").replace(",", ".").strip()
        try:
            val = float(raw) if raw else 0.0
        except Exception:
            val = 0.0
        by_province[prov] = by_province.get(prov, 0) + val
    total_facturacion = sum(by_province.values()) or 1
    by_province_list = sorted(
        [{"province": p, "amount": round(v, 2), "pct": round((v / total_facturacion) * 100, 1)} for p, v in by_province.items()],
        key=lambda x: x["amount"], reverse=True
    )[:8]
    geo_distribution = {
        "top_cities": top_cities,
        "by_province": by_province_list,
        "total_amount": round(total_facturacion, 2),
    }

    return {
        "projects_by_status": projects_by_status,
        "manager_hours": manager_hours[:8],
        "sat_by_month": sat_by_month,
        "projects_by_month": projects_by_month,
        "total_active_hours": total_active_hours,
        "projects_over_hours": projects_over_hours,
        "total_over_hours": total_over_hours,
        "top_over_hours": top_over,
        "total_imputadas_hours": total_imputadas,
        "total_previstas_hours": total_previstas,
        "today": {
            "events": today_events,
            "pending_sat": pending_sat,
            "pending_budgets": pending_budgets,
        },
        # New extended fields
        "week_summary": week_summary,
        "tech_three_weeks": tech_three_weeks,
        "top_technicians": top_technicians,
        "alerts": alerts,
        "budget_pipeline": budget_pipeline,
        "sat_health": sat_health,
        "active_projects_map": active_projects_map,
        "yoy_comparison": yoy_comparison,
        "geo_distribution": geo_distribution,
    }

# ---------------- Budgets (Presupuestos) ----------------
async def current_budget_view(user: dict = Depends(current_user)):
    perms = await get_user_permissions(user)
    if "presupuestos.view" not in perms:
        raise HTTPException(403, "Requiere permiso de Presupuestos")
    return user

async def current_budget_edit(user: dict = Depends(current_user)):
    perms = await get_user_permissions(user)
    if "presupuestos.edit" not in perms:
        raise HTTPException(403, "Requiere permiso para editar Presupuestos")
    return user

async def current_budget_export(user: dict = Depends(current_user)):
    perms = await get_user_permissions(user)
    if "presupuestos.export" not in perms and "presupuestos.edit" not in perms:
        raise HTTPException(403, "Requiere permiso para exportar Presupuestos")
    return user

class EquipmentRow(BaseModel):
    elemento: str = ""
    cantidad: Optional[str] = ""
    ubicacion: Optional[str] = ""
    observaciones: Optional[str] = ""

def _serialize_equipos(equipos: list) -> list:
    return [e if isinstance(e, dict) else (e.dict() if hasattr(e, "dict") else e) for e in equipos]

def _validate_date(value: Optional[str]) -> Optional[str]:
    if value is None or value.strip() == "":
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value.strip(), fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    raise HTTPException(422, f"Formato de fecha inválido: {value}. Usá DD/MM/AAAA")

class BudgetCreate(BaseModel):
    # general
    n_proyecto: Optional[str] = ""
    cliente: Optional[str] = ""
    nombre_instalacion: Optional[str] = ""
    direccion: Optional[str] = ""
    contacto_1: Optional[str] = ""
    contacto_2: Optional[str] = ""
    observaciones_presupuesto: Optional[str] = ""
    fecha_inicio: Optional[str] = ""
    fecha_fin: Optional[str] = ""
    observaciones_ejecucion: Optional[str] = ""
    # equipos
    equipos: List[EquipmentRow] = []
    # deliveries
    entrega_tarjeta_mantenimiento: bool = False
    entrega_llave_salto: bool = False
    entrega_eps100: bool = False
    # signatures (base64 PNG)
    firma_isai: Optional[str] = ""
    nombre_isai: Optional[str] = ""
    cargo_isai: Optional[str] = ""
    firma_cliente: Optional[str] = ""
    nombre_cliente: Optional[str] = ""
    cargo_cliente: Optional[str] = ""
    # link to existing material (proyecto)
    material_id: Optional[str] = None

class BudgetPatch(BudgetCreate):
    pass

class BudgetStatusUpdate(BaseModel):
    status: str

class BudgetTemplateCreate(BaseModel):
    name: str
    equipos: List[dict] = []

@api_router.post("/budgets")
async def create_budget(payload: BudgetCreate, user: dict = Depends(current_budget_edit)):
    now = datetime.now(timezone.utc).isoformat()
    doc = payload.dict(exclude_none=True)
    doc["equipos"] = _serialize_equipos(payload.equipos or [])
    doc["fecha_inicio"] = _validate_date(doc.get("fecha_inicio"))
    doc["fecha_fin"] = _validate_date(doc.get("fecha_fin"))
    if doc["fecha_inicio"] and doc["fecha_fin"]:
        try:
            fi = datetime.strptime(doc["fecha_inicio"], "%d/%m/%Y")
            ff = datetime.strptime(doc["fecha_fin"], "%d/%m/%Y")
            if ff < fi:
                raise HTTPException(422, "La fecha de fin no puede ser anterior a la fecha de inicio")
        except HTTPException:
            raise
        except Exception:
            pass
    doc.update({
        "id": str(uuid.uuid4()),
        "created_at": now,
        "updated_at": now,
        "created_by": user["email"],
        "created_by_name": user.get("name") or user["email"],
        "status": "pendiente",
    })
    await db.budgets.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/budgets")
async def list_budgets(skip: int = 0, limit: int = 100, user: dict = Depends(current_budget_view)):
    items = await db.budgets.find({}, {"_id": 0, "firma_isai": 0, "firma_cliente": 0, "attachments.data": 0}).sort("updated_at", -1).skip(skip).limit(limit).to_list(limit)
    return items

@api_router.get("/budgets/accepted")
async def list_accepted_budgets(user: dict = Depends(current_budget_view)):
    items = await db.budgets.find(
        {"status": "aceptado"},
        {"_id": 0, "firma_isai": 0, "firma_cliente": 0, "attachments.data": 0},
    ).sort("updated_at", -1).to_list(200)
    return items

@api_router.get("/budgets/{bid}")
async def get_budget(bid: str, user: dict = Depends(current_budget_view)):
    b = await db.budgets.find_one({"id": bid}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Presupuesto no encontrado")
    if b.get("attachments"):
        for a in b["attachments"]:
            a.pop("data", None)
    return b

@api_router.patch("/budgets/{bid}/status")
async def update_budget_status(bid: str, payload: BudgetStatusUpdate, user: dict = Depends(current_budget_edit)):
    if payload.status not in ("pendiente", "en_revision", "aceptado", "rechazado", "facturado"):
        raise HTTPException(422, "Estado inválido")
    b = await db.budgets.find_one({"id": bid})
    if not b:
        raise HTTPException(404, "Presupuesto no encontrado")
    await db.budgets.update_one({"id": bid}, {"$set": {
        "status": payload.status,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": user["email"],
        "updated_by_name": user.get("name") or user["email"],
    }})
    return {"ok": True, "status": payload.status}

@api_router.patch("/budgets/{bid}")
async def update_budget(bid: str, payload: BudgetPatch, user: dict = Depends(current_budget_edit)):
    upd = payload.dict(exclude_unset=True)
    if "equipos" in upd:
        upd["equipos"] = _serialize_equipos(upd["equipos"])
    if "fecha_inicio" in upd:
        upd["fecha_inicio"] = _validate_date(upd["fecha_inicio"])
    if "fecha_fin" in upd:
        upd["fecha_fin"] = _validate_date(upd["fecha_fin"])
    upd["updated_at"] = datetime.now(timezone.utc).isoformat()
    upd["updated_by"] = user["email"]
    upd["updated_by_name"] = user.get("name") or user["email"]
    old = await db.budgets.find_one({"id": bid})
    fi = upd.get("fecha_inicio", old.get("fecha_inicio") if old else None)
    ff = upd.get("fecha_fin", old.get("fecha_fin") if old else None)
    if fi and ff:
        try:
            if datetime.strptime(ff, "%d/%m/%Y") < datetime.strptime(fi, "%d/%m/%Y"):
                raise HTTPException(422, "La fecha de fin no puede ser anterior a la fecha de inicio")
        except HTTPException:
            raise
        except Exception:
            pass
    if old:
        version_doc = {k: v for k, v in old.items() if k != "_id"}
        version_doc["budget_id"] = bid
        version_doc["version_id"] = str(uuid.uuid4())
        version_doc["saved_at"] = datetime.now(timezone.utc).isoformat()
        version_doc["saved_by"] = user["email"]
        version_doc["saved_by_name"] = user.get("name") or user["email"]
        await db.budget_versions.insert_one(version_doc)
    res = await db.budgets.update_one({"id": bid}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(404, "Presupuesto no encontrado")
    b = await db.budgets.find_one({"id": bid}, {"_id": 0})
    return b

@api_router.delete("/budgets/{bid}")
async def delete_budget(bid: str, user: dict = Depends(current_budget_edit)):
    res = await db.budgets.delete_one({"id": bid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Presupuesto no encontrado")
    return {"ok": True}

@api_router.get("/budgets/{bid}/versions")
async def list_budget_versions(bid: str, user: dict = Depends(current_budget_view)):
    items = await db.budget_versions.find(
        {"budget_id": bid},
        {"_id": 0, "firma_isai": 0, "firma_cliente": 0, "attachments": 0},
    ).sort("saved_at", -1).to_list(50)
    return items

@api_router.get("/budgets/{bid}/versions/{vid}")
async def get_budget_version(bid: str, vid: str, user: dict = Depends(current_budget_view)):
    v = await db.budget_versions.find_one({"budget_id": bid, "version_id": vid}, {"_id": 0})
    if not v:
        raise HTTPException(404, "Versión no encontrada")
    return v

@api_router.post("/budget-templates")
async def create_budget_template(payload: BudgetTemplateCreate, user: dict = Depends(current_budget_view)):
    doc = {
        "id": str(uuid.uuid4()),
        "name": payload.name,
        "equipos": payload.equipos,
        "created_by": user["email"],
        "created_by_name": user.get("name") or user["email"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.budget_templates.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/budget-templates")
async def list_budget_templates(user: dict = Depends(current_budget_view)):
    items = await db.budget_templates.find({}, {"_id": 0}).sort("created_at", -1).to_list(50)
    return items

@api_router.delete("/budget-templates/{tid}")
async def delete_budget_template(tid: str, user: dict = Depends(current_budget_view)):
    res = await db.budget_templates.delete_one({"id": tid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Plantilla no encontrada")
    return {"ok": True}

@api_router.get("/budgets/stats")
async def budgets_stats(user: dict = Depends(current_budget_view)):
    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()

    all_budgets = await db.budgets.find({}, {"status": 1, "created_by": 1, "created_by_name": 1, "updated_at": 1}).to_list(2000)

    by_status = {"pendiente": 0, "en_revision": 0, "aceptado": 0, "rechazado": 0, "facturado": 0}
    by_commercial: dict = {}
    accepted_this_month = 0

    for b in all_budgets:
        st = b.get("status", "pendiente")
        if st in by_status:
            by_status[st] += 1

        email = b.get("created_by", "desconocido")
        name = b.get("created_by_name", email)
        if email not in by_commercial:
            by_commercial[email] = {"name": name, "pendiente": 0, "en_revision": 0, "aceptado": 0, "rechazado": 0, "facturado": 0, "total": 0}
        by_commercial[email][st] = by_commercial[email].get(st, 0) + 1
        by_commercial[email]["total"] += 1

        if st == "aceptado" and b.get("updated_at", "") >= month_start:
            accepted_this_month += 1

    return {
        "by_status": by_status,
        "by_commercial": list(by_commercial.values()),
        "accepted_this_month": accepted_this_month,
        "total": len(all_budgets),
    }

class BudgetAttachmentUpload(BaseModel):
    name: str
    data: str  # base64
    mime: str = "application/octet-stream"

@api_router.post("/budgets/{bid}/duplicate")
async def duplicate_budget(bid: str, user: dict = Depends(require_permission("presupuestos.edit"))):
    original = await db.budgets.find_one({"id": bid}, {"_id": 0})
    if not original:
        raise HTTPException(404, "Presupuesto no encontrado")
    now = datetime.now(timezone.utc).isoformat()
    new_doc = {k: v for k, v in original.items()
               if k not in ("id", "status", "created_at", "updated_at", "created_by", "created_by_name",
                            "firma_isai", "firma_cliente")}
    new_doc.update({
        "id": str(uuid.uuid4()),
        "status": "pendiente",
        "created_at": now,
        "updated_at": now,
        "created_by": user["email"],
        "created_by_name": user.get("name") or user["email"],
        "firma_isai": "",
        "firma_cliente": "",
    })
    await db.budgets.insert_one(new_doc)
    return new_doc

@api_router.post("/budgets/{bid}/attachments")
async def upload_budget_attachment(bid: str, payload: BudgetAttachmentUpload, user: dict = Depends(require_permission("presupuestos.edit"))):
    b = await db.budgets.find_one({"id": bid}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Presupuesto no encontrado")
    try:
        raw_bytes = base64.b64decode(payload.data)
    except Exception:
        raise HTTPException(400, "Base64 inválido")
    if len(raw_bytes) > 10 * 1024 * 1024:
        raise HTTPException(413, "El archivo excede 10MB")
    now = datetime.now(timezone.utc).isoformat()
    att = {
        "id": str(uuid.uuid4()),
        "name": payload.name[:200],
        "mime": payload.mime,
        "data": payload.data,
        "size": len(raw_bytes),
        "created_at": now,
    }
    await db.budgets.update_one({"id": bid}, {"$push": {"attachments": att}})
    return {
        "ok": True,
        "attachment": {
            "id": att["id"],
            "name": att["name"],
            "mime": att["mime"],
            "size": att["size"],
            "created_at": att["created_at"],
        },
    }

@api_router.get("/budgets/{bid}/attachments/{aid}")
async def get_budget_attachment(bid: str, aid: str, user: dict = Depends(require_permission("presupuestos.view"))):
    b = await db.budgets.find_one({"id": bid}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Presupuesto no encontrado")
    for a in (b.get("attachments") or []):
        if a.get("id") == aid:
            return {"id": a["id"], "name": a["name"], "mime": a["mime"], "data": a["data"], "size": a["size"]}
    raise HTTPException(404, "Adjunto no encontrado")

@api_router.delete("/budgets/{bid}/attachments/{aid}")
async def delete_budget_attachment(bid: str, aid: str, user: dict = Depends(require_permission("presupuestos.edit"))):
    res = await db.budgets.update_one({"id": bid}, {"$pull": {"attachments": {"id": aid}}})
    if res.modified_count == 0:
        raise HTTPException(404, "Adjunto no encontrado")
    return {"ok": True}

DEFAULT_EQUIPMENT_LIST = [
    "Cilindro electrónico Salto XS4",
    "Lector de muro Salto XS4",
    "Cerradura electrónica Salto XS4",
    "Escudo electrónico Salto XS4",
    "Mini cilindro Salto",
    "Tarjeta MIFARE 1K",
    "Llavero MIFARE",
    "Portal encodificador HAMS",
    "PPD (Portable Programming Device)",
    "Controlador Salto CU5000",
    "Fuente de alimentación 12V",
    "Cableado estructurado",
]

@api_router.get("/budgets-defaults/equipos")
async def budgets_default_equipos(user: dict = Depends(current_budget_view)):
    return {"items": DEFAULT_EQUIPMENT_LIST}


@api_router.get("/budgets/{bid}/pdf")
async def get_budget_pdf(bid: str, user: dict = Depends(current_budget_export)):
    """
    Genera el PDF 'Hoja de instalación' rellenado con los datos del presupuesto,
    manteniendo el layout exacto del template y los campos editables (AcroForm).
    """
    b = await db.budgets.find_one({"id": bid}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Presupuesto no encontrado")
    try:
        pdf_bytes = build_budget_pdf(b)
    except Exception as e:
        logging.exception("PDF generation failed")
        raise HTTPException(500, f"No se pudo generar el PDF: {e}")
    filename = f"hoja_instalacion_{(b.get('n_proyecto') or bid)[:40]}.pdf"
    safe = "".join(c for c in filename if c.isalnum() or c in "._-") or "hoja_instalacion.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="{safe}"',
            "Cache-Control": "no-store",
        },
    )


@api_router.post("/budgets/pdf-preview")
async def post_budget_pdf_preview(payload: BudgetCreate,
                                  user: dict = Depends(current_budget_edit)):
    """
    Genera el PDF sin guardar el presupuesto. Útil para previsualizar antes de
    guardar. Recibe en el body los mismos campos que BudgetCreate.
    """
    data = payload.dict()
    data["equipos"] = _serialize_equipos(payload.equipos or [])
    try:
        pdf_bytes = build_budget_pdf(data)
    except Exception as e:
        logging.exception("PDF preview failed")
        raise HTTPException(500, f"No se pudo generar el PDF: {e}")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": 'inline; filename="hoja_instalacion_preview.pdf"',
            "Cache-Control": "no-store",
        },
    )


# ---------------- Utilities: image -> PDF ----------------
class ImageToPdfBody(BaseModel):
    base64: str
    mime_type: str = "image/jpeg"   # image/jpeg or image/png
    filename: Optional[str] = "plano.pdf"
    orientation: Optional[str] = None   # "portrait" o "landscape" para forzar A3


@api_router.post("/utils/image-to-pdf")
async def utils_image_to_pdf(body: ImageToPdfBody, user: dict = Depends(current_user)):
    """
    Convierte una imagen base64 (JPEG/PNG) a un PDF de una sola página
    dimensionada al aspecto de la imagen. Devuelve application/pdf.
    Utilidad para el editor de Planos al guardar de vuelta al evento.
    """
    if body.mime_type not in ("image/jpeg", "image/png"):
        raise HTTPException(400, "Solo JPEG o PNG")
    try:
        raw = base64.b64decode(body.base64)
    except Exception:
        raise HTTPException(400, "Base64 inválido")
    try:
        img = Image.open(io.BytesIO(raw))
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
        buf = io.BytesIO()
        if body.orientation in ("portrait", "landscape"):
            DPI = 150
            A3_PORTRAIT = (int(297 * DPI / 25.4), int(420 * DPI / 25.4))  # 1754x2480
            A3_LANDSCAPE = (int(420 * DPI / 25.4), int(297 * DPI / 25.4))  # 2480x1754
            canvas_size = A3_PORTRAIT if body.orientation == "portrait" else A3_LANDSCAPE
            img.thumbnail(canvas_size, Image.Resampling.LANCZOS)
            canvas = Image.new("RGB", canvas_size, (255, 255, 255))
            offset_x = (canvas_size[0] - img.width) // 2
            offset_y = (canvas_size[1] - img.height) // 2
            canvas.paste(img, (offset_x, offset_y))
            canvas.save(buf, format="PDF", resolution=DPI)
        else:
            img.save(buf, format="PDF", resolution=300.0)
        data = buf.getvalue()
    except Exception as e:
        logging.exception("image-to-pdf failed")
        raise HTTPException(500, f"No se pudo generar PDF: {e}")
    filename = body.filename or "plano.pdf"
    if not filename.lower().endswith(".pdf"):
        filename += ".pdf"
    safe = "".join(c for c in filename if c.isalnum() or c in "._-") or "plano.pdf"
    return Response(
        content=data,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{safe}"'},
    )





# ---------------- Seed ----------------
async def seed_initial_data():
    count = await db.materiales.count_documents({})
    if count > 0:
        return
    path = Path(INITIAL_EXCEL_PATH)
    if not path.exists():
        return
    with open(path, "rb") as f:
        rows = parse_workbook(f.read())
    docs = [{
        "id": str(uuid.uuid4()),
        **r,
        "sync_status": "synced",
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": "system",
    } for r in rows]
    if docs:
        await db.materiales.insert_many(docs)
    logging.getLogger(__name__).info(f"Seeded {len(docs)} materials from {path}")

async def seed_admin_user():
    if not ENABLE_DEMO_SEED:
        return
    existing = await db.users.find_one({"email": DEMO_ADMIN_EMAIL})
    if existing:
        if DEMO_ADMIN_PASSWORD:
            role = await db.roles.find_one({"key": "admin"})
            await db.users.update_one(
                {"email": DEMO_ADMIN_EMAIL},
                {"$set": {
                    "password": hash_password(DEMO_ADMIN_PASSWORD),
                    "name": DEMO_ADMIN_NAME,
                    "role": "admin",
                    "role_id": role.get("id") if role else None,
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                }}
            )
        else:
            logging.getLogger(__name__).warning(
                "DEMO_ADMIN_PASSWORD no configurada; no se puede actualizar el password del usuario admin de demo"
            )
        return
    if not DEMO_ADMIN_PASSWORD:
        logging.getLogger(__name__).warning(
            "DEMO_ADMIN_PASSWORD no configurada; no se puede crear el usuario admin de demo"
        )
        return
    role = await db.roles.find_one({"key": "admin"})
    user = {
        "id": str(uuid.uuid4()),
        "email": DEMO_ADMIN_EMAIL,
        "name": DEMO_ADMIN_NAME,
        "password": hash_password(DEMO_ADMIN_PASSWORD),
        "role": "admin",
        "role_id": role.get("id") if role else None,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(user)

async def _seed_demo_user(email: str, name: str, role_key: str, color: str) -> Optional[str]:
    existing = await db.users.find_one({"email": email})
    if existing:
        return existing.get("id")
    if not DEMO_USER_PASSWORD:
        logging.getLogger(__name__).warning(
            "DEMO_USER_PASSWORD no configurada; no se crean usuarios demo adicionales"
        )
        return None
    role = await db.roles.find_one({"key": role_key})
    legacy_role = "admin" if role_key in ("admin", "gestor") else ("comercial" if role_key == "comercial" else "user")
    user_slug = email.split("@", 1)[0].replace(".", "-").replace("_", "-")
    user_id = f"demo-user-{user_slug}"
    await db.users.insert_one({
        "id": user_id,
        "email": email,
        "name": name,
        "password": hash_password(DEMO_USER_PASSWORD),
        "role": legacy_role,
        "role_id": role.get("id") if role else None,
        "color": color,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "demo_seed": True,
    })
    return user_id

async def seed_demo_data():
    if not ENABLE_DEMO_SEED:
        return

    now = datetime.now(timezone.utc)
    now_iso = now.isoformat()
    admin = await db.users.find_one({"email": DEMO_ADMIN_EMAIL}, {"_id": 0})
    manager_id = admin.get("id") if admin else None
    gestor_id = await _seed_demo_user("gestor.demo@materiales.com", "Gestor Demo", "gestor", "#3B82F6")
    gestor_obras_id = await _seed_demo_user("gestor.obras.demo@materiales.com", "Gestor Obras Demo", "gestor", "#06B6D4")
    tecnico_id = await _seed_demo_user("tecnico.demo@materiales.com", "Tecnico Demo", "tecnico", "#10B981")
    tecnico_refuerzo_id = await _seed_demo_user("tecnico.refuerzo.demo@materiales.com", "Tecnico Refuerzo Demo", "tecnico", "#14B8A6")
    tecnico_mantenimiento_id = await _seed_demo_user("tecnico.mantenimiento.demo@materiales.com", "Tecnico Mantenimiento Demo", "tecnico", "#84CC16")
    await _seed_demo_user("comercial.demo@materiales.com", "Comercial Demo", "comercial", "#F59E0B")
    await _seed_demo_user("comercial.obras.demo@materiales.com", "Comercial Obras Demo", "comercial", "#F97316")
    sat_id = await _seed_demo_user("sat.demo@materiales.com", "SAT Demo", "sat", "#8B5CF6")
    await _seed_demo_user("sat.guardias.demo@materiales.com", "SAT Guardias Demo", "sat", "#EC4899")

    await db.materiales.delete_many({"demo_seed": True, "id": {"$regex": "^demo-material-"}})
    await db.budgets.update_many({"demo_seed": True, "material_id": {"$regex": "^demo-material-"}}, {"$set": {"material_id": None}})
    await db.events.update_many({"demo_seed": True, "material_id": {"$regex": "^demo-material-"}}, {"$set": {"material_id": None}})

    budget_id = "demo-budget-oficina-centro"
    await db.budgets.update_one(
        {"id": budget_id},
        {"$setOnInsert": {
            "id": budget_id,
            "n_proyecto": "POC-001",
            "cliente": "Comunidad Edificio Centro",
            "nombre_instalacion": "PCI zonas comunes y garaje",
            "direccion": "Av. Principal 123",
            "contacto_1": "Administracion Fincas Centro",
            "contacto_2": "600 123 456",
            "observaciones_presupuesto": "Presupuesto demo para mostrar flujo comercial, ejecucion y SAT.",
            "fecha_inicio": (now + timedelta(days=3)).date().isoformat(),
            "fecha_fin": (now + timedelta(days=5)).date().isoformat(),
            "observaciones_ejecucion": "Coordinar acceso a garaje con conserjeria.",
            "equipos": [
                {"elemento": "Central PCI", "cantidad": "1", "ubicacion": "Cuarto tecnico", "observaciones": "Incluye configuracion inicial"},
                {"elemento": "Detector optico", "cantidad": "18", "ubicacion": "Zonas comunes", "observaciones": "Reposicion por sectores"},
                {"elemento": "Sirena interior", "cantidad": "4", "ubicacion": "Escaleras", "observaciones": "Prueba acustica final"},
            ],
            "entrega_tarjeta_mantenimiento": True,
            "entrega_llave_salto": False,
            "entrega_eps100": True,
            "firma_isai": "",
            "nombre_isai": DEMO_ADMIN_NAME,
            "cargo_isai": "Responsable tecnico",
            "firma_cliente": "",
            "nombre_cliente": "Administracion Fincas Centro",
            "cargo_cliente": "Cliente",
            "material_id": None,
            "created_at": now_iso,
            "updated_at": now_iso,
            "created_by": DEMO_ADMIN_EMAIL,
            "created_by_name": DEMO_ADMIN_NAME,
            "status": "pendiente",
            "demo_seed": True,
        }},
        upsert=True,
    )

    demo_budgets = [
        ("demo-budget-hotel-aurora", "POC-002", "Hotel Aurora", "Renovacion PCI recepcion y parking", "pendiente", 2),
        ("demo-budget-talleres-meridian", "POC-003", "Talleres Meridian", "Adecuacion nave industrial", "aceptado", -6),
        ("demo-budget-logistica-eurosur", "POC-004", "Logistica Eurosur", "Ampliacion deteccion muelles", "aceptado", -14),
        ("demo-budget-colegio-san-mateo", "POC-005", "Colegio San Mateo", "Revision anual y mejoras evacuacion", "pendiente", 7),
        ("demo-budget-reformas-horizonte", "POC-006", "Reformas Horizonte", "Instalacion oficinas nueva sede", "pendiente", 12),
    ]
    for demo_budget_id, project_number, cliente, instalacion, status, day_offset in demo_budgets:
        await db.budgets.update_one(
            {"id": demo_budget_id},
            {"$setOnInsert": {
                "id": demo_budget_id,
                "n_proyecto": project_number,
                "cliente": cliente,
                "nombre_instalacion": instalacion,
                "direccion": "Direccion demo PoC",
                "contacto_1": f"Responsable {cliente}",
                "contacto_2": "600 000 000",
                "observaciones_presupuesto": "Presupuesto demo con estado y proyecto vinculado.",
                "fecha_inicio": (now + timedelta(days=day_offset)).date().isoformat(),
                "fecha_fin": (now + timedelta(days=day_offset + 3)).date().isoformat(),
                "observaciones_ejecucion": "Coordinar accesos, cortes y pruebas con cliente.",
                "equipos": [
                    {"elemento": "Central PCI", "cantidad": "1", "ubicacion": "Cuarto tecnico", "observaciones": "Configuracion demo"},
                    {"elemento": "Detector optico", "cantidad": "12", "ubicacion": "Zonas comunes", "observaciones": "Instalacion por sectores"},
                    {"elemento": "Prueba funcional", "cantidad": "1", "ubicacion": "Toda la instalacion", "observaciones": "Acta final"},
                ],
                "entrega_tarjeta_mantenimiento": status == "aceptado",
                "entrega_llave_salto": False,
                "entrega_eps100": status == "aceptado",
                "firma_isai": "",
                "nombre_isai": DEMO_ADMIN_NAME,
                "cargo_isai": "Responsable tecnico",
                "firma_cliente": "",
                "nombre_cliente": f"Responsable {cliente}",
                "cargo_cliente": "Cliente",
                "material_id": None,
                "created_at": now_iso,
                "updated_at": now_iso,
                "created_by": DEMO_ADMIN_EMAIL,
                "created_by_name": DEMO_ADMIN_NAME,
                "status": status,
                "demo_seed": True,
            }},
            upsert=True,
        )

    event_id = "demo-event-instalacion-centro"
    start_at = (now + timedelta(days=3)).replace(hour=9, minute=0, second=0, microsecond=0)
    end_at = start_at + timedelta(hours=4)
    await db.events.update_one(
        {"id": event_id},
        {"$setOnInsert": {
            "id": event_id,
            "title": "Instalacion PCI - Comunidad Edificio Centro",
            "start_at": start_at.isoformat(),
            "end_at": end_at.isoformat(),
            "description": "Demo PoC: instalacion planificada con tecnico asignado y presupuesto vinculado.",
            "material_id": None,
            "assigned_user_ids": [uid for uid in [tecnico_id, tecnico_refuerzo_id] if uid],
            "manager_id": gestor_id or manager_id,
            "recurrence": None,
            "attachments": [],
            "created_by": DEMO_ADMIN_EMAIL,
            "created_at": now_iso,
            "status": "in_progress",
            "seguimiento": "Pendiente de confirmar acceso a garaje.",
            "budget_id": budget_id,
            "demo_seed": True,
        }},
        upsert=True,
    )

    tech_pool = [uid for uid in [tecnico_id, tecnico_refuerzo_id, tecnico_mantenimiento_id] if uid]
    manager_pool = [uid for uid in [gestor_id, gestor_obras_id, manager_id] if uid]
    if tech_pool and manager_pool:
        await db.events.delete_many({"demo_seed": True, "id": {"$regex": "^demo-calendar-"}})
        calendar_start = now.replace(hour=8, minute=0, second=0, microsecond=0)
        calendar_end = calendar_start + timedelta(days=29)
        workdays = []
        cursor = calendar_start
        while cursor <= calendar_end:
            if cursor.weekday() < 5:
                workdays.append(cursor)
            cursor += timedelta(days=1)

        event_templates = [
            ("Revision trimestral PCI", "Mantenimiento preventivo y checklist de central.", 8, 0, 150, "in_progress"),
            ("Instalacion detectores", "Montaje y prueba por sectores.", 10, 30, 180, "in_progress"),
            ("Puesta en marcha", "Configuracion final, pruebas y entrega al cliente.", 13, 30, 150, "pending_completion"),
            ("Visita SAT programada", "Revision correctiva con parte de trabajo.", 16, 0, 120, "in_progress"),
        ]
        client_names = [
            "Comunidad Edificio Centro",
            "Hotel Aurora",
            "Talleres Meridian",
            "Logistica Eurosur",
            "Colegio San Mateo",
            "Reformas Horizonte",
        ]

        for day_idx, day in enumerate(workdays):
            for block_idx, (title, description, hour, minute, duration_minutes, status) in enumerate(event_templates):
                start = day.replace(hour=hour, minute=minute, second=0, microsecond=0)
                end = start + timedelta(minutes=duration_minutes)
                assigned_count = ((day_idx + block_idx) % min(3, len(tech_pool))) + 1
                assigned = [tech_pool[(day_idx + block_idx + offset) % len(tech_pool)] for offset in range(assigned_count)]
                manager = manager_pool[(day_idx + block_idx) % len(manager_pool)]
                client = client_names[(day_idx + block_idx) % len(client_names)]
                event_id_block = f"demo-calendar-workday-{day.date().isoformat()}-{block_idx + 1}"
                await db.events.update_one(
                    {"id": event_id_block},
                    {"$setOnInsert": {
                        "id": event_id_block,
                        "title": f"{title} - {client}",
                        "start_at": start.isoformat(),
                        "end_at": end.isoformat(),
                        "description": description,
                        "material_id": None,
                        "assigned_user_ids": assigned,
                        "manager_id": manager,
                        "recurrence": None,
                        "attachments": [],
                        "created_by": DEMO_ADMIN_EMAIL,
                        "created_at": now_iso,
                        "status": status,
                        "seguimiento": "Bloque laboral demo para calendario PoC.",
                        "budget_id": budget_id if block_idx == 0 and day_idx % 3 == 0 else None,
                        "demo_seed": True,
                    }},
                    upsert=True,
                )

        recurrent_specs = [
            ("demo-calendar-daily-standup", "Coordinacion diaria de tecnicos", "Revision de agenda, materiales y accesos del dia.", "daily", calendar_start, 8, 0, 30, tech_pool[:1], manager_pool[0]),
            ("demo-calendar-weekly-planning", "Planificacion semanal de obras", "Reunion semanal de gestor con equipo tecnico.", "weekly", calendar_start, 12, 30, 60, tech_pool[:min(3, len(tech_pool))], manager_pool[min(1, len(manager_pool) - 1)]),
            ("demo-calendar-monthly-review", "Revision mensual de cartera", "Seguimiento mensual de presupuestos, SAT y obras activas.", "monthly", calendar_start, 9, 30, 90, tech_pool[:min(2, len(tech_pool))], manager_pool[0]),
        ]
        for event_id_rec, title, description, recurrence_type, base_day, hour, minute, duration_minutes, assigned, manager in recurrent_specs:
            start = base_day.replace(hour=hour, minute=minute, second=0, microsecond=0)
            end = start + timedelta(minutes=duration_minutes)
            await db.events.update_one(
                {"id": event_id_rec},
                {"$setOnInsert": {
                    "id": event_id_rec,
                    "title": title,
                    "start_at": start.isoformat(),
                    "end_at": end.isoformat(),
                    "description": description,
                    "material_id": None,
                    "assigned_user_ids": assigned,
                    "manager_id": manager,
                    "recurrence": {"type": recurrence_type, "until": calendar_end.date().isoformat()},
                    "attachments": [],
                    "created_by": DEMO_ADMIN_EMAIL,
                    "created_at": now_iso,
                    "status": "in_progress",
                    "seguimiento": "Evento recurrente demo para calendario PoC.",
                    "budget_id": None,
                    "demo_seed": True,
                }},
                upsert=True,
            )

    client_id = "demo-sat-client-centro"
    await db.sat_clients.update_one(
        {"id": client_id},
        {"$setOnInsert": {
            "id": client_id,
            "cliente": "Comunidad Edificio Centro",
            "direccion": "Av. Principal 123",
            "contacto": "Administracion Fincas Centro",
            "telefono": "600 123 456",
            "created_at": now_iso,
            "updated_at": now_iso,
            "demo_seed": True,
        }},
        upsert=True,
    )

    incident_id = "demo-sat-incident-centro"
    await db.sat_incidents.update_one(
        {"id": incident_id},
        {"$setOnInsert": {
            "id": incident_id,
            "cliente": "Comunidad Edificio Centro",
            "direccion": "Av. Principal 123",
            "telefono": "600 123 456",
            "observaciones": "Aviso demo: revisar sirena de garaje tras prueba de mantenimiento.",
            "comentarios_sat": "Caso preparado para mostrar seguimiento SAT.",
            "status": "pendiente",
            "client_id": client_id,
            "created_at": now_iso,
            "updated_at": now_iso,
            "resolved_at": None,
            "resolved_by": None,
            "history": [{
                "id": "demo-sat-history-centro",
                "action": "note",
                "comment": "Incidencia demo generada para PoC.",
                "user_id": sat_id,
                "user_name": "SAT Demo",
                "created_at": now_iso,
            }],
            "demo_seed": True,
        }},
        upsert=True,
    )

    sat_clients_demo = [
        ("demo-sat-client-hotel-aurora", "Hotel Aurora", "C/ Marina 48", "Marta Ruiz", "600 111 222"),
        ("demo-sat-client-talleres-meridian", "Talleres Meridian", "Poligono Norte nave 12", "Carlos Molina", "600 222 333"),
        ("demo-sat-client-logistica-eurosur", "Logistica Eurosur", "Centro logistico Sur", "Nuria Campos", "600 333 444"),
        ("demo-sat-client-colegio-san-mateo", "Colegio San Mateo", "Av. Educacion 22", "Secretaria tecnica", "600 444 555"),
        ("demo-sat-client-reformas-horizonte", "Reformas Horizonte", "C/ Norte 7", "Sergio Vidal", "600 555 666"),
        ("demo-sat-client-clinica-norte", "Clinica Norte", "Paseo Salud 3", "Recepcion mantenimiento", "600 666 777"),
    ]
    for client_id_demo, cliente, direccion, contacto, telefono in sat_clients_demo:
        await db.sat_clients.update_one(
            {"id": client_id_demo},
            {"$setOnInsert": {
                "id": client_id_demo,
                "cliente": cliente,
                "direccion": direccion,
                "contacto": contacto,
                "telefono": telefono,
                "created_at": now_iso,
                "updated_at": now_iso,
                "demo_seed": True,
            }},
            upsert=True,
        )

    sat_incidents_demo = [
        ("demo-sat-incident-hotel-pendiente", "demo-sat-client-hotel-aurora", "Hotel Aurora", "C/ Marina 48", "600 111 222", "Fallo intermitente en sirena de parking durante prueba semanal.", "pendiente", None, None, None, 0),
        ("demo-sat-incident-taller-resuelta", "demo-sat-client-talleres-meridian", "Talleres Meridian", "Poligono Norte nave 12", "600 222 333", "Pulsador golpeado en zona de carga. Sustituido y probado.", "resuelta", None, True, sat_id, -3),
        ("demo-sat-incident-logistica-agendada", "demo-sat-client-logistica-eurosur", "Logistica Eurosur", "Centro logistico Sur", "600 333 444", "Detector lineal con falsas alarmas en muelle 4. Visita programada.", "agendada", now + timedelta(days=2, hours=9), None, None, -1),
        ("demo-sat-incident-colegio-pendiente", "demo-sat-client-colegio-san-mateo", "Colegio San Mateo", "Av. Educacion 22", "600 444 555", "Revision de senaletica de evacuacion tras inspeccion interna.", "pendiente", None, None, None, -2),
        ("demo-sat-incident-reformas-resuelta", "demo-sat-client-reformas-horizonte", "Reformas Horizonte", "C/ Norte 7", "600 555 666", "Central sin alimentacion auxiliar. Bateria sustituida.", "resuelta", None, False, sat_id, -7),
        ("demo-sat-incident-clinica-agendada", "demo-sat-client-clinica-norte", "Clinica Norte", "Paseo Salud 3", "600 666 777", "Comprobar sectorizacion de zona consultas antes de auditoria.", "agendada", now + timedelta(days=5, hours=10), None, None, -4),
    ]
    for incident_demo_id, client_id_demo, cliente, direccion, telefono, observaciones, status, scheduled_for, facturable, resolved_by, created_offset in sat_incidents_demo:
        created_at = (now + timedelta(days=created_offset)).isoformat()
        history = [{
            "id": f"{incident_demo_id}-history-open",
            "action": "created",
            "comment": "Incidencia demo creada para mostrar flujo SAT.",
            "user_id": None,
            "user_name": cliente,
            "created_at": created_at,
        }]
        if status == "resuelta":
            history.append({
                "id": f"{incident_demo_id}-history-resolved",
                "action": "status_change",
                "from_status": "pendiente",
                "to_status": "resuelta",
                "comment": "Caso resuelto en visita demo.",
                "facturable": facturable,
                "user_id": sat_id,
                "user_name": "SAT Demo",
                "created_at": now_iso,
            })
        if status == "agendada":
            history.append({
                "id": f"{incident_demo_id}-history-scheduled",
                "action": "scheduled",
                "from_status": "pendiente",
                "to_status": "agendada",
                "scheduled_for": scheduled_for.isoformat() if scheduled_for else None,
                "comment": "Visita agendada para resolver incidencia demo.",
                "user_id": sat_id,
                "user_name": "SAT Demo",
                "created_at": now_iso,
            })
        await db.sat_incidents.update_one(
            {"id": incident_demo_id},
            {"$setOnInsert": {
                "id": incident_demo_id,
                "cliente": cliente,
                "direccion": direccion,
                "telefono": telefono,
                "observaciones": observaciones,
                "comentarios_sat": "Caso demo con historial para CRM SAT.",
                "status": status,
                "client_id": client_id_demo,
                "scheduled_for": scheduled_for.isoformat() if scheduled_for else None,
                "facturable": facturable,
                "created_at": created_at,
                "updated_at": now_iso,
                "resolved_at": now_iso if status == "resuelta" else None,
                "resolved_by": resolved_by,
                "history": history,
                "demo_seed": True,
            }},
            upsert=True,
        )

# ---------------- Root/Health ----------------
@api_router.get("/")
async def root():
    return {"status": "ok", "service": "Materiales OneDrive App"}


# ====================  CRM SAT — incidencias  ====================
# Flujo:
#   1) El cliente abre un enlace público (`/aviso-sat`), rellena un formulario
#      y envía. Esto crea un registro en la colección `sat_incidents` vía
#      POST /sat/public (SIN autenticación).
#   2) El equipo SAT (usuarios autenticados) consulta la lista con
#      GET /sat/incidents, añade comentarios internos y marca como
#      "pendiente" o "resuelta" con PATCH /sat/incidents/{id}.

class SATPublicIn(BaseModel):
    cliente: str = Field(..., min_length=1, max_length=200)
    direccion: str = Field("", max_length=400)
    telefono: str = Field("", max_length=60)
    observaciones: str = Field(..., min_length=1, max_length=4000)

class SATUpdateIn(BaseModel):
    cliente: Optional[str] = None
    direccion: Optional[str] = None
    telefono: Optional[str] = None
    observaciones: Optional[str] = None
    comentarios_sat: Optional[str] = None
    status: Optional[str] = None  # "pendiente" | "resuelta"

_sat_rl: dict = {}  # IP → list of epoch timestamps

_SAT_RL_WINDOW_S = 30
_SAT_RL_MAX = 3

def _request_client_ip(request: Request) -> str:
    if TRUST_PROXY_HEADERS:
        forwarded = request.headers.get("x-forwarded-for", "").split(",")[0].strip()
        if forwarded:
            return forwarded
    return request.client.host if request.client else "unknown"

def _check_sat_rate_limit(ip: str) -> None:
    now = time.monotonic()
    entries = [t for t in _sat_rl.get(ip, []) if now - t < _SAT_RL_WINDOW_S]
    _sat_rl[ip] = entries
    if len(entries) >= _SAT_RL_MAX:
        raise HTTPException(429, "Demasiadas solicitudes. Esperá unos segundos.")
    _sat_rl[ip].append(now)

@api_router.post("/sat/public")
async def sat_public_create(body: SATPublicIn, request: Request):
    """Endpoint PÚBLICO — no requiere login. Lo usa el enlace que se envía
    al cliente para que abra la incidencia."""
    ip = _request_client_ip(request)
    _check_sat_rate_limit(ip)
    now = datetime.now(timezone.utc)
    doc = {
        "id": str(uuid.uuid4()),
        "cliente": body.cliente.strip(),
        "direccion": (body.direccion or "").strip(),
        "telefono": (body.telefono or "").strip(),
        "observaciones": body.observaciones.strip(),
        "comentarios_sat": "",
        "status": "pendiente",
        "created_at": now.isoformat(),
        "updated_at": now.isoformat(),
        "resolved_at": None,
        "resolved_by": None,
    }
    # Auto-link to catalog client if the name matches an existing client.
    if doc["cliente"]:
        matched = await db.sat_clients.find_one(
            {"cliente": {"$regex": f"^{re.escape(doc['cliente'])}$", "$options": "i"}},
            {"_id": 0, "id": 1},
        )
        if matched:
            doc["client_id"] = matched["id"]
    await db.sat_incidents.insert_one(doc)
    doc.pop("_id", None)
    # Notificar a los administradores
    admins = await db.users.find({"role": "admin"}, {"_id": 0, "id": 1}).to_list(500)
    for a in admins:
        await db.notifications.insert_one({
            "id": str(uuid.uuid4()),
            "user_id": a["id"],
            "event_id": None,
            "type": "sat_new",
            "title": f"Nuevo aviso SAT: {doc['cliente']}",
            "message": (doc["observaciones"][:160] + ("…" if len(doc["observaciones"]) > 160 else "")),
            "read": False,
            "created_at": now.isoformat(),
            "from_user_id": None,
            "from_user_name": doc["cliente"],
            "link": f"/sat?openIncident={doc['id']}",
        })
    return {"ok": True, "id": doc["id"]}

@api_router.get("/sat/incidents")
async def sat_list(
    user: dict = Depends(require_permission("sat.view")),
    status: Optional[str] = None,
    client_id: Optional[str] = None,
    year: Optional[str] = None,
    month: Optional[str] = None,
):
    # Lazy auto-revive: any incident scheduled for past time returns as "pendiente".
    now_iso = datetime.now(timezone.utc).isoformat()
    await _sat_auto_revive(now_iso)

    q: dict = {}
    if status in {"pendiente", "resuelta", "agendada"}:
        q["status"] = status
    if client_id:
        # Match by client_id OR by cliente name (case-insensitive) for incidents
        # submitted via the public form that didn't have client_id linked.
        client_doc = await db.sat_clients.find_one({"id": client_id}, {"_id": 0})
        client_name = (client_doc or {}).get("cliente", "")
        if client_name:
            q["$or"] = [
                {"client_id": client_id},
                {"cliente": {"$regex": f"^{re.escape(client_name)}$", "$options": "i"}},
            ]
        else:
            q["client_id"] = client_id
    if year and year != "todos":
        q["created_at"] = {"$regex": f"^{year}"}
    if month:
        # month viene como "01" a "12", coincidir con "-MM" en ISO timestamp
        q["created_at"] = {"$regex": f"-{month}"} if not q.get("created_at") else q["created_at"]
        # NOTA: si se pasan ambos year y month, esto pisa el year regex. Se combinan mejor así:
    if year and year != "todos" and month:
        q["created_at"] = {"$regex": f"^{year}-{month}"}
    rows = await db.sat_incidents.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return rows

async def _sat_auto_revive(now_iso: str):
    """Move incidents whose scheduled_for date has passed back to pendiente
    and leave a history note describing the auto-revival."""
    overdue = await db.sat_incidents.find({
        "status": "agendada",
        "scheduled_for": {"$lte": now_iso},
    }, {"_id": 0}).to_list(500)
    for ev in overdue:
        entry = {
            "id": str(uuid.uuid4()),
            "action": "auto_revive",
            "from_status": "agendada",
            "to_status": "pendiente",
            "comment": "Ha llegado la fecha programada. La incidencia vuelve a Pendiente.",
            "user_id": None,
            "user_name": "Sistema",
            "created_at": now_iso,
        }
        await db.sat_incidents.update_one(
            {"id": ev["id"]},
            {
                "$set": {"status": "pendiente", "updated_at": now_iso, "scheduled_for": None},
                "$push": {"history": entry},
            },
        )
        # Also notify admins about the revived incident
        admins = await db.users.find({"role": "admin"}, {"_id": 0, "id": 1}).to_list(200)
        for a in admins:
            await db.notifications.insert_one({
                "id": str(uuid.uuid4()),
                "user_id": a["id"],
                "event_id": None,
                "type": "sat_revived",
                "title": f"Incidencia reactivada: {ev.get('cliente','')}",
                "message": ev.get("observaciones", "")[:200],
                "read": False,
                "created_at": now_iso,
                "from_user_id": None,
                "from_user_name": "Sistema",
                "link": f"/sat?openIncident={ev['id']}",
            })

@api_router.get("/sat/incidents/{iid}")
async def sat_get(iid: str, user: dict = Depends(require_permission("sat.view"))):
    doc = await db.sat_incidents.find_one({"id": iid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Incidencia no encontrada")
    return doc

@api_router.patch("/sat/incidents/{iid}")
async def sat_update(iid: str, body: SATUpdateIn, user: dict = Depends(require_permission("sat.edit"))):
    existing = await db.sat_incidents.find_one({"id": iid})
    if not existing:
        raise HTTPException(404, "Incidencia no encontrada")
    patch: dict = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if body.cliente is not None: patch["cliente"] = body.cliente.strip()
    if body.direccion is not None: patch["direccion"] = body.direccion.strip()
    if body.telefono is not None: patch["telefono"] = body.telefono.strip()
    if body.observaciones is not None: patch["observaciones"] = body.observaciones.strip()
    if body.comentarios_sat is not None: patch["comentarios_sat"] = body.comentarios_sat
    if body.status in {"pendiente", "resuelta"}:
        patch["status"] = body.status
        if body.status == "resuelta" and existing.get("status") != "resuelta":
            patch["resolved_at"] = datetime.now(timezone.utc).isoformat()
            patch["resolved_by"] = user.get("id")
        elif body.status == "pendiente":
            patch["resolved_at"] = None
            patch["resolved_by"] = None
    await db.sat_incidents.update_one({"id": iid}, {"$set": patch})
    doc = await db.sat_incidents.find_one({"id": iid}, {"_id": 0})
    return doc

@api_router.delete("/sat/incidents/{iid}")
async def sat_delete(iid: str, admin: dict = Depends(require_permission("sat.edit"))):
    res = await db.sat_incidents.delete_one({"id": iid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Incidencia no encontrada")
    return {"ok": True}

class SATStatusChangeIn(BaseModel):
    status: str  # "pendiente" | "resuelta"
    comment: str = Field("", max_length=4000)
    facturable: Optional[bool] = None  # required when status == "resuelta"

@api_router.post("/sat/incidents/{iid}/status")
async def sat_change_status(iid: str, body: SATStatusChangeIn, user: dict = Depends(require_permission("sat.edit"))):
    """Change the status of an incident and append a history entry with the
    user's comment. This is the canonical way to toggle pendiente↔resuelta
    from the UI so that every change is traceable."""
    existing = await db.sat_incidents.find_one({"id": iid})
    if not existing:
        raise HTTPException(404, "Incidencia no encontrada")
    if body.status not in {"pendiente", "resuelta"}:
        raise HTTPException(400, "Estado inválido")
    if body.status == "resuelta" and body.facturable is None:
        raise HTTPException(400, "Debes indicar si la incidencia es facturable o no")

    now = datetime.now(timezone.utc).isoformat()
    prev = existing.get("status") or "pendiente"
    entry = {
        "id": str(uuid.uuid4()),
        "action": "status_change",
        "from_status": prev,
        "to_status": body.status,
        "comment": (body.comment or "").strip(),
        "facturable": body.facturable if body.status == "resuelta" else None,
        "user_id": user.get("id"),
        "user_name": user.get("name") or user.get("email") or "usuario",
        "created_at": now,
    }
    patch = {
        "status": body.status,
        "updated_at": now,
    }
    if body.status == "resuelta" and prev != "resuelta":
        patch["resolved_at"] = now
        patch["resolved_by"] = user.get("id")
        patch["facturable"] = body.facturable
    elif body.status == "pendiente":
        patch["resolved_at"] = None
        patch["resolved_by"] = None
    await db.sat_incidents.update_one(
        {"id": iid},
        {"$set": patch, "$push": {"history": entry}},
    )
    doc = await db.sat_incidents.find_one({"id": iid}, {"_id": 0})
    return doc

class SATScheduleIn(BaseModel):
    scheduled_for: str  # ISO datetime
    comment: str = Field("", max_length=4000)

@api_router.post("/sat/incidents/{iid}/schedule")
async def sat_schedule(iid: str, body: SATScheduleIn, user: dict = Depends(require_permission("sat.edit"))):
    """Reschedule an incident. Moves it to status='agendada' with the chosen
    date/time. When that date arrives, `_sat_auto_revive` flips it back to
    pendiente and notifies the admins."""
    existing = await db.sat_incidents.find_one({"id": iid})
    if not existing:
        raise HTTPException(404, "Incidencia no encontrada")
    try:
        sched = datetime.fromisoformat(body.scheduled_for.replace("Z", "+00:00"))
        if sched.tzinfo is None:
            sched = sched.replace(tzinfo=timezone.utc)
    except Exception:
        raise HTTPException(400, "Fecha inválida")
    now = datetime.now(timezone.utc).isoformat()
    prev = existing.get("status") or "pendiente"
    pretty = sched.strftime("%d/%m/%Y %H:%M")
    entry = {
        "id": str(uuid.uuid4()),
        "action": "scheduled",
        "from_status": prev,
        "to_status": "agendada",
        "scheduled_for": sched.isoformat(),
        "comment": (body.comment or "").strip() or f"Incidencia reagendada para el {pretty}.",
        "user_id": user.get("id"),
        "user_name": user.get("name") or user.get("email") or "usuario",
        "created_at": now,
    }
    await db.sat_incidents.update_one(
        {"id": iid},
        {
            "$set": {
                "status": "agendada",
                "scheduled_for": sched.isoformat(),
                "resolved_at": None,
                "resolved_by": None,
                "updated_at": now,
            },
            "$push": {"history": entry},
        },
    )
    doc = await db.sat_incidents.find_one({"id": iid}, {"_id": 0})
    return doc

@api_router.post("/sat/incidents/{iid}/note")
async def sat_add_note(iid: str, body: SATStatusChangeIn, user: dict = Depends(require_permission("sat.edit"))):
    """Add a free-form note to the history without changing status. Handy if
    the SAT team wants to leave intermediate comments."""
    existing = await db.sat_incidents.find_one({"id": iid})
    if not existing:
        raise HTTPException(404, "Incidencia no encontrada")
    if not (body.comment or "").strip():
        raise HTTPException(400, "El comentario no puede estar vacío")
    now = datetime.now(timezone.utc).isoformat()
    entry = {
        "id": str(uuid.uuid4()),
        "action": "note",
        "comment": body.comment.strip(),
        "user_id": user.get("id"),
        "user_name": user.get("name") or user.get("email") or "usuario",
        "created_at": now,
    }
    await db.sat_incidents.update_one(
        {"id": iid},
        {"$set": {"updated_at": now}, "$push": {"history": entry}},
    )
    doc = await db.sat_incidents.find_one({"id": iid}, {"_id": 0})
    return doc


# ====================  CRM SAT — clientes  ====================
# Catálogo de clientes. Se puede importar desde Excel o mantener
# manualmente. Cada incidencia puede (opcionalmente) enlazar a un cliente
# vía `client_id` para centralizar información.

class SATClientIn(BaseModel):
    cliente: str = Field(..., min_length=1, max_length=240)
    direccion: str = Field("", max_length=400)
    contacto: str = Field("", max_length=160)
    telefono: str = Field("", max_length=80)

@api_router.get("/sat/clients")
async def sat_clients_list(user: dict = Depends(require_permission("sat.view"))):
    rows = await db.sat_clients.find({}, {"_id": 0}).sort("cliente", 1).to_list(5000)
    return rows

@api_router.get("/sat/clients/{cid}")
async def sat_client_get(cid: str, user: dict = Depends(require_permission("sat.view"))):
    doc = await db.sat_clients.find_one({"id": cid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Cliente no encontrado")
    return doc

@api_router.post("/sat/clients")
async def sat_client_create(body: SATClientIn, admin: dict = Depends(require_permission("sat.edit"))):
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "cliente": body.cliente.strip(),
        "direccion": body.direccion.strip(),
        "contacto": body.contacto.strip(),
        "telefono": body.telefono.strip(),
        "created_at": now,
        "updated_at": now,
    }
    await db.sat_clients.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.patch("/sat/clients/{cid}")
async def sat_client_update(cid: str, body: SATClientIn, admin: dict = Depends(require_permission("sat.edit"))):
    existing = await db.sat_clients.find_one({"id": cid})
    if not existing:
        raise HTTPException(404, "Cliente no encontrado")
    patch = {
        "cliente": body.cliente.strip(),
        "direccion": body.direccion.strip(),
        "contacto": body.contacto.strip(),
        "telefono": body.telefono.strip(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.sat_clients.update_one({"id": cid}, {"$set": patch})
    doc = await db.sat_clients.find_one({"id": cid}, {"_id": 0})
    return doc

@api_router.delete("/sat/clients/{cid}")
async def sat_client_delete(cid: str, admin: dict = Depends(require_permission("sat.edit"))):
    res = await db.sat_clients.delete_one({"id": cid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Cliente no encontrado")
    return {"ok": True}

def _match_header(headers: list, *needles: str) -> Optional[int]:
    """Return the index of the first header that contains any of the
    needles (case/accent-insensitive, substring match). Used to support
    Excel files with slight typos in column names (e.g. 'Responsablñe')."""
    import unicodedata
    def norm(s: str) -> str:
        s = str(s or "").strip().lower()
        return "".join(c for c in unicodedata.normalize("NFD", s)
                       if unicodedata.category(c) != "Mn")
    for i, h in enumerate(headers):
        nh = norm(h)
        if any(n in nh for n in needles):
            return i
    return None

@api_router.post("/sat/clients/import")
async def sat_client_import(
    file: UploadFile = File(...),
    replace: bool = False,
    admin: dict = Depends(require_permission("sat.edit")),
):
    """Admin/SAT-only: import clients from an uploaded .xlsx/.xlsm file.
    - Auto-detects columns using keyword matching on headers (tolerant to
      typos like 'Responsablñe' in the real data).
    - `replace=true` wipes the collection before importing.
    - Default behaviour = upsert by `cliente` name (case-insensitive).
    Returns {created, updated, skipped}.
    """
    try:
        from openpyxl import load_workbook
    except Exception:
        raise HTTPException(500, "openpyxl no instalado en el servidor")
    try:
        raw = await file.read()
        import io
        wb = load_workbook(io.BytesIO(raw), data_only=True, keep_vba=False, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Archivo Excel inválido: {e}")

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows_iter))
    except StopIteration:
        raise HTTPException(400, "La hoja está vacía")

    col_cli = _match_header(headers, "client", "nombre")
    col_dir = _match_header(headers, "direcc", "address")
    col_con = _match_header(headers, "responsab", "contact")
    col_tel = _match_header(headers, "tel", "phone", "movil")
    if col_cli is None:
        raise HTTPException(400, "No se ha encontrado una columna 'Cliente' en el Excel")

    if replace:
        await db.sat_clients.delete_many({})

    created = updated = skipped = 0
    now = datetime.now(timezone.utc).isoformat()
    for r in rows_iter:
        try:
            name = str(r[col_cli]).strip() if col_cli is not None and r[col_cli] is not None else ""
        except IndexError:
            name = ""
        if not name:
            skipped += 1
            continue
        def _cell(ix):
            if ix is None: return ""
            try:
                v = r[ix]
                return str(v).strip() if v is not None else ""
            except IndexError:
                return ""
        doc = {
            "cliente": name,
            "direccion": _cell(col_dir),
            "contacto": _cell(col_con),
            "telefono": _cell(col_tel),
            "updated_at": now,
        }
        # Upsert by case-insensitive name match.
        existing = await db.sat_clients.find_one({
            "cliente": {"$regex": f"^{re.escape(name)}$", "$options": "i"}
        })
        if existing:
            await db.sat_clients.update_one({"id": existing["id"]}, {"$set": doc})
            updated += 1
        else:
            doc["id"] = str(uuid.uuid4())
            doc["created_at"] = now
            await db.sat_clients.insert_one(doc)
            created += 1
    return {"ok": True, "created": created, "updated": updated, "skipped": skipped}



class ChatCreate(BaseModel):
    participant_ids: List[str] = Field(..., min_length=1)
    name: Optional[str] = None
    project_id: Optional[str] = None
    event_id: Optional[str] = None

class MessageCreate(BaseModel):
    text: Optional[str] = None
    file_base64: Optional[str] = None
    file_name: Optional[str] = None
    file_mime: Optional[str] = None

# -----------------------------------------------------------------------------
# Portfolio — self-contained HTML presentation for clients.
# -----------------------------------------------------------------------------
from portfolio_view import build_portfolio_html  # noqa: E402
from pathlib import Path as _Path  # noqa: E402

@api_router.get("/portfolio", response_class=HTMLResponse)
async def portfolio_html():
    """Public URL that renders the i-SAI portfolio (no auth, self-contained HTML)."""
    return HTMLResponse(content=build_portfolio_html(), media_type="text/html")


@api_router.get("/portfolio/demo.mp4")
async def portfolio_demo_video(request: Request):
    """Serve the demo video with HTTP byte-range support (needed by <video>)."""
    p = _Path(__file__).resolve().parent.parent / "frontend" / "assets" / "portfolio" / "demo.mp4"
    if not p.exists():
        raise HTTPException(status_code=404, detail="demo video not built")
    file_size = p.stat().st_size
    range_header = request.headers.get("range")
    if range_header:
        # Basic single-range support: "bytes=<start>-<end>"
        try:
            units, rng = range_header.split("=", 1)
            start_s, end_s = rng.split("-", 1)
            start = int(start_s) if start_s else 0
            end = int(end_s) if end_s else file_size - 1
            if start >= file_size:
                raise HTTPException(status_code=416)
            end = min(end, file_size - 1)
            length = end - start + 1
        except Exception:
            raise HTTPException(status_code=416, detail="Invalid Range header")

        async def iterfile():
            with open(str(p), "rb") as f:
                f.seek(start)
                remaining = length
                while remaining > 0:
                    chunk_size = min(65536, remaining)
                    chunk = f.read(chunk_size)
                    if not chunk:
                        break
                    remaining -= len(chunk)
                    yield chunk

        headers = {
            "Content-Range": f"bytes {start}-{end}/{file_size}",
            "Accept-Ranges": "bytes",
            "Content-Length": str(length),
            "Content-Type": "video/mp4",
        }
        return StreamingResponse(iterfile(), status_code=206, headers=headers)

    return FileResponse(p, media_type="video/mp4", headers={"Accept-Ranges": "bytes"})


@api_router.get("/sat/export-excel")
async def sat_export_excel(user: dict = Depends(current_user)):
    perms = await get_user_permissions(user)
    if "sat.export" not in perms and "sat.edit" not in perms:
        raise HTTPException(403, "No tienes permiso para exportar SAT")
    """Export all SAT incidents as an Excel file."""
    incidents = await db.sat_incidents.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Incidencias SAT"
    headers = ["Cliente", "Dirección", "Teléfono", "Observaciones", "Estado", "Creado"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    for i, inc in enumerate(incidents, 2):
        ws.cell(row=i, column=1, value=inc.get("cliente", ""))
        ws.cell(row=i, column=2, value=inc.get("direccion", ""))
        ws.cell(row=i, column=3, value=inc.get("telefono", ""))
        ws.cell(row=i, column=4, value=inc.get("observaciones", ""))
        ws.cell(row=i, column=5, value=inc.get("status", ""))
        ws.cell(row=i, column=6, value=(inc.get("created_at") or "")[:10])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"incidencias_sat_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@api_router.post("/chats")
async def chat_create(payload: ChatCreate, user: dict = Depends(current_user)):
    if user["id"] not in payload.participant_ids:
        payload.participant_ids.append(user["id"])
    pids = list(dict.fromkeys(payload.participant_ids))
    if len(pids) < 2:
        raise HTTPException(400, "Mínimo 2 participantes")
    # For 1-on-1 chats, check if chat already exists
    if len(pids) == 2 and not payload.name:
        existing = await db.chats.find_one({
            "participant_ids": {"$all": pids, "$size": 2},
            "name": {"$in": [None, ""]},
        })
        if existing:
            existing.pop("_id", None)
            return existing
    now = datetime.now(timezone.utc).isoformat()
    chat = {
        "id": str(uuid.uuid4()),
        "participant_ids": pids,
        "name": payload.name or None,
        "project_id": payload.project_id,
        "event_id": payload.event_id,
        "created_by": user["id"],
        "created_at": now,
        "updated_at": now,
    }
    await db.chats.insert_one(chat)
    chat.pop("_id", None)
    return chat

@api_router.get("/chats")
async def chat_list(user: dict = Depends(current_user)):
    chats = await db.chats.find(
        {"participant_ids": user["id"]},
        {"_id": 0},
    ).sort("updated_at", -1).to_list(200)
    # Enrich with last message, participant names, and unread count
    out = []
    for c in chats:
        last_msg = await db.messages.find_one(
            {"chat_id": c["id"]},
            {"_id": 0},
            sort=[("created_at", -1)],
        )
        c["last_message"] = last_msg
        unread = await db.messages.count_documents({
            "chat_id": c["id"],
            "sender_id": {"$ne": user["id"]},
            "read_by": {"$ne": user["id"]},
        })
        c["unread"] = unread
        # Participant names
        users = await db.users.find(
            {"id": {"$in": c["participant_ids"]}},
            {"_id": 0, "id": 1, "name": 1, "email": 1, "color": 1},
        ).to_list(50)
        c["participants"] = users
        out.append(c)
    return out

@api_router.get("/chats/{cid}/messages")
async def chat_messages(cid: str, user: dict = Depends(current_user), limit: int = 50, before: Optional[str] = None):
    chat = await db.chats.find_one({"id": cid})
    if not chat or user["id"] not in chat.get("participant_ids", []):
        raise HTTPException(404, "Chat no encontrado")
    q: dict = {"chat_id": cid}
    if before:
        q["created_at"] = {"$lt": before}
    msgs = await db.messages.find(q, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    # Mark as read
    await db.messages.update_many(
        {"chat_id": cid, "sender_id": {"$ne": user["id"]}, "read_by": {"$ne": user["id"]}},
        {"$push": {"read_by": user["id"]}},
    )
    return list(reversed(msgs))

@api_router.post("/chats/{cid}/messages")
async def chat_send_message(cid: str, payload: MessageCreate, user: dict = Depends(current_user)):
    perms = await get_user_permissions(user)
    if "chat.edit" not in perms and "chat.view" not in perms:
        raise HTTPException(403, "No tienes permiso para enviar mensajes")
    chat = await db.chats.find_one({"id": cid})
    if not chat or user["id"] not in chat.get("participant_ids", []):
        raise HTTPException(404, "Chat no encontrado")
    now = datetime.now(timezone.utc).isoformat()
    msg = {
        "id": str(uuid.uuid4()),
        "chat_id": cid,
        "sender_id": user["id"],
        "sender_name": user.get("name") or user.get("email"),
        "text": (payload.text or "").strip(),
        "created_at": now,
        "read_by": [user["id"]],
    }
    if payload.file_base64 and payload.file_name:
        msg["file_base64"] = payload.file_base64
        msg["file_name"] = payload.file_name
        msg["file_mime"] = payload.file_mime or "application/octet-stream"
    if not msg["text"] and not msg.get("file_base64"):
        raise HTTPException(400, "El mensaje no puede estar vacío")
    await db.messages.insert_one(msg)
    await db.chats.update_one({"id": cid}, {"$set": {"updated_at": now}})
    msg.pop("_id", None)
    # Notify other participants
    sender_name = user.get("name") or user.get("email")
    for pid in chat["participant_ids"]:
        if pid == user["id"]:
            continue
        if await should_notify_user(pid, "chat_message"):
            await db.notifications.insert_one({
                "id": str(uuid.uuid4()),
                "user_id": pid,
                "event_id": None,
                "type": "chat_message",
                "title": f"💬 {sender_name}",
                "message": payload.text.strip()[:200],
                "read": False,
                "created_at": now,
                "from_user_id": user["id"],
                "from_user_name": sender_name,
                "link": f"/chat/{cid}",
            })
    return msg

@api_router.get("/chats/unread-total")
async def chat_unread_total(user: dict = Depends(current_user)):
    total = await db.messages.count_documents({
        "sender_id": {"$ne": user["id"]},
        "read_by": {"$ne": user["id"]},
    })
    return {"unread": total}

@api_router.get("/preciario/productos")
async def get_preciario_productos(
    q: Optional[str] = Query(None, description="Buscar por referencia o descripción"),
    stock_min: Optional[int] = Query(None, ge=0, description="Filtrar por stock mínimo (ej: 1 = solo productos con stock)"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    user: dict = Depends(require_permission("preciario.view")),
):
    user_perms = await get_user_permissions(user)
    puede_ver_precios = "preciario.ver_precios" in user_perms or "preciario.edit" in user_perms
    # Cache del Excel: evitar leer y parsear 56k filas en cada request
    if not hasattr(get_preciario_productos, "_cache"):
        get_preciario_productos._cache = {"at": 0, "items": []}
    cache = get_preciario_productos._cache
    now_ts = time.time()
    if now_ts - cache["at"] < 300 and cache["items"]:
        items = list(cache["items"])  # copia para no mutar
    else:
        wb = load_workbook(ROOT_DIR / "Tarifas2025 para isai.xlsx", data_only=True)
        try:
            ws = wb["tar2025"]
            items = []
            for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
                ref_raw, descripcion_raw, precio_raw = row[0], row[1], row[2]
                if ref_raw is None and descripcion_raw is None and precio_raw is None:
                    continue
                ref = str(ref_raw).strip() if ref_raw is not None else ""
                descripcion = str(descripcion_raw).strip() if descripcion_raw is not None else ""
                precio_str = str(precio_raw).strip() if precio_raw is not None else ""
                try:
                    # El Excel puede tener coma como separador decimal y punto como miles: 1.234,56
                    precio_str = precio_str.replace(".", "").replace(",", ".")
                    precio_unitario = float(precio_str)
                except (ValueError, TypeError):
                    precio_unitario = 0.0
                items.append({
                    "ref": ref,
                    "descripcion": descripcion,
                    "precio_unitario": precio_unitario,
                })
            cache["at"] = now_ts
            cache["items"] = items
        finally:
            wb.close()
    # Aplicar visibilidad de precios (si no tiene permiso, ocultar en la copia de respuesta)
    items_out = []
    for it in items:
        items_out.append({
            **it,
            "precio_unitario": it["precio_unitario"] if puede_ver_precios else 0.0,
        })
    items = items_out

    # Cargar descuentos y stock guardados
    docs = await db.preciario_descuentos.find({}, {"_id": 0}).to_list(length=None)
    descuentos = {d["ref"]: d["descuento"] for d in docs}
    stocks = {d["ref"]: d.get("stock", 0) for d in docs}

    # Filtro por texto
    if q:
        q_lower = q.lower()
        items = [i for i in items if q_lower in i["ref"].lower() or q_lower in i["descripcion"].lower()]

    # Filtro por stock mínimo
    if stock_min is not None:
        items = [i for i in items if stocks.get(i["ref"], 0) >= stock_min]

    total = len(items)
    start = (page - 1) * page_size
    items_pagina = items[start:start + page_size]

    return {
        "items": items_pagina,
        "descuentos": descuentos,
        "stocks": stocks,
        "total": total,
        "page": page,
        "page_size": page_size,
    }

@api_router.get("/preciario/descuentos")
async def get_preciario_descuentos(
    user: dict = Depends(require_permission("preciario.view")),
):
    """Devuelve todos los descuentos guardados: { ref: descuento }"""
    docs = await db.preciario_descuentos.find({}, {"_id": 0}).to_list(length=None)
    return {d["ref"]: d["descuento"] for d in docs}

class DescuentoBody(BaseModel):
    ref: str
    descuento: int = Field(ge=0, le=100)

@api_router.patch("/preciario/descuentos")
async def update_preciario_descuento(
    body: DescuentoBody,
    user: dict = Depends(require_permission("preciario.edit")),
):
    """Guarda o actualiza el descuento de un producto. Requiere permiso preciario.edit."""
    await db.preciario_descuentos.update_one(
        {"ref": body.ref},
        {"$set": {"descuento": body.descuento}},
        upsert=True,
    )
    return {"ok": True}

class StockBody(BaseModel):
    ref: str
    stock: int = Field(ge=0, le=100)

@api_router.patch("/preciario/stock")
async def update_preciario_stock(
    body: StockBody,
    user: dict = Depends(require_permission("preciario.edit")),
):
    """Guarda o actualiza el stock de un producto. Requiere permiso preciario.edit."""
    await db.preciario_descuentos.update_one(
        {"ref": body.ref},
        {"$set": {"stock": body.stock}},
        upsert=True,
    )
    return {"ok": True}


# ---------------- Notas ----------------

class NotaCreate(BaseModel):
    titulo: str = ""
    contenido: str = ""
    fecha: Optional[str] = None
    material_id: Optional[str] = None
    marcada: bool = False
    color: Optional[str] = None
    priority: Optional[str] = None
    tags: Optional[list] = []
    pinned: bool = False
    archived: bool = False

class NotaUpdate(BaseModel):
    titulo: Optional[str] = None
    contenido: Optional[str] = None
    fecha: Optional[str] = None
    material_id: Optional[str] = None
    marcada: Optional[bool] = None
    color: Optional[str] = None
    priority: Optional[str] = None
    tags: Optional[list] = None
    pinned: Optional[bool] = None
    archived: Optional[bool] = None

@api_router.get("/notas")
async def list_notas(
    user: dict = Depends(require_permission("notas.view")),
    fecha: Optional[str] = Query(None, description="Filtrar por fecha YYYY-MM-DD"),
    marcada: Optional[bool] = Query(None, description="Filtrar por marcadas (true/false)"),
    search: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    tag: Optional[str] = Query(None),
    material_id: Optional[str] = Query(None),
    pinned: Optional[bool] = Query(None),
    archived: Optional[bool] = Query(False),
):
    q: dict = {"user_id": user["id"]}
    if fecha:
        q["fecha"] = fecha
    if marcada is not None:
        q["marcada"] = marcada
    if search:
        q["$or"] = [{"titulo": {"$regex": search, "$options": "i"}}, {"contenido": {"$regex": search, "$options": "i"}}]
    if priority:
        q["priority"] = priority
    if tag:
        q["tags"] = tag
    if material_id:
        q["material_id"] = material_id
    if pinned is not None:
        q["pinned"] = pinned
    if archived is not None:
        q["archived"] = archived
    else:
        q["archived"] = {"$ne": True}
    items = await db.notas.find(q, {"_id": 0}).sort("updated_at", -1).to_list(500)
    # Enriquecer con nombre de proyecto
    mids = {n["material_id"] for n in items if n.get("material_id")}
    if mids:
        mats = await db.materiales.find({"id": {"$in": list(mids)}}, {"_id": 0, "id": 1, "materiales": 1, "cliente": 1}).to_list(500)
        mat_map = {m["id"]: f"{m.get('materiales') or ''} — {m.get('cliente') or ''}" for m in mats}
        for n in items:
            if n.get("material_id") in mat_map:
                n["material_name"] = mat_map[n["material_id"]]
    return items

@api_router.post("/notas")
async def create_nota(
    body: NotaCreate,
    user: dict = Depends(require_permission("notas.view")),
):
    nid = str(uuid.uuid4())
    doc = {
        "_id": nid,
        "id": nid,
        "user_id": user["id"],
        "titulo": body.titulo or "",
        "contenido": body.contenido or "",
        "fecha": body.fecha,
        "material_id": body.material_id or None,
        "marcada": body.marcada or False,
        "color": body.color or None,
        "priority": body.priority or None,
        "tags": body.tags or [],
        "pinned": body.pinned or False,
        "archived": body.archived or False,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.notas.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.patch("/notas/{nid}")
async def update_nota(
    nid: str,
    body: NotaUpdate,
    user: dict = Depends(require_permission("notas.view")),
):
    doc = await db.notas.find_one({"id": nid, "user_id": user["id"]})
    if not doc:
        raise HTTPException(404, "Nota no encontrada")
    upd = {k: v for k, v in body.dict().items() if v is not None}
    if not upd:
        raise HTTPException(400, "Nada que actualizar")
    upd["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.notas.update_one({"id": nid}, {"$set": upd})
    doc.update(upd)
    return doc

@api_router.delete("/notas/{nid}")
async def delete_nota(
    nid: str,
    user: dict = Depends(require_permission("notas.view")),
):
    res = await db.notas.delete_one({"id": nid, "user_id": user["id"]})
    if res.deleted_count == 0:
        raise HTTPException(404, "Nota no encontrada")
    return {"ok": True}


# ---------------- Documentos (fichas técnicas / manuales) ----------------

DOCUMENTOS_DIR = ROOT_DIR / "uploads" / "documentos"
DOCUMENTOS_DIR.mkdir(parents=True, exist_ok=True)

# Archivos ordenados por proyecto (interno)
ARCHIVOS_DIR = ROOT_DIR / "Archivos ordenados"
ARCHIVOS_DIR.mkdir(parents=True, exist_ok=True)

async def _sync_project_folder(material: dict):
    """Genera/actualiza la carpeta del proyecto con PDF resumen y adjuntos."""
    try:
        code = (material.get("materiales") or "").strip()
        name = (material.get("cliente") or "Sin cliente").strip()
        # Limpiar saltos de línea y tabulaciones que romperían el nombre de carpeta
        code = re.sub(r'[\r\n\t]+', ' ', code).strip()
        name = re.sub(r'[\r\n\t]+', ' ', name).strip()
        pid = (material.get("id") or "")[:8]
        # Extraer año de la fecha del proyecto
        fecha = material.get("fecha") or material.get("created_at") or ""
        year = fecha[:4] if len(fecha) >= 4 and fecha[:4].isdigit() else "Sin año"
        folder_name = f"{code} - {name} ({pid})" if code else f"{name} ({pid})"
        # Sanitizar nombre de carpeta
        folder_name = re.sub(r'[<>:"/\\|?*\r\n\t]', '_', folder_name)[:200]
        proj_dir = ARCHIVOS_DIR / year / folder_name
        proj_dir.mkdir(parents=True, exist_ok=True)
        # Generar PDF resumen
        buf = io.BytesIO()
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas as rl_canvas
        c = rl_canvas.Canvas(buf, pagesize=A4)
        w, h = A4
        y = h - 20 * mm
        # Helper para evitar errores de encoding con caracteres no-Latin-1 y saltos de línea
        def _safe_text(s: str) -> str:
            return re.sub(r'[\r\n\t]+', ' ', s).encode("latin-1", errors="replace").decode("latin-1")
        c.setFont("Helvetica-Bold", 16)
        c.drawString(15 * mm, y, _safe_text(f"Proyecto: {code or 'Sin código'}"))
        y -= 10 * mm
        c.setFont("Helvetica", 11)
        campos = [
            ("Código", material.get("materiales")),
            ("Cliente", material.get("cliente")),
            ("Ubicación", material.get("ubicacion")),
            ("Horas previstas", material.get("horas_prev")),
            ("Comercial", material.get("comercial")),
            ("Gestor", material.get("gestor") or material.get("manager_name")),
            ("Fecha", material.get("fecha")),
            ("Entrega/Recogida", material.get("entrega_recogida")),
            ("Total/Parcial", material.get("total_parcial")),
            ("Técnico", material.get("tecnico")),
            ("Comentarios", material.get("comentarios")),
            ("Estado", material.get("project_status")),
        ]
        for label, value in campos:
            if value:
                c.drawString(15 * mm, y, _safe_text(f"{label}: {value}"))
                y -= 7 * mm
                if y < 25 * mm:
                    c.showPage()
                    y = h - 20 * mm
        c.save()
        buf.seek(0)
        pdf_path = proj_dir / _safe_text(f"{folder_name}.pdf")
        pdf_path.write_bytes(buf.read())

        # Copiar adjuntos del proyecto
        for att in (material.get("attachments") or []):
            try:
                att_data = att.get("base64") or att.get("data")
                if att_data:
                    raw = base64.b64decode(att_data)
                    att_path = proj_dir / (att.get("filename") or f"adjunto_{att.get('id','')}")
                    att_path.write_bytes(raw)
            except Exception:
                pass
    except Exception as e:
        logging.warning(f"sync_project_folder error for {material.get('id','')}: {e}")


@api_router.get("/archivos")
async def browse_archivos(
    request: Request,
    path: str = Query("", description="Ruta relativa dentro de Archivos ordenados"),
    token: Optional[str] = Query(None, description="Token JWT para descarga directa de archivos"),
):
    """Devuelve la lista de carpetas/archivos o descarga un archivo."""
    # Autenticación: por cabecera Bearer o por query param ?token=
    user = None
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        try:
            payload = pyjwt.decode(auth_header[7:], JWT_SECRET, algorithms=[JWT_ALGORITHM])
            user = await db.users.find_one({"id": payload["sub"]})
        except Exception:
            pass
    if not user and token:
        try:
            payload = pyjwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            user = await db.users.find_one({"id": payload["sub"]})
        except Exception:
            pass
    if not user:
        raise HTTPException(401, "Autenticación requerida")
    perms = await get_user_permissions(user)
    if "proyectos.view" not in perms:
        raise HTTPException(403, "No tienes permiso")
    base = ARCHIVOS_DIR
    # Normalizar el path: quitar saltos de línea que pueden venir por nombre de carpeta corrupto
    clean_path = re.sub(r'[\r\n\t]+', ' ', path or "")
    target = (base / clean_path).resolve()
    if not str(target).startswith(str(base.resolve())) or not target.exists():
        raise HTTPException(404, "Ruta no encontrada")
    if target.is_file():
        # Descargar archivo
        mt = "application/pdf" if target.suffix == ".pdf" else "application/octet-stream"
        return StreamingResponse(open(target, "rb"), media_type=mt, headers={"Content-Disposition": f'inline; filename="{target.name}"'})
    items = []
    for child in sorted(target.iterdir(), key=lambda x: (not x.is_dir(), x.name)):
        safe_name = re.sub(r'[\r\n\t]+', ' ', child.name)
        safe_path = re.sub(r'[\r\n\t]+', ' ', str(child.relative_to(base)))
        item = {
            "name": safe_name,
            "type": "folder" if child.is_dir() else "file",
            "path": safe_path,
            "size": child.stat().st_size if child.is_file() else None,
        }
        if child.is_dir():
            item["count"] = len(list(child.iterdir()))
        items.append(item)
    return {"path": path or "", "items": items}

class DocumentoUpload(BaseModel):
    titulo: str
    categoria: str  # "fichas" o "manuales"
    filename: str
    file_base64: str

@api_router.get("/documentos")
async def list_documentos(
    user: dict = Depends(current_user),
    categoria: Optional[str] = Query(None),
):
    q: dict = {}
    if categoria:
        q["categoria"] = categoria
    docs = await db.documentos.find(q, {"file_path": 0}).sort("created_at", -1).to_list(500)
    for d in docs:
        d.pop("file_path", None)  # no enviar ruta interna
    return docs

@api_router.post("/documentos")
async def create_documento(
    body: DocumentoUpload,
    user: dict = Depends(require_permission("documentos.manage")),
):
    if body.categoria not in ("fichas", "manuales"):
        raise HTTPException(400, "Categoría inválida")
    did = str(uuid.uuid4())
    # Guardar archivo en disco
    cat_dir = DOCUMENTOS_DIR / body.categoria
    cat_dir.mkdir(parents=True, exist_ok=True)
    file_path = cat_dir / f"{did}_{body.filename}"
    raw = base64.b64decode(body.file_base64)
    file_path.write_bytes(raw)
    doc = {
        "_id": did,
        "id": did,
        "titulo": body.titulo.strip(),
        "categoria": body.categoria,
        "filename": body.filename,
        "file_path": str(file_path),
        "created_by": user["email"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.documentos.insert_one(doc)
    doc.pop("file_path", None)
    return doc

@api_router.get("/documentos/{did}/file")
async def get_documento_file(
    did: str,
    user: dict = Depends(current_user),
):
    doc = await db.documentos.find_one({"_id": did}, {"_id": 0, "file_path": 1, "filename": 1})
    if not doc or not doc.get("file_path"):
        raise HTTPException(404, "Documento no encontrado")
    path = Path(doc["file_path"])
    if not path.exists():
        raise HTTPException(404, "Archivo no encontrado en disco")
    return StreamingResponse(
        open(path, "rb"),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{doc.get("filename", "documento.pdf")}"'},
    )

@api_router.get("/documentos/{did}")
async def get_documento(
    did: str,
    user: dict = Depends(current_user),
):
    doc = await db.documentos.find_one({"_id": did}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Documento no encontrado")
    doc.pop("file_path", None)
    return doc

@api_router.delete("/documentos/{did}")
async def delete_documento(
    did: str,
    user: dict = Depends(require_permission("documentos.manage")),
):
    doc = await db.documentos.find_one({"_id": did}, {"file_path": 1})
    if doc and doc.get("file_path"):
        try:
            Path(doc["file_path"]).unlink(missing_ok=True)
        except Exception:
            pass
    res = await db.documentos.delete_one({"_id": did})
    if res.deleted_count == 0:
        raise HTTPException(404, "Documento no encontrado")
    return {"ok": True}


app.include_router(api_router)
cors_origins = [FRONTEND_URL]
if CORS_ORIGINS:
    cors_origins.extend(o.strip() for o in CORS_ORIGINS.split(',') if o.strip())

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=cors_origins,
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def on_startup():
    try:
        await db.preciario_descuentos.create_index("ref", unique=True)
    except Exception:
        pass  # el índice ya existe o hay duplicados; la app sigue funcionando
    await ensure_default_roles_and_migrate()
    await seed_admin_user()
    await seed_initial_data()
    await backfill_user_colors()
    await seed_demo_data()

async def backfill_user_colors():
    """Assign a color to any user that doesn't have one."""
    users_without = await db.users.find(
        {"$or": [{"color": {"$exists": False}}, {"color": None}, {"color": ""}]},
        {"_id": 0, "id": 1}
    ).to_list(1000)
    if not users_without:
        return
    existing_colors = [u.get("color") for u in await db.users.find(
        {"color": {"$exists": True, "$ne": None, "$nin": [""]}},
        {"_id": 0, "color": 1}
    ).to_list(1000)]
    used = [c for c in existing_colors if c]
    for u in users_without:
        c = _next_default_color(used)
        await db.users.update_one({"id": u["id"]}, {"$set": {"color": c}})
        used.append(c)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
